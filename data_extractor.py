"""
DATAS PARA CAPTURACION
- Hojas WK  (Mi hoja Principal)
- Hojas PR  (pandas + requests)
- Hojas MP  (MANTENIMIENTO)
- Hojas ME  (MATERIAL DE EMPAQUE)
- Hojas MV  (MATERIAL VEGETAL)
"""

import streamlit as st

import re
import requests
import pandas as pd
import openpyxl
from copy import copy
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor, as_completed

#  URLs de SharePoint (leídas desde st.secrets["urls"]) 
_urls = st.secrets["urls"]
SHAREPOINT_URL_WK              = _urls["SHAREPOINT_URL_WK"]
SHAREPOINT_URL_PR              = _urls["SHAREPOINT_URL_PR"]
SHAREPOINT_URL_CONTEO          = _urls["SHAREPOINT_URL_CONTEO"]
SHAREPOINT_URL_CONTEO_MARLEN   = _urls["SHAREPOINT_URL_CONTEO_MARLEN"]
SHAREPOINT_URL_NOMINA          = _urls["SHAREPOINT_URL_NOMINA"]
SHAREPOINT_URL_SIEMBRA_DETALLE = _urls["SHAREPOINT_URL_SIEMBRA_DETALLE"]
SHAREPOINT_URL_WEEKLY          = _urls["SHAREPOINT_URL_WEEKLY"]
SHAREPOINT_URL_WEEKLY_2026     = _urls["SHAREPOINT_URL_WEEKLY_2026"]
GOOGLE_DRIVE_URL_TRANSPORTE    = _urls["GOOGLE_DRIVE_URL_TRANSPORTE"]

# Índices de columna (0-based) dentro de cada hoja WEEKLY####
# Usados como FALLBACK si no se detecta encabezado dinámico
_WK_COL_FLOR    = 5   # F   Variedad de flor
_WK_COL_INI     = 6   # G   INV. INICIAL
_WK_COL_CEC     = 7   # H   Recepción CECILIA
_WK_COL_RAM     = 8   # I   Recepción RAMONA
_WK_COL_ISA     = 9   # J   Recepción ISABELA
_WK_COL_CHR     = 10  # K   Recepción CRISTINA
_WK_COL_C25     = 11  # L   Recepción CECILIA 25
_WK_COL_COMP    = 14  # O   PRAS A TERC (DAMIAN) = Tallos Comprados
_WK_COL_EXPORT  = 18  # S   Exportación
_WK_COL_MUEST   = 19  # T   Muestras
_WK_COL_DES     = 21  # V   Desechos = Tallos Desechados
_WK_COL_INV_FIN = 22  # W   INV. FINAL Cálculo

# Nombres de columna a buscar dinámicamente en el encabezado de cada hoja WEEKLY####
# Cada entrada: clave interna  lista de posibles nombres en el Excel (sin importar mayúsculas/acentos)
_WK_COL_NAMES = {
    "flor":    ["flor", "variedad", "flower", "nombre"],
    "ini":     ["inv. inicial", "inventario inicial", "inv inicial", "inv.inicial"],
    "cec":     ["cecilia", "cec"],
    "ram":     ["ramona", "ram"],
    "isa":     ["isabela", "isa"],
    "chr":     ["cristina", "chr"],
    "c25":     ["cecilia 25", "cec 25", "c25"],
    "export":  ["exportacion", "exportación", "exportacion", "export"],
    "muest":   ["muestras", "muestra", "sample"],
    "des":     ["desechos", "desechados", "descarte"],
    "inv_fin": ["inv. final", "inventario final", "inv final", "inv.final", "calculo", "inv.final calculo"],
}

# Palabras clave para detectar la fila de encabezados de GRUPO
_WK_GROUP_KEYS = [
    "recepcion de flor", "recepcion flor",
    "compras a terceros", "compras terceros",
]

def _norm_cell(s: str) -> str:
    """Normaliza una celda: minúsculas, sin acentos, sin espacios/puntos extra."""
    s = s.lower().strip()
    for a, b in [("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u"),("ü","u"),("ñ","n")]:
        s = s.replace(a, b)
    s = re.sub(r'[\s.]+', ' ', s).strip()
    return s

def _celda_coincide(cell_norm: str, keywords: list) -> bool:
    """
    Devuelve True si la celda normalizada coincide EXACTAMENTE con alguna keyword.
    Evita falsos positivos como 'recepcion de flor'  'flor'.
    """
    for kw in keywords:
        kw_norm = _norm_cell(kw)
        if cell_norm == kw_norm:
            return True
        if cell_norm.startswith(kw_norm + " "):
            return True
    return False

def _detectar_columnas_weekly(df) -> dict:
    """
    Detecta dinámicamente los índices de columna en una hoja WEEKLY####.

    Estrategia de dos filas:
      · row_g  fila de encabezados de GRUPO ("RECEPCION DE FLOR", "COMPRAS A TERCEROS")
                Permite determinar el rango de columnas de cada grupo.
      · row_h  fila de encabezados INDIVIDUALES ("FLOR", "CECILIA", "EXPORTACION")
                Permite mapear columnas fijas por nombre exacto.

    Para COMPRAS A TERCEROS se retorna 'comp_cols': lista de todos los índices de columna
    bajo ese encabezado de grupo, sin importar cuántos proveedores haya.
    Si no se detecta el grupo, se usa el fallback de columna única.
    """
    fallbacks = {
        "flor": _WK_COL_FLOR, "ini": _WK_COL_INI,
        "cec":  _WK_COL_CEC,  "ram": _WK_COL_RAM,
        "isa":  _WK_COL_ISA,  "chr": _WK_COL_CHR,
        "c25":  _WK_COL_C25,
        "comp_cols":  [_WK_COL_COMP],   # lista de índices, no índice único
        "comp_names": ["COMPRAS"],       # nombre legible por cada col en comp_cols
        "export": _WK_COL_EXPORT, "muest": _WK_COL_MUEST,
        "des":  _WK_COL_DES,  "inv_fin": _WK_COL_INV_FIN,
    }

    #  1. Buscar fila de grupos (row_g) 
    row_g_idx  = None
    row_g_norm = []
    for row_idx in range(min(10, len(df))):
        fila_norm = [_norm_cell(str(v)) for v in df.iloc[row_idx]]
        if any(any(gk in cell for gk in _WK_GROUP_KEYS) for cell in fila_norm if cell):
            row_g_idx  = row_idx
            row_g_norm = fila_norm
            print(f"    Fila de grupos detectada en fila {row_g_idx}")
            break

    #  2. Buscar fila de encabezados individuales (row_h): la de más hits exactos 
    best_row_idx = None
    best_hits    = 0
    best_vals    = []
    for row_idx in range(min(15, len(df))):
        fila = [_norm_cell(str(v)) for v in df.iloc[row_idx]]
        hits = sum(
            1 for cell in fila if cell and
            any(_celda_coincide(cell, kws) for kws in _WK_COL_NAMES.values())
        )
        if hits > best_hits:
            best_hits    = hits
            best_row_idx = row_idx
            best_vals    = fila

    if best_hits < 4:
        print(f"   [WARN] Encabezado WEEKLY no detectado (hits={best_hits})  usando índices fijos")
        return fallbacks

    print(f"   [OK] Encabezado individual detectado en fila {best_row_idx} ({best_hits} columnas)")

    #  3. Mapear columnas fijas desde row_h 
    cols = dict(fallbacks)
    for clave, keywords in _WK_COL_NAMES.items():
        for col_idx, cell in enumerate(best_vals):
            if cell and _celda_coincide(cell, keywords):
                cols[clave] = col_idx
                break

    #  4. Detectar rango de COMPRAS A TERCEROS desde row_g 
    #
    # PROBLEMA CONOCIDO: EXPORTACION, MUESTRAS, DESECHOS e INV.FINAL no tienen
    # encabezado de grupo en row_g  quedan como celdas vacías. Eso hace que
    # comp_end llegue hasta el final de la fila capturando esas columnas también.
    #
    # SOLUCIÓN: después de acotar el rango por row_g, se filtran las columnas
    # cuyos nombres en row_h coincidan con cualquier campo "fijo" conocido
    # (export, muest, des, inv_fin, flor, ini, rancho, etc.).
    # Así el código es 100 % dinámico: acepta N proveedores nuevos en COMPRAS A
    # TERCEROS sin tocar nada, y nunca arrastra columnas que no corresponden.
    #
    # Palabras clave que identifican columnas que NO son proveedores de compra:
    _NON_COMP_KEYWORDS = [
        kw
        for key in ["flor", "ini", "cec", "ram", "isa", "chr", "c25",
                    "export", "muest", "des", "inv_fin"]
        for kw in _WK_COL_NAMES.get(key, [])
    ] + [
        # términos adicionales que pueden aparecer como encabezado individual
        "calculo", "fisico", "inv. final fisico", "inv. final calculo",
        "total", "desechados", "desechos", "descarte",
        "exportacion", "exportación", "muestras", "muestra",
        "inv. inicial", "inv inicial", "inventario inicial",
        "inv. final", "inv final", "inventario final",
    ]

    if row_g_idx is not None:
        comp_start = None
        for col_idx, cell in enumerate(row_g_norm):
            if cell and "compras" in cell:
                comp_start = col_idx
                break

        if comp_start is not None:
            # Límite superior: próxima celda no-vacía en row_g (otro grupo)
            comp_end = len(row_g_norm)
            for col_idx in range(comp_start + 1, len(row_g_norm)):
                if row_g_norm[col_idx]:
                    comp_end = col_idx
                    break

            # comp_cols = columnas en row_h dentro del rango con encabezado no vacío
            # EXCLUYENDO cualquier columna cuyo nombre coincida con un campo fijo conocido
            comp_cols = []
            excluidas = []
            for col_idx in range(comp_start, comp_end):
                if col_idx >= len(best_vals) or not best_vals[col_idx]:
                    continue
                cell_norm = best_vals[col_idx]   # ya está normalizado
                # ¿Coincide con algún campo que NO es un proveedor?
                es_no_comp = any(
                    _celda_coincide(cell_norm, [kw]) for kw in _NON_COMP_KEYWORDS
                )
                if es_no_comp:
                    excluidas.append((col_idx, cell_norm))
                else:
                    comp_cols.append(col_idx)

            if excluidas:
                print(f"    COMPRAS A TERCEROS  columnas excluidas (no son proveedores): "
                      f"{[(c, best_vals[c]) for c, _ in excluidas]}")

            if comp_cols:
                cols["comp_cols"]  = comp_cols
                cols["comp_names"] = [best_vals[c].upper() for c in comp_cols]
                print(f"    COMPRAS A TERCEROS: {len(comp_cols)} proveedore(s)  "
                      f"cols {comp_cols} ({cols['comp_names']})")
            else:
                print(f"   [WARN] COMPRAS A TERCEROS grupo encontrado pero sin columnas de proveedor  fallback")
        else:
            print(f"   [WARN] No se encontró 'COMPRAS A TERCEROS' en row_g  fallback col única")

    # Log columnas fijas
    fixed_keys = ["flor","ini","cec","ram","isa","chr","c25","export","muest","des","inv_fin"]
    for clave in fixed_keys:
        origen = "" if cols[clave] != fallbacks[clave] else "~"
        print(f"      {origen} {clave:10s}  col {cols[clave]}")

    return cols

#  Constantes 
RANCH_CONFIG = {
    "Prop-RM":     {"color": "#047857", "codes": ["VIV"], "keywords": ["PROP"]},
    "PosCo-RM":    {"color": "#1d4ed8", "codes": ["POS", "LIM"], "keywords": ["POSCO"]},
    "Campo-RM":    {"color": "#b45309", "codes": ["CAM", "RAM"], "keywords": ["CAMPO"]},
    "Isabela":     {"color": "#7c3aed", "codes": ["ISA"], "keywords": ["ISABEL"]},
    "HOOPS":       {"color": "#c2410c", "codes": ["HOO"], "keywords": ["HOOPS"]},
    "Cecilia":     {"color": "#be185d", "codes": ["CEC"], "keywords": ["CECILIA"]},
    "Cecilia 25":  {"color": "#047857", "codes": ["C25"], "keywords": ["CECILIA 25"]},
    "Christina":   {"color": "#0369a1", "codes": ["CHR"], "keywords": ["CHRISTINA"]},
    "Albahaca-RM": {"color": "#6d28d9", "codes": ["ALB"], "keywords": ["ALBAHACA"]},
    "Campo-VI":    {"color": "#64748b", "codes": [], "keywords": ["CAMPO-VI", "CAMPO-IV"]}
}

RANCH_KEYS = []
for data in RANCH_CONFIG.values():
    RANCH_KEYS.extend(data["keywords"])

RANCH_CODE_MAP = {}
for ranch, data in RANCH_CONFIG.items():
    for code in data["codes"]:
        RANCH_CODE_MAP[code] = ranch

CATEGORIAS_ORDEN = [
    "DESINFECCION Y FERTILIZACION",
    "AMPLIACION",
    "CULTIVO TIERRA, CHAROLAS",
    "MATERIAL VEGETAL",
    "PREPARACION DE SUELO",
    "FERTILIZANTES",
    "DESINFECCION / PLAGUICIDAS",
    "MANTENIMIENTO",
    "EXPANSION CECILIA 25",
    "RENOVACION DE SIEMBRA",
    "MATERIAL DE EMPAQUE",
    "COSTO SERVICIOS",
    "COSTO MANO DE OBRA",
]

WK_MXN_RANCH_COLS = {
    "Prop-RM": "E",
    "PosCo-RM": "F",
    "Campo-RM": "G",
    "Isabela": "H",
    "Christina": "I",
    "Cecilia": "J",
    "Cecilia 25": "K",
}

WK_MATERIAL_AUTOFILL = {
    "FERTILIZANTES": {"row": 15, "prefix": "PR", "tipo": "MIRFE"},
    "DESINFECCION / PLAGUICIDAS": {"row": 16, "prefix": "PR", "tipo": "MIPE"},
    "MANTENIMIENTO": {"row": 17, "prefix": "MP", "tipo": None},
    "MATERIAL DE EMPAQUE": {"row": 20, "prefix": "ME", "tipo": None},
    "MATERIAL VEGETAL": {"row": 13, "prefix": "MV", "tipo": None},
}

NOMINA_BD_FINCA_TO_WK_RANCH = {
    "VIVERO": "Prop-RM",
    "POSCOSECHA": "PosCo-RM",
    "RAMONA": "Campo-RM",
    "ISABELA": "Isabela",
    "CHRISTINA": "Christina",
    "CECILIA": "Cecilia",
    "CECILIA 25": "Cecilia 25",
    "CECILIA25": "Cecilia 25",
}

NOMINA_WK_ROWS = {
    24: "NOMINA ADMON Oficina, Jefes de Finca, Ingenieros",
    27: "NOMINA PRODUCCION",
    30: "NOMINA PRODUCCION CORTE",
    33: "NOMINA PRODUCCION TRANSPLANTE",
    36: "NOMINA PRODUCCION MANEJO PLANTA",
    39: "NOMINA HOOPS",
    42: "NOMINA (MIPE,MIRFE,)",
    45: "NOMINA OPERATIVOS (TRACTORES, CAMEROS)",
    48: "NOMINA OPERATIVOS (CHOFER)",
    51: "NOMINA OPERATIVOS (VELADORES)",
    54: "NOMINA OPERATIVOS (SOLDADOR)",
    57: "NOMINA PRODUCCION Contratista y comisiones",
}

SKIP = {"ACUMULADO", "GRAFICOS I-IV", "COMPARATIVO", "DATOS", "HOJA1", "SHEET1"}


#  Descarga de Excel desde SharePoint 
def _descargar_excel(url: str, label: str = "archivo") -> BytesIO | None:
    """
    Descarga un archivo .xlsx desde un link público de SharePoint/OneDrive.
    Agrega el parámetro download=1 necesario para la descarga directa.
    Funciona con URLs que tengan o no el token ?e=...
    """
    url = url.strip()
    if "?e=" in url:
        download_url = url.replace("?e=", "?download=1&e=")
    elif "?" in url:
        download_url = url + "&download=1"
    else:
        download_url = url + "?download=1"
    try:
        response = requests.get(download_url, timeout=30)
        response.raise_for_status()
        # SharePoint a veces devuelve HTML si la URL no es válida
        content_type = response.headers.get("Content-Type", "")
        if "html" in content_type.lower():
            print(f"[ERR] {label}: SharePoint devolvió HTML en vez de Excel. Verifica que el link sea público.")
            return None
        return BytesIO(response.content)
    except Exception as e:
        print(f"[ERR] Error descargando {label}: {e}")
        return None


# Alias para compatibilidad con get_sheet_xlsx
def descargar_excel() -> BytesIO | None:
    return _descargar_con_graph(SHAREPOINT_URL_WK, "Excel WK")


def _leer_hoja(xls: pd.ExcelFile, titulo: str, rango_filas: int = 60,
               rango_cols: int = 35) -> list[list]:
    """
    Lee una hoja del ExcelFile y la retorna como lista de listas.
    Las celdas vacías / NaN se convierten a "".
    """
    try:
        df = pd.read_excel(
            xls,
            sheet_name=titulo,
            header=None,
            nrows=rango_filas,
        ).fillna("")
        if df.shape[1] > rango_cols:
            df = df.iloc[:, :rango_cols]
        return df.values.tolist()
    except Exception as e:
        print(f"   [WARN]  Error leyendo hoja '{titulo}': {e}")
        return []


#  Helpers de normalización 
def norm_ranch(s: str):
    s = str(s).upper().strip()
    if "CAMPO-VI" in s or "CAMPO-IV" in s:               return "Campo-VI"
    if "CECILIA 25" in s or "25 CECILIA" in s:           return "Cecilia 25"
    if "CECILIA" in s and "25" not in s:                 return "Cecilia"
    if "CAMPO" in s and "VI" not in s and "IV" not in s: return "Campo-RM"
    if "CRISTINA" in s:                                  return "Christina"
    if "PROPAGACION" in s:                               return "Prop-RM"
    
    for ranch, data in RANCH_CONFIG.items():
        if ranch in ["Campo-VI", "Cecilia 25", "Cecilia", "Campo-RM"]: 
            continue
        for kw in data["keywords"]:
            if kw in s:
                return ranch
    return None


def norm_cat(s: str):
    s = str(s).upper().strip()
    if "DESINFECCION" in s and "FERTILIZ" in s:  return "DESINFECCION Y FERTILIZACION"
    if s.startswith("AMPLIACION"):                return "AMPLIACION"
    if "CULTIVO" in s:                            return "CULTIVO TIERRA, CHAROLAS"
    if "MATERIAL VEG" in s:                       return "MATERIAL VEGETAL"
    if "PREPARACION" in s:                        return "PREPARACION DE SUELO"
    if "FERTILIZANTE" in s:                       return "FERTILIZANTES"
    if "SANIDAD" in s or "PLAGUICIDA" in s:       return "DESINFECCION / PLAGUICIDAS"
    if "MANTENIMIENTO" in s:                      return "MANTENIMIENTO"
    if "EXPANSION" in s:                          return "EXPANSION CECILIA 25"
    if "RENOVACION" in s:                         return "RENOVACION DE SIEMBRA"
    if "MATERIAL DE EMP" in s:                    return "MATERIAL DE EMPAQUE"
    if "COSTO DE MAT" in s:                       return "COSTO_STOP"
    if "COSTO DE MANO DE OBRA" in s:              return "COSTO MANO DE OBRA"
    if "COSTO DE SERV" in s:                      return "COSTO SERVICIOS"
    if s.startswith("ELECTRICIDAD"):                        return "SV:Electricidad"
    if s.startswith("FLETES Y ACARREOS"):                   return "SV:Fletes y Acarreos"
    if s.startswith("GASTOS DE EXPORTACION"):               return "SV:Gastos de Exportación"
    if s.startswith("CERTIFICADO DE FITOSANITARIO"):        return "SV:Certificado Fitosanitario"
    if s.startswith("TRANSPORTE DE PERSONAL"):              return "SV:Transporte de Personal"
    if s.startswith("COMPRA DE FLOR"):                      return "SV:Compra de Flor a Terceros"
    if s.startswith("COMIDA PARA EL PERSONAL"):             return "SV:Comida para el Personal"
    if s.startswith("RO, TEL") or s.startswith("RO , TEL"): return "SV:RO, TEL, RTA.Alim"
    #  MANO DE OBRA 
    # Nómina + H.Extra + Bonos se colapsan al mismo subcat que BD CONTEO
    # para que semanas 1-13 (conteo.xlsx) y 14+ (WK Excel) sean consistentes.
    _is_nomina = "NOMINA" in s or "NÓMINA" in s
    _is_hextra = "HORAS EXTR" in s
    _is_bonos  = "BONOS ASISIT" in s

    if _is_nomina or _is_hextra or _is_bonos:
        # Admon Posco antes de Admon genérico (contiene "ADMON")
        if "ADMON" in s and ("POSCO" in s or "POSCOSECHA" in s):
            return "MO:Admon Posco"
        # Administración: Nómina/H.Extra dom.festivos/Bonos despensa
        if "ADMON" in s or ("FESTIVOS" in s and "FEST." not in s) or "DESPENSA" in s:
            return "MO:Ing. Y Admon."
        if "SUPERVISOR" in s:                              return "MO:Supervisores"
        if "CORTE" in s:                                   return "MO:Corte"
        if "TRANSPLANTE" in s:                             return "MO:Trasplante"
        if "MANEJO" in s and "PLANTA" in s:                return "MO:Manejo P."
        if "CONSOLIDAC" in s:                              return "MO:Consolidacion"
        if "SIEMBRA" in s:                                 return "MO:Siembra"
        if "MOV" in s and "CHAROLA" in s:                  return "MO:Mov. Charolas"
        if "RIEGO" in s:                                   return "MO:Riego"
        if "PHLOX" in s or "ESQUEJE" in s:                 return "MO:Esquejes"
        if "HOOPS" in s:                                   return "MO:Hoops"
        if "MIPE" in s or "MIRFE" in s:                    return "MO:MIPE Y MIRFE"
        if "TRACTORES" in s or "CAMEROS" in s:             return "MO:Tract. Y Cameros"
        if "VELADOR" in s:                                 return "MO:Veladores"
        if "SOLDADOR" in s:                                return "MO:Soldadores"
        if "CHOFER" in s:                                  return "MO:Transporte"
        if "CONTRATISTA" in s:                             return "MO:Contratista y com."
        if "ALM" in s and ("UPC" in s or "EMPAQ" in s):   return "MO:Alm.upc y empaq"
        # Producción general, Dom y Fest. Prod., Nómina Producción  Prod. Patina y rec
        return "MO:Prod. Patina y rec"
    if "IMSS" in s or "INFONAVIT" in s:                    return "MO:IMSS,INFO Y RCV"
    if "1.8%" in s or "TASA EFECTIVA" in s:                return "MO:Imp. 1.8%"
    return None


#  Mapa de nombres de rancho WK  BD (para consistencia en mano_obra_data) 
WK_RANCH_TO_BD = {
    "Prop-RM":  "Propagacion",
    "PosCo-RM": "Poscosecha",
    "Campo-RM": "Ramona",
}


def _area_from_concepto_rancho(label: str, ranch: str) -> str | None:
    """
    Dado el label (Concepto) de una fila WK y el nombre del rancho (norm_ranch),
    retorna el Área equivalente al BD para que semanas 1-13 y 14+ sean consistentes.
    Retorna None si el concepto no aplica para ese rancho (se debe omitir).

    Tabla de mapeo (fuente: columna Área del Conteo Personal BD):
      Concepto WK              | Prop-RM       | PosCo-RM          | Otros
      -------------------------|---------------|-------------------|-------------------
      NOMINA PRODUCCION        | Supervisores  | Admon Posco       | (vacío/omitir)
      PRODUCCION CORTE         | Phlox         | Supervisores      | Corte
      PRODUCCION TRANSPLANTE   | Siembra       | Prod. Patina y rec| Trasplante
      PRODUCCION MANEJO PLANTA | Consolidacion | Alm.upc y empaq   | Manejo P.
      HOOPS                    | Mov. Charolas | (omitir)          | Hoops
      (MIPE,MIRFE)             | Riego         | (omitir)          | MIPE Y MIRFE
      TRACTORES/CAMEROS        | (omitir)      | (omitir)          | Tract. Y Cameros
      CHOFER                   | (omitir)      | Transporte        | Transporte
      VELADORES                | (omitir)      | (omitir)          | Veladores
      SOLDADOR                 | (omitir)      | (omitir)          | Soldadores
    """
    s = str(label).upper().strip()
    r = str(ranch).upper().strip() if ranch else ""

    _is_nomina = "NOMINA" in s or "NÓMINA" in s
    _is_hextra = "HORAS EXTR" in s
    _is_bonos  = "BONOS ASISIT" in s
    _is_prop   = "PROP" in r      # Prop-RM   Propagacion
    _is_posco  = "POSCO" in r     # PosCo-RM  Poscosecha

    if not (_is_nomina or _is_hextra or _is_bonos):
        if "IMSS" in s or "INFONAVIT" in s:        return "IMSS,INFO Y RCV"
        if "1.8%" in s or "TASA EFECTIVA" in s:    return "Imp. 1.8%"
        return None

    #  1. Actividades específicas PRIMERO (antes de chequeos genéricos) 

    # CORTE: Prop=Esquejes, PosCo=Supervisores, otros=Corte
    if "CORTE" in s:
        if _is_prop:  return "Esquejes"
        if _is_posco: return "Supervisores"
        return "Corte"

    # TRANSPLANTE: Prop=Siembra, PosCo=Prod. Patina y rec, otros=Trasplante
    if "TRANSPLANTE" in s or "TRASPLANTE" in s:
        if _is_prop:  return "Siembra"
        if _is_posco: return "Prod. Patina y rec"
        return "Trasplante"

    # MANEJO PLANTA: Prop=Consolidacion, PosCo=Alm.upc y empaq, otros=Manejo P.
    if "MANEJO" in s and "PLANTA" in s:
        if _is_prop:  return "Consolidacion"
        if _is_posco: return "Alm.upc y empaq"
        return "Manejo P."

    # HOOPS: Prop=Mov. Charolas, PosCo=omitir, otros=Hoops
    if "HOOPS" in s:
        if _is_prop:  return "Mov. Charolas"
        if _is_posco: return None
        return "Hoops"

    # MIPE/MIRFE: Prop=Riego, PosCo=omitir, otros=MIPE Y MIRFE
    if "MIPE" in s or "MIRFE" in s:
        if _is_prop:  return "Riego"
        if _is_posco: return None
        return "MIPE Y MIRFE"

    # TRACTORES/CAMEROS: Prop=omitir, PosCo=omitir, otros=Tract. Y Cameros
    if "TRACTORES" in s or "CAMEROS" in s:
        if _is_prop or _is_posco: return None
        return "Tract. Y Cameros"

    # CHOFER: Prop=omitir, PosCo=Transporte, otros=Transporte
    if "CHOFER" in s:
        if _is_prop: return None
        return "Transporte"

    # VELADORES: Prop=omitir, PosCo=Veladores, otros=Veladores
    if "VELADOR" in s:
        if _is_prop: return None
        return "Veladores"

    # SOLDADOR: Prop=omitir, PosCo=Soldadores, otros=Soldadores
    if "SOLDADOR" in s:
        if _is_prop: return None
        return "Soldadores"

    #  2. Conceptos varios 
    if "CONTRATISTA" in s:                           return "Contratista y com."
    if "ALM" in s and ("UPC" in s or "EMPAQ" in s): return "Alm.upc y empaq"
    if "CONSOLIDAC" in s:                            return "Consolidacion"
    if "SIEMBRA" in s:                               return "Siembra"
    if "MOV" in s and "CHAROLA" in s:                return "Mov. Charolas"
    if "RIEGO" in s:                                 return "Riego"
    if "PHLOX" in s or "ESQUEJE" in s:               return "Esquejes"
    if "SUPERVISOR" in s:                            return "Supervisores"

    #  3. Administración (DESPUÉS de actividades específicas) 
    if "ADMON" in s and ("POSCO" in s or "POSCOSECHA" in s):
        return "Admon Posco"
    if "ADMON" in s or "DESPENSA" in s:
        return "Ing. Y Admon."

    # HORAS EXTR. DOM. Y FESTIVOS (palabra completa, sin qualifier de actividad)
    #  horas extra de Administración
    if _is_hextra and "FESTIVOS" in s:
        return "Ing. Y Admon."

    #  4. NOMINA PRODUCCION genérica / HORAS EXTR. DOM. Y FEST. / BONOS 
    # Sin actividad específica  mapeo según rancho
    # Prop-RM=Supervisores, PosCo-RM=Admon Posco, otros=omitir (vacío en BD)
    if _is_posco: return "Admon Posco"
    if _is_prop:  return "Supervisores"
    return None   # otros ranches: vacío  omitir


def sv(v) -> float:
    try:
        if isinstance(v, str):
            v = v.replace("$", "").replace(",", "").strip()
        f = float(v)
        return f if f == f else 0.0
    except (TypeError, ValueError):
        return 0.0


#  Detección de rancho por palabras clave (exclusivo para MV) 
def _ranch_from_ubicacion_mv(ubicacion: str):
    """
    Para hojas MV (Material Vegetal), detecta el rancho usando palabras clave
    en el campo UBICACION, sin importar mayúsculas/minúsculas.
    Cubre errores tipográficos comunes (ej: 'Prop', 'propaga', 'Cristina').
      - Empieza con 'PROPAGA'    Prop-RM
      - Empieza con 'CRISTINA'   Christina
      - Empieza con 'CECILIA25'  Cecilia 25  (con o sin espacio, cualquier capitalización)
      - Empieza con 'CECILIA'    Cecilia
      - Empieza con 'RAMONA'     Campo-RM
    Si no coincide ninguno, retorna None y se intenta el mapa de 3 letras normal.
    """
    u = ubicacion.upper().replace(' ', '')  # elimina espacios para cubrir "Cecilia 25" y "Cecilia25"
    if u.startswith('PROPAGA'):   return 'Prop-RM'
    if u.startswith('CRISTINA'):  return 'Christina'
    if u.startswith('CECILIA25'): return 'Cecilia 25'  # debe ir ANTES que 'CECILIA'
    if u.startswith('CECILIA'):   return 'Cecilia'
    if u.startswith('RAMONA'):    return 'Campo-RM'
    return None


#  Parser genérico compartido (PR / MP / ME tienen el mismo formato) 
def _parse_generic(rows: list, mv_mode: bool = False) -> dict:
    """
    Formato común a PR####, MP####, ME####:
      Col 2: UBICACION  (ej: RAMMIPRNN, CECMIPSNF)
      Col 5: PRODUCTO
      Col 7: UNIDADES
      Col 9: GASTO
    mv_mode=True: detección de rancho por palabras clave para MV (Material Vegetal).
    Retorna: { rancho: { tipo: [[producto, unidades, gasto, ubicacion], ...] } }
    """
    UBICACION_COL = 2
    PRODUCTO_COL  = 5
    UNIDADES_COL  = 7
    GASTO_COL     = 9

    # Autodetectar columnas si existe fila de encabezado (soporte para formato limpio)
    for i in range(min(15, len(rows))):
        if not rows[i]: continue
        r_str = [str(c).strip().upper() for c in rows[i]]
        if "UBICACION" in r_str and ("GASTO" in r_str or "COSTO" in r_str):
            UBICACION_COL = r_str.index("UBICACION")
            if "GASTO" in r_str:
                GASTO_COL = r_str.index("GASTO")
            elif "COSTO" in r_str:
                GASTO_COL = r_str.index("COSTO")
                
            if "PRODUCTO" in r_str: 
                PRODUCTO_COL = r_str.index("PRODUCTO")
            if "UNIDADES" in r_str: 
                UNIDADES_COL = r_str.index("UNIDADES")
            elif "CANTIDAD" in r_str:
                UNIDADES_COL = r_str.index("CANTIDAD")
            break

    result = {}
    accum  = {}   # (rancho, tipo, producto, ubicacion)  [u_total, g_total]

    for row in rows:
        if not row or len(row) <= max(UBICACION_COL, PRODUCTO_COL, UNIDADES_COL, GASTO_COL):
            continue

        ubicacion = str(row[UBICACION_COL]).strip().upper() if len(row) > UBICACION_COL else ''
        ubicacion = re.sub(r'\s+', '', ubicacion)

        if not ubicacion or len(ubicacion) < 6:
            continue
        if not re.match(r'^[A-Z0-9]+$', ubicacion):
            continue

        # MV: primero intentar detección por palabras clave (cubre errores tipográficos)
        if mv_mode:
            rancho = _ranch_from_ubicacion_mv(ubicacion)
        else:
            rancho = None

        # Fallback al mapa de 3 letras (aplica siempre si no se resolvió arriba)
        if not rancho:
            ranch_code = ubicacion[:3]
            rancho = RANCH_CODE_MAP.get(ranch_code)

        if not rancho and ubicacion.startswith('VIV'):
            rancho = 'Prop-RM'

        if not rancho:
            continue

        tipo = 'MIPE' if 'MIP' in ubicacion else 'MIRFE'

        producto = str(row[PRODUCTO_COL]).strip() if len(row) > PRODUCTO_COL else ''
        if not producto or producto.upper() in ('PRODUCTO', 'NOMBRE', ''):
            continue

        unidades = str(row[UNIDADES_COL]).strip() if len(row) > UNIDADES_COL else ''
        try:
            u = float(str(unidades).replace(',', ''))
            unidades = str(int(u)) if u == int(u) else str(round(u, 2))
        except Exception:
            unidades = '0'

        gasto = str(row[GASTO_COL]).strip() if len(row) > GASTO_COL else ''
        try:
            g = float(str(gasto).replace(',', ''))
            gasto = str(round(g, 2))
        except Exception:
            gasto = '0'

        u_f = float(unidades) if unidades else 0.0
        g_f = float(gasto)    if gasto    else 0.0

        key = (rancho, tipo, producto, ubicacion)
        if key in accum:
            accum[key][0] += u_f
            accum[key][1] += g_f
        else:
            accum[key] = [u_f, g_f]

    for (rancho, tipo, producto, ubicacion), (u_tot, g_tot) in accum.items():
        u_str = str(int(u_tot)) if u_tot == int(u_tot) else str(round(u_tot, 2))
        g_str = str(round(g_tot, 2))
        result.setdefault(rancho, {}).setdefault(tipo, []).append([producto, u_str, g_str, ubicacion])

    return result


#  Fetch hojas PR / MP / ME desde el segundo Excel de SharePoint 
def _fetch_desde_sharepoint(prefix: str, parser_fn, label: str) -> tuple[dict, dict]:
    """
    Descarga el Excel secundario de SharePoint y extrae todas las hojas
    que coincidan con el patrón  {PREFIX}####  (ej: PR2611, MP2608, ME2610).

    Args:
        prefix:    "PR", "MP" o "ME"
        parser_fn: función que convierte list[list]  dict de ranchos
        label:     nombre legible para logs

    Returns:
        (datos, debug)  con el mismo formato que antes usaban las funciones gspread
    """
    datos = {}
    debug = {f"hojas_{prefix.lower()}_encontradas": []}

    archivo = _descargar_con_graph(SHAREPOINT_URL_PR, f"Excel {label}")
    if archivo is None:
        print(f"[WARN]  No se pudo descargar el archivo para hojas {prefix}")
        return datos, debug

    try:
        xls = pd.ExcelFile(archivo)
    except Exception as e:
        print(f"[WARN]  No se pudo abrir el Excel de {label}: {e}")
        return datos, debug

    hojas_encontradas = []
    pat = re.compile(rf'^{prefix}\s*\d{{4}}$', re.IGNORECASE)

    for sname in xls.sheet_names:
        sname = sname.strip()
        if pat.match(sname):
            raw_code = re.sub(rf'{prefix}\s*', '', sname, flags=re.IGNORECASE).strip()
            try:
                code = int(raw_code)
                year = 2000 + (code // 100)
                if 2018 <= year <= 2030:
                    print(f"   [OK] {prefix}{code} encontrada en SharePoint: {sname}")
                    hojas_encontradas.append((sname, code))
                else:
                    print(f"   [ERR] {prefix}{code} año {year} fuera de rango")
            except ValueError as e:
                print(f"   [ERR] Error código '{raw_code}': {e}")

    debug[f"hojas_{prefix.lower()}_encontradas"] = [t for t, _ in hojas_encontradas]

    if not hojas_encontradas:
        print(f"   [i]  No hay hojas {prefix} en el Excel de SharePoint")
        return datos, debug

    for titulo, code in hojas_encontradas:
        vals   = _leer_hoja(xls, titulo, rango_filas=500, rango_cols=11)
        parsed = parser_fn(vals)
        datos[code] = parsed
        debug[f"{prefix}{code}_ranchos"] = list(parsed.keys()) if parsed else []
        print(f"    {prefix}{code} ranchos detectados: {list(parsed.keys())}")

    return datos, debug


def _normalizar_week_code(week_code: str) -> str:
    code = str(week_code or "").strip().upper()
    if code.startswith("WK"):
        code = code[2:]
    return code


def _buscar_hoja_por_prefijo(sheet_names: list[str], prefix: str, week_code: str) -> str | None:
    code = _normalizar_week_code(week_code)
    pat = re.compile(rf'^{prefix}\s*{re.escape(code)}$', re.IGNORECASE)
    for sname in sheet_names:
        if pat.match(str(sname).strip()):
            return str(sname).strip()
    return None


def _nomina_wk_row_from_departamento(departamento: str) -> int | None:
    dep = str(departamento).strip().upper()
    if not dep:
        return None
    if dep in {"ADMON", "ADMON CAMPO"}:
        return 24
    if dep in {"VIVERO", "POSCOSECHA"}:
        return 27
    if dep.startswith("CORTE "):
        return 30
    if dep.startswith("TRANSPLANTE "):
        return 33
    if dep.startswith("MANEJO "):
        return 36
    if dep.startswith("HOOPS "):
        return 39
    if dep.startswith("MIPE ") or dep.startswith("MIRFE "):
        return 42
    if dep.startswith("CAMERO ") or dep.startswith("TRACTORISTA "):
        return 45
    if dep.startswith("CHOFER "):
        return 48
    if dep.startswith("VELADOR "):
        return 51
    if dep.startswith("SOLDADOR "):
        return 54
    if "COMISIONES" in dep:
        return 57
    return None


def _nomina_wk_ranch_from_bd(finca: str, departamento: str) -> str | None:
    finca_norm = str(finca).strip().upper()
    if finca_norm in NOMINA_BD_FINCA_TO_WK_RANCH:
        return NOMINA_BD_FINCA_TO_WK_RANCH[finca_norm]

    dep = str(departamento).strip().upper().replace("  ", " ")
    if "CECILIA 25" in dep or "CECILIA25" in dep:
        return "Cecilia 25"
    if dep.endswith("CECILIA"):
        return "Cecilia"
    if dep.endswith("CHRISTINA"):
        return "Christina"
    if dep.endswith("ISABELA"):
        return "Isabela"
    if dep.endswith("RAMONA"):
        return "Campo-RM"
    if dep.endswith("POSCOSECHA") or dep == "POSCOSECHA":
        return "PosCo-RM"
    if dep.endswith("VIVERO") or dep == "VIVERO":
        return "Prop-RM"

    # Casos administrativos: se alinean con el mismo bloque E:K del autorrelleno WK.
    if dep == "ADMON":
        return "Prop-RM"
    if dep == "ADMON CAMPO":
        return "Campo-RM"

    return None


def _sumar_gasto_por_rancho(parsed: dict, tipo: str | None = None) -> tuple[dict, dict]:
    totales = {rancho: 0.0 for rancho in WK_MXN_RANCH_COLS}
    omitidos = {}
    for rancho, by_tipo in (parsed or {}).items():
        subtotal = 0.0
        tipos = [tipo] if tipo else list((by_tipo or {}).keys())
        for tipo_name in tipos:
            for item in (by_tipo or {}).get(tipo_name, []):
                try:
                    subtotal += abs(float(str(item[2]).replace(",", "")))
                except Exception:
                    continue
        subtotal = round(subtotal, 2)
        if rancho in totales:
            totales[rancho] = round(totales[rancho] + subtotal, 2)
        elif subtotal:
            omitidos[rancho] = round(omitidos.get(rancho, 0.0) + subtotal, 2)
    return totales, omitidos


#  Extractor principal 
def extraer_datos(xls: pd.ExcelFile) -> dict:
    all_data       = []
    servicios_data = []
    mano_obra_data = []

    hojas_validas = []
    pr_hojas      = []

    print("\n" + "=" * 60)
    print("[FIND] DETECTANDO HOJAS EN EL EXCEL WK")
    print("=" * 60)

    for sname in xls.sheet_names:
        sname = sname.strip()
        print(f"\n[DOC] Hoja: '{sname}'")

        if sname.upper() in SKIP:
            print("     SKIP (en lista de exclusión)")
            continue

        pr_match = re.match(r'^PR\s*\d{4}$', sname, re.IGNORECASE)
        if pr_match:
            pr_raw = re.sub(r'PR\s*', '', sname, flags=re.IGNORECASE).strip()
            try:
                pr_code = int(pr_raw)
                pr_year = 2000 + (pr_code // 100)
                if 2018 <= pr_year <= 2030:
                    print("   [OK] PR DETECTADA Y VÁLIDA (en WK Excel)")
                    pr_hojas.append((sname, pr_code))
                    continue
            except ValueError:
                pass

        wk_match = re.match(r'^WK\s*\d{4}$', sname, re.IGNORECASE)
        if wk_match:
            code_raw = re.sub(r"WK\s*", "", sname, flags=re.IGNORECASE).strip()
            try:
                code = int(code_raw)
                year = 2000 + (code // 100)
                if 2018 <= year <= 2030:
                    print("   [OK] WK DETECTADA Y VÁLIDA")
                    hojas_validas.append((sname, code))
                else:
                    print(f"   [ERR] Año {year} fuera de rango")
            except ValueError:
                print("   [ERR] Error convirtiendo código")
        else:
            if not pr_match:
                print("   [i]  No es WK ni PR")

    print("\n" + "=" * 60)
    print("[CHART] RESUMEN:")
    print(f"    Hojas WK encontradas: {len(hojas_validas)}")
    print(f"    Hojas PR en WK Excel: {len(pr_hojas)}")
    print("=" * 60 + "\n")

    if not hojas_validas:
        return {"error": "No se encontraron hojas WK validas."}

    # 2. Leer hojas WK
    batch_data = {}
    for titulo, _ in hojas_validas:
        batch_data[titulo] = _leer_hoja(xls, titulo, rango_filas=125, rango_cols=35)

    # 2b. Leer hojas PR que estén en el Excel WK (fallback)
    productos       = {}
    productos_debug = {"hojas_pr_encontradas": [t for t, _ in pr_hojas]}
    for titulo, pr_code in pr_hojas:
        vals   = _leer_hoja(xls, titulo, rango_filas=500, rango_cols=11)
        parsed = _parse_generic(vals)
        productos[pr_code] = parsed
        productos_debug[f"PR{pr_code}_ranchos"] = list(parsed.keys()) if parsed else []

    # 3. Procesar cada hoja WK
    siembra_data: dict = {}  # {wk_code: {ranch: {charolas,esquejes,metros,hectareas}}}
    unit_costs_data: dict = {}
    SIEMBRA_LABELS = [
        ("inv_inicial",     "INVENTARIO INICIAL"),
        ("tallos_cos",      "TALLOS COSECHADOS"),
        ("tallos_des",      "TALLOS DESECHADOS"),
        ("tallos_des_sf",   "TALLOS DESECHADOS SF"),
        ("tallos_comp",     "TALLOS COMPRADOS"),
        ("tallos_bouq",     "TALLOS EN BOUQUETS O PROCESADOS"),
        ("tallos_desp",     "TALLOS DESPACHADOS"),
        ("libras_alb",      "LIBRAS DESPACHADAS ALBAHACA"),
        ("tallos_mues",     "TALLOS MUESTRA"),
        ("inv_final",       "INVENTARIO FINAL"),
        ("tallos_proc",     "TALLOS PROCESADOS TOTALES"),
        ("charolas_288",    "CHAROLAS SEMBRADAS"),
        ("charolas",        "NUMERO DE CHAROLAS SEMBRADAS"),
        ("esquejes",        "NUMERO DE ESQUEJES SEMBRADOS"),
        ("metros",          "METROS DE SIEMBRA"),
        ("hectareas",       "HECTAREAS EN SIEMBRA"),
    ]

    for titulo, code in hojas_validas:
        raw = batch_data.get(titulo, [])
        if not raw:
            continue

        yy   = code // 100
        ww   = code % 100
        year = 2000 + yy

        max_cols = max((len(r) for r in raw), default=0)
        data     = [r + [""] * (max_cols - len(r)) for r in raw]

        date_range = ""
        for _dr in range(min(8, len(data))):
            for _dc in range(min(5, len(data[_dr]))):
                _v = str(data[_dr][_dc]).strip()
                if _v and " al " in _v.lower() and len(_v) > 8:
                    date_range = _v
                    break
            if date_range:
                break

        exec_idx = -1
        for i, row in enumerate(data):
            if any(isinstance(c, str) and "EJECUCION SEMANAL" in c.upper() for c in row):
                exec_idx = i
                break
        if exec_idx < 0:
            continue

        header_idx = -1
        for i in range(exec_idx - 1, max(0, exec_idx - 6) - 1, -1):
            if any(isinstance(v, str) and any(k in v.upper() for k in RANCH_KEYS) for v in data[i]):
                header_idx = i
                break
        if header_idx < 0:
            continue

        header = data[header_idx]

        total_cols = [j for j, v in enumerate(header)
                      if isinstance(v, str) and v.strip().upper() == "TOTAL"]
        if not total_cols:
            continue
        mxn_total_col = total_cols[0]
        usd_total_col = total_cols[1] if len(total_cols) >= 2 else None

        mxn_ranch_cols, usd_ranch_cols = {}, {}
        for j, v in enumerate(header):
            rn = norm_ranch(str(v)) if v else None
            if not rn:
                continue
            if j < mxn_total_col:
                mxn_ranch_cols[j] = rn
            elif usd_total_col and mxn_total_col < j < usd_total_col:
                mxn_ranch_cols[j] = rn
            elif usd_total_col and j > usd_total_col:
                usd_ranch_cols[j] = rn

        print(f"\n[DEBUG {titulo}]")
        print(f"   exec_idx={exec_idx}, header_idx={header_idx}")
        print(f"   mxn_total_col={mxn_total_col}, usd_total_col={usd_total_col}")
        print(f"   mxn_ranch_cols={mxn_ranch_cols}")
        print(f"   usd_ranch_cols={usd_ranch_cols}")
        hdr_vals = [(j, str(header[j])[:15]) for j in range(len(header)) if str(header[j]).strip()]
        print(f"   header non-empty: {hdr_vals}")

        for i in range(exec_idx + 1, min(exec_idx + 120, len(data))):
            row   = data[i]
            label = next((str(row[c]).strip() for c in range(5)
                          if c < len(row) and row[c] and len(str(row[c]).strip()) > 3), None)
            if not label:
                continue

            cat = norm_cat(label)
            if not cat:
                continue
            if cat == "COSTO_STOP":
                continue

            mxn_ranches = {rn: sv(row[j]) for j, rn in mxn_ranch_cols.items() if j < len(row)}
            usd_ranches = {rn: sv(row[j]) for j, rn in usd_ranch_cols.items() if j < len(row)}

            if cat.startswith("SV:"):
                print(f"   [SV] fila={i} label='{label[:30]}' cat='{cat}' mxn_ranches={mxn_ranches}")
                servicios_data.append({
                    "semana":      code,
                    "year":        year,
                    "week":        ww,
                    "date_range":  date_range,
                    "subcat":      cat[3:],
                    "mxn_total":   round(sv(row[mxn_total_col]) if mxn_total_col < len(row) else 0, 2),
                    "usd_total":   round(sv(row[usd_total_col]) if usd_total_col and usd_total_col < len(row) else 0, 2),
                    "mxn_ranches": mxn_ranches,
                    "usd_ranches": usd_ranches,
                })
            elif cat.startswith("MO:"):
                # Agrupar ranchos por su Área correcta (varía según rancho)
                # y normalizar nombre de rancho de WK a nombre BD
                area_groups: dict = {}
                for rn, mxn_val in mxn_ranches.items():
                    if mxn_val == 0.0:
                        continue   # omitir ranchos sin costo
                    area = _area_from_concepto_rancho(label, rn)
                    if area is None:
                        continue   # este concepto no aplica para este rancho (vacío en BD)
                    rn_bd = WK_RANCH_TO_BD.get(rn, rn)   # normalizar nombre
                    if area == "Ing. Y Admon." and rn_bd == "Poscosecha":
                        rn_bd = "Administracion"
                        
                    ag = area_groups.setdefault(area, {
                        "mxn_ranches": {}, "usd_ranches": {},
                        "mxn_t": 0.0, "usd_t": 0.0,
                    })
                    ag["mxn_ranches"][rn_bd] = round(ag["mxn_ranches"].get(rn_bd, 0.0) + mxn_val, 2)
                    ag["mxn_t"] = round(ag["mxn_t"] + mxn_val, 2)
                    usd_val = usd_ranches.get(rn, 0.0)
                    if usd_val:
                        ag["usd_ranches"][rn_bd] = round(ag["usd_ranches"].get(rn_bd, 0.0) + usd_val, 2)
                        ag["usd_t"] = round(ag["usd_t"] + usd_val, 2)

                # Sin ranchos: un registro con el total de la fila
                if not area_groups:
                    mxn_t = round(sv(row[mxn_total_col]) if mxn_total_col < len(row) else 0, 2)
                    usd_t = round(sv(row[usd_total_col]) if usd_total_col and usd_total_col < len(row) else 0, 2)
                    if mxn_t or usd_t:
                        area = _area_from_concepto_rancho(label, "") or cat[3:]
                        area_groups[area] = {
                            "mxn_ranches": {}, "usd_ranches": {},
                            "mxn_t": mxn_t, "usd_t": usd_t,
                        }

                for area, ag in area_groups.items():
                    mano_obra_data.append({
                        "semana":      code,
                        "year":        year,
                        "week":        ww,
                        "date_range":  date_range,
                        "subcat":      area,
                        "mxn_total":   ag["mxn_t"],
                        "usd_total":   ag["usd_t"],
                        "mxn_ranches": ag["mxn_ranches"],
                        "usd_ranches": ag["usd_ranches"],
                    })
            else:
                all_data.append({
                    "semana":      code,
                    "year":        year,
                    "week":        ww,
                    "date_range":  date_range,
                    "categoria":   cat,
                    "mxn_total":   round(sv(row[mxn_total_col]) if mxn_total_col < len(row) else 0, 2),
                    "usd_total":   round(sv(row[usd_total_col]) if usd_total_col and usd_total_col < len(row) else 0, 2),
                    "mxn_ranches": mxn_ranches,
                    "usd_ranches": usd_ranches,
                })

        #  Extraer filas de siembra (89-92 Excel = labels fijos) 
        wk_siembra: dict = {}
        for field_key, field_label in SIEMBRA_LABELS:
            for row in data:
                cell_text = " ".join(str(row[c]).strip().upper() for c in range(min(5, len(row))))
                if field_label in cell_text:
                    # Total MXN (columna total)
                    total_val = sv(row[mxn_total_col]) if mxn_total_col < len(row) else 0
                    wk_siembra.setdefault("TOTAL", {})[field_key] = total_val
                    # Por rancho
                    for j, rn in mxn_ranch_cols.items():
                        if j < len(row):
                            wk_siembra.setdefault(rn, {})[field_key] = sv(row[j])
                    break
        if wk_siembra:
            siembra_data[code] = wk_siembra

        #  Extraer Indicadores y Costos Unitarios 
        wk_unit_costs = {}
        curr_section = None
        for i_r, row in enumerate(data):
            cell_text = " ".join(str(row[c]).strip().upper() for c in range(min(5, len(row))))
            if not cell_text.strip(): continue
            
            if "$ / TALLO" in cell_text or "$ / LIBRAS" in cell_text or "COSTOS UNITARIOS" in cell_text:
                curr_section = "tallo"
            elif "$ / HECTAREA" in cell_text or "$ / HECTÁREA" in cell_text:
                curr_section = "ha"
            elif "<<< INDICADORES" in cell_text or "KPI'S" in cell_text or "PROYECTOS" in cell_text:
                curr_section = None
                
            label = str(row[1]).strip().upper() if len(row) > 1 else ""
            if not label:
                label = str(row[0]).strip().upper() if len(row) > 0 else ""
                
            key = None
            #  COSTOS UNITARIOS $ / TALLO PROCESADO 
            # Los labels en el Excel son simples: "Materiales", "Mano de Obra", etc.
            # curr_section ya identifica la sección; NO requieren "TALLO" en el label.
            if curr_section == "tallo":
                if "MANO DE OBRA PROD" in label:
                    key = "mano_obra_prod_tallo"
                elif "$ / TALLO" in label:
                    key = "tallo_procesados"
                elif "MATERIAL DE EMPAQUE" in label or "EMPAQUE" in label:
                    key = "empaque_tallo"
                elif "SANIDAD VEGETAL" in label or "SANIDAD" in label:
                    key = "sanidad_tallo"
                elif "FERTILIZ" in label or "FERTLIZ" in label:
                    key = "fertilizacion_tallo"
                elif "MANO DE OBRA" in label:
                    key = "mano_obra_tallo"
                elif "MATERIALES" in label or (label == "MATERIALES"):
                    key = "materiales_tallo"
                elif "SERVICIOS" in label or "FLETES" in label:
                    key = "servicios_tallo"
                elif "COSTO DE PRODUCCION" in label or "COSTO DE PRODUCCIÓN" in label:
                    if "VENTAS" in label:
                        key = "cpv_tallo"

            #  COSTOS UNITARIOS $ / HECTÁREA 
            # Misma lógica: labels simples, curr_section identifica la sección.
            # Bug anterior: precedencia `and` > `or` causaba matches incorrectos.
            elif curr_section == "ha":
                if "$ / HECTAREA" in label or "$ / HECTÁREA" in label:
                    key = "hectareas_ha"   # fila-encabezado con los totales de Ha por rancho
                elif "MANO DE OBRA PROD" in label:
                    key = "mano_obra_prod_ha"
                elif "MATERIAL DE EMPAQUE" in label or "EMPAQUE" in label:
                    key = "empaque_ha"
                elif "SANIDAD VEGETAL" in label or "SANIDAD" in label:
                    key = "sanidad_ha"
                elif "FERTILIZ" in label or "FERTLIZ" in label:
                    key = "fertilizacion_ha"
                elif "MANO DE OBRA" in label:
                    key = "mano_obra_ha"
                elif "MATERIALES" in label or (label == "MATERIALES"):
                    key = "materiales_ha"
                elif "SERVICIOS" in label or "FLETES" in label:
                    key = "servicios_ha"
                elif "COSTO DE PRODUCCION" in label or "COSTO DE PRODUCCIÓN" in label:
                    if "VENTAS" in label:
                        key = "cpv_ha"
                    
            if key:
                if wk_unit_costs.get("TOTAL", {}).get(key) is not None:
                    continue
                total_val = sv(row[mxn_total_col]) if mxn_total_col < len(row) else 0
                wk_unit_costs.setdefault("TOTAL", {})[key] = total_val
                for j, rn in mxn_ranch_cols.items():
                    if j < len(row):
                        wk_unit_costs.setdefault(rn, {})[key] = sv(row[j])
                        
        if wk_unit_costs:
            unit_costs_data[code] = wk_unit_costs

    print(f"\n[OK] servicios_data: {len(servicios_data)} registros encontrados")
    if servicios_data:
        print(f"   subcats: {list({r['subcat'] for r in servicios_data})}")
    print(f"[OK] mano_obra_data: {len(mano_obra_data)} registros encontrados")

    cats_found = {r["categoria"] for r in all_data}
    if mano_obra_data:
        cats_found.add("COSTO MANO DE OBRA")
    cats  = [c for c in CATEGORIAS_ORDEN if c in cats_found]
    years = sorted({r["year"] for r in all_data})

    ranches_seen: set = set()
    for r in all_data:
        ranches_seen.update(r["mxn_ranches"])
        ranches_seen.update(r["usd_ranches"])
    ranches = sorted(ranches_seen)

    summary: dict = {cat: {yr: {"usd": 0.0, "mxn": 0.0, "ranches": {}, "ranches_mxn": {}}
                            for yr in years} for cat in cats}
    for r in all_data:
        s = summary.get(r["categoria"], {}).get(r["year"])
        if not s:
            continue
        s["usd"] += r["usd_total"]
        s["mxn"] += r["mxn_total"]
        for rn, v in r["usd_ranches"].items():
            s["ranches"][rn] = round(s["ranches"].get(rn, 0) + v, 2)
        for rn, v in r["mxn_ranches"].items():
            s["ranches_mxn"][rn] = round(s["ranches_mxn"].get(rn, 0) + v, 2)
    for cat in cats:
        for yr in years:
            d = summary[cat][yr]
            d["usd"] = round(d["usd"], 2)
            d["mxn"] = round(d["mxn"], 2)

    weeks_per_year: dict = {}
    week_date_ranges: dict = {}
    for r in all_data:
        weeks_per_year.setdefault(r["year"], set()).add(r["week"])
        key = f"{r['year']}-{r['week']}"
        if key not in week_date_ranges and r.get("date_range"):
            week_date_ranges[key] = r["date_range"]
    weeks_per_year = {yr: sorted(wks) for yr, wks in weeks_per_year.items()}

    return {
        "years":            years,
        "categories":       cats,
        "ranches":          ranches,
        "summary":          summary,
        "weeks_per_year":   weeks_per_year,
        "week_date_ranges": week_date_ranges,
        "weekly_detail":    all_data,
        "productos":        productos,
        "productos_debug":  productos_debug,
        "servicios_data":   servicios_data,
        "mano_obra_data":   mano_obra_data,
        "siembra_data":     siembra_data,
        "unit_costs_data":  unit_costs_data,
    }


#  Punto de entrada público 
def _extraer_conteo_marlen() -> list:
    """
    Lee la hoja 'Conteo' del Excel de Nómina/Conteo de Marlen en SharePoint.

    Formato esperado de la hoja:
      Fila 1: "CENTRO FLORICULTOR DE B.C."  (título)
      Fila 2: "CONTEO DE PERSONAL..."        (subtítulo)
      Fila 3: vacía
      Fila 4: encabezados  Sem | UBICACIÓN | ÁREA / DEPARTAMENTO | CONTEO
      Fila 5+: datos      2615 | PROPAGACION | ING. Y ADMON. | 3

    Sem usa el formato YYWW (ej: 2615 = año 2026, semana 15), igual que los códigos WK.

    Retorna lista de registros con el mismo esquema que _extraer_mano_obra_conteo():
      semana, year, week, date_range, subcat (Área), mxn_total=0, usd_total=0,
      hc_total, mxn_ranches={}, usd_ranches={}, hc_ranches={rancho: conteo}
    """
    # Mapa UBICACIÓN del Excel  nombre de rancho normalizado del sistema
    _UBICACION_MAP = {
        "PROPAGACION":   "Propagacion",
        "PROPAGACIÓN":   "Propagacion",
        "ADMINISTRACION":"Administracion",
        "ADMINISTRACIÓN":"Administracion",
        "POSCOSECHA":    "Poscosecha",
        "POSCO":         "Poscosecha",
        "RAMONA":        "Ramona",
        "ISABELA":       "Isabela",
        "ISABELLA":      "Isabela",
        "CHRISTINA":     "Christina",
        "CRISTINA":      "Christina",
        "CECILIA 25":    "Cecilia 25",
        "CECILIA25":     "Cecilia 25",
        "CECILIA":       "Cecilia",
    }

    # Mapa ÁREA / DEPARTAMENTO  subcat normalizado (mismo que usa el BD existente)
    _AREA_MAP = {
        "ING. Y ADMON.":      "Ing. Y Admon.",
        "ING Y ADMON":        "Ing. Y Admon.",
        "SUPERVISORES":       "Supervisores",
        "CORTE":              "Corte",
        "TRASPLANTE":         "Trasplante",
        "TRANSPLANTE":        "Trasplante",
        "MANEJO PLANTA":      "Manejo P.",
        "CONSOLIDACIÓN":      "Consolidacion",
        "CONSOLIDACION":      "Consolidacion",
        "SIEMBRA":            "Siembra",
        "MOV. CHAROLAS":      "Mov. Charolas",
        "MOV CHAROLAS":       "Mov. Charolas",
        "RIEGO":              "Riego",
        "PHLOX":              "Esquejes",
        "ESQUEJES":           "Esquejes",
        "HOOPS":              "Hoops",
        "MIPE / MIRFE":       "MIPE Y MIRFE",
        "MIPE/MIRFE":         "MIPE Y MIRFE",
        "MIPE Y MIRFE":       "MIPE Y MIRFE",
        "TRACTORES/CAMEROS":  "Tract. Y Cameros",
        "TRACTORES Y CAMEROS":"Tract. Y Cameros",
        "VELADORES":          "Veladores",
        "SOLDADORES":         "Soldadores",
        "TRANSPORTE":         "Transporte",
        "ADMON POSCO":        "Admon Posco",
        "ADMON. POSCO":       "Admon Posco",
        "ALM. UPC Y EMPAQUE": "Alm.upc y empaq",
        "ALM UPC Y EMPAQUE":  "Alm.upc y empaq",
        "PROD. PÁTINA Y REC": "Prod. Patina y rec",
        "PROD. PATINA Y REC": "Prod. Patina y rec",
        "PROD PATINA Y REC":  "Prod. Patina y rec",
    }

    print(f"[DL] Descargando Conteo Marlen desde: {SHAREPOINT_URL_CONTEO_MARLEN}")
    archivo = _descargar_con_graph(SHAREPOINT_URL_CONTEO_MARLEN, "Conteo Marlen")
    if archivo is None:
        print("[WARN]  No se pudo descargar Conteo Marlen  retornando vacío")
        return []
    print(f"[OK] Descarga OK, tamaño={len(archivo.getvalue())} bytes")

    try:
        df_raw = pd.read_excel(archivo, sheet_name="Conteo", header=None).fillna("")
        print(f"[OK] Hoja 'Conteo' leída: {df_raw.shape[0]} filas × {df_raw.shape[1]} cols")
    except Exception as e:
        print(f"[WARN]  Error leyendo hoja 'Conteo': {e}")
        return []

    #  Detectar fila de encabezados 
    header_idx = None
    for i in range(min(10, len(df_raw))):
        row_up = [str(v).strip().upper() for v in df_raw.iloc[i].values]
        if "SEM" in row_up and ("UBICACIÓN" in row_up or "UBICACION" in row_up):
            header_idx = i
            break

    if header_idx is None:
        print("[WARN]  No se encontró fila de encabezados en hoja 'Conteo'")
        return []

    print(f"    Encabezados detectados en fila {header_idx}")

    # Mapear columnas por nombre
    hdrs = [str(v).strip().upper() for v in df_raw.iloc[header_idx].values]
    def _find_col(keywords):
        for kw in keywords:
            if kw in hdrs:
                return hdrs.index(kw)
        return None

    col_sem   = _find_col(["SEM", "SEMANA"])
    col_ubic  = _find_col(["UBICACIÓN", "UBICACION"])
    col_area  = _find_col(["ÁREA / DEPARTAMENTO", "AREA / DEPARTAMENTO",
                            "ÁREA/DEPARTAMENTO",   "AREA/DEPARTAMENTO",
                            "ÁREA", "AREA"])
    col_cont  = _find_col(["CONTEO"])

    missing = [n for n, c in [("Sem", col_sem), ("Ubicación", col_ubic),
                               ("Área", col_area), ("Conteo", col_cont)] if c is None]
    if missing:
        print(f"[WARN]  Columnas faltantes en 'Conteo': {missing}")
        return []

    #  Leer datos 
    # Acumular por (semana_code, area)  {rancho: conteo_sum}
    from collections import defaultdict
    acum: dict = defaultdict(lambda: defaultdict(float))  # (code, area)  {rancho: float}

    for i in range(header_idx + 1, len(df_raw)):
        row = df_raw.iloc[i].values

        sem_raw  = str(row[col_sem]).strip()  if col_sem  < len(row) else ""
        ubic_raw = str(row[col_ubic]).strip() if col_ubic < len(row) else ""
        area_raw = str(row[col_area]).strip() if col_area < len(row) else ""
        cont_raw = row[col_cont]              if col_cont < len(row) else ""

        # Código de semana: debe ser numérico de 4 dígitos tipo YYWW
        try:
            code = int(float(sem_raw))
        except (ValueError, TypeError):
            continue
        if not (1800 <= code <= 9999):  # sanity check
            continue

        # Normalizar ubicación y área
        ubic_up = re.sub(r'\s+', ' ', ubic_raw.upper().strip())
        area_up = re.sub(r'\s+', ' ', area_raw.upper().strip())

        rancho = _UBICACION_MAP.get(ubic_up)
        if not rancho:
            # Búsqueda parcial como fallback
            for k, v in _UBICACION_MAP.items():
                if k in ubic_up or ubic_up in k:
                    rancho = v
                    break
        if not rancho:
            continue

        area = _AREA_MAP.get(area_up)
        if not area:
            # Búsqueda parcial
            for k, v in _AREA_MAP.items():
                if k in area_up:
                    area = v
                    break
        if not area:
            area = area_raw.strip()  # mantener tal cual si no mapea

        # Conteo numérico
        try:
            conteo = float(str(cont_raw).replace(",", "").strip()) if str(cont_raw).strip() not in ("", "nan") else 0.0
        except (ValueError, TypeError):
            conteo = 0.0

        acum[(code, area)][rancho] += conteo

    #  Construir registros 
    result = []
    for (code, area), hc_ranches in acum.items():
        yy   = code // 100
        ww   = code % 100
        year = 2000 + yy
        hc_total = sum(hc_ranches.values())
        result.append({
            "semana":      code,
            "year":        year,
            "week":        ww,
            "date_range":  "",
            "subcat":      area,
            "mxn_total":   0.0,
            "usd_total":   0.0,
            "hc_total":    hc_total,
            "mxn_ranches": {},
            "usd_ranches": {},
            "hc_ranches":  dict(hc_ranches),
        })

    print(f"[OK] Conteo Marlen: {len(result)} registros ({len({r['semana'] for r in result})} semanas)")
    return result


def _extraer_mano_obra_conteo() -> list:
    """
    Lee el Excel de conteo de personal desde SharePoint.
    """
    print(f"[DL] Descargando conteo desde: {SHAREPOINT_URL_CONTEO}")
    archivo = _descargar_con_graph(SHAREPOINT_URL_CONTEO, "Conteo Personal")
    if archivo is None:
        print("[WARN]  No se pudo descargar conteo.xlsx  mano_obra_data vacío")
        return []
    print(f"[OK] Descarga OK, tamaño={len(archivo.getvalue())} bytes")

    try:
        df = pd.read_excel(archivo, sheet_name="BD", header=2)
        print(f"[OK] Excel leído: {df.shape[0]} filas, columnas={list(df.columns)}")
    except Exception as e:
        print(f"[WARN]  Error leyendo conteo.xlsx: {e}")
        return []

    df.columns = [str(c).strip() for c in df.columns]
    needed = {"Año", "Semana", "Área", "Rancho", "Costo MN", "Costo DLLS", "Conteo"}
    missing = needed - set(df.columns)
    if missing:
        print(f"[WARN]  Conteo.xlsx  columnas faltantes: {missing}")
        return []

    #  Diagnóstico columna Conteo 
    print(f"[FIND] Conteo dtype  : {df['Conteo'].dtype}")
    print(f"[FIND] Conteo sample : {df['Conteo'].head(10).tolist()}")
    print(f"[FIND] Conteo no-nulos: {df['Conteo'].notna().sum()} de {len(df)} filas")
    # Forzar conversión numérica por si llegan como string o formula
    df["Conteo"] = pd.to_numeric(df["Conteo"], errors="coerce").fillna(0.0)

    df = df.dropna(subset=["Año", "Semana", "Área"])
    df["Año"]    = pd.to_numeric(df["Año"],    errors="coerce")
    df["Semana"] = pd.to_numeric(df["Semana"], errors="coerce")
    df = df.dropna(subset=["Año", "Semana"])
    df["Año"]    = df["Año"].astype(int)
    df["Semana"] = df["Semana"].astype(int)

    def _sv(v):
        try:
            s = str(v).strip().replace(",", "").replace(" ", "")
            if not s or s in ("-", "-   ", " -   "):
                return 0.0
            return round(float(s), 2)
        except:
            return 0.0

    result = []
    for (anio, semana, area), grp in df.groupby(["Año", "Semana", "Área"]):
        code = (int(anio) - 2000) * 100 + int(semana)
        mxn_ranches, usd_ranches, hc_ranches = {}, {}, {}
        mxn_total = usd_total = hc_total = 0.0
        for _, row in grp.iterrows():
            rancho     = str(row.get("Rancho", "")).strip()
            costo_mn   = _sv(row.get("Costo MN",   0))
            costo_dlls = _sv(row.get("Costo DLLS", 0))
            conteo_val = _sv(row.get("Conteo", 0))
            mxn_total += costo_mn
            usd_total += costo_dlls
            hc_total  += conteo_val
            if rancho:
                mxn_ranches[rancho] = round(mxn_ranches.get(rancho, 0.0) + costo_mn,   2)
                usd_ranches[rancho] = round(usd_ranches.get(rancho, 0.0) + costo_dlls, 2)
                hc_ranches[rancho]  = hc_ranches.get(rancho, 0.0) + conteo_val
        result.append({
            "semana":      code,
            "year":        int(anio),
            "week":        int(semana),
            "date_range":  "",
            "subcat":      str(area).strip(),
            "mxn_total":   round(mxn_total, 2),
            "usd_total":   round(usd_total, 2),
            "hc_total":    hc_total,
            "mxn_ranches": mxn_ranches,
            "usd_ranches": usd_ranches,
            "hc_ranches":  hc_ranches,
        })

    print(f"[OK] conteo mano_obra_data: {len(result)} registros desde conteo.xlsx")
    return result


def _descargar_con_graph(url: str, label: str) -> BytesIO | None:
    import base64
    try:
        import streamlit as st
        tenant_id     = st.secrets["sharepoint"]["tenant_id"]
        client_id     = st.secrets["sharepoint"]["client_id"]
        client_secret = st.secrets["sharepoint"]["client_secret"]
        
        token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
        resp = requests.post(token_url, data={
            "client_id": client_id,
            "scope": "https://graph.microsoft.com/.default",
            "client_secret": client_secret,
            "grant_type": "client_credentials"
        })
        token = resp.json().get("access_token")
        if not token:
            print(f"[ERR] {label}: MS Graph auth falló (revisa secrets)")
            return None
            
        base64_value = base64.urlsafe_b64encode(url.encode("utf-8")).decode("utf-8").rstrip("=")
        graph_url = f"https://graph.microsoft.com/v1.0/shares/u!{base64_value}/driveItem/content"
        
        dl_resp = requests.get(graph_url, headers={"Authorization": f"Bearer {token}"})
        if dl_resp.status_code == 200:
            return BytesIO(dl_resp.content)
        else:
            print(f"[ERR] {label}: Falló ({dl_resp.status_code}) -> {dl_resp.text[:200]}")
            return None
    except Exception as e:
        print(f"[ERR] Error en _descargar_con_graph para {label}: {e}")
        return None

#  Detalle de Metros Acumulados (Material Vegetal) 
def _extraer_metros_acumulados() -> list:
    """
    Lee la hoja 'Mtrs Acumulados' del Excel de siembra.
    Se usa MS Graph API porque este link no es público y requiere autenticación.
    """
    print(f"[DL] Descargando Metros Acumulados desde SharePoint vía MS Graph...")
    url_limpia = SHAREPOINT_URL_SIEMBRA_DETALLE.split("?")[0]
    archivo = _descargar_con_graph(url_limpia, "Metros Acumulados")
    if archivo is None:
        print("[WARN]  No se pudo descargar el Excel de Metros Acumulados")
        return []

    try:
        xls = pd.ExcelFile(archivo)
    except Exception as e:
        print(f"[WARN]  No se pudo abrir el Excel de Metros Acumulados: {e}")
        return []

    SHEET_NAME = "Mtrs Acumulados"
    if SHEET_NAME not in xls.sheet_names:
        # intento case-insensitive
        match = next((s for s in xls.sheet_names if s.strip().lower() == SHEET_NAME.lower()), None)
        if not match:
            print(f"[WARN]  Hoja '{SHEET_NAME}' no encontrada. Hojas disponibles: {xls.sheet_names}")
            return []
        SHEET_NAME = match

    try:
        df_raw = pd.read_excel(xls, sheet_name=SHEET_NAME, header=None).fillna("")
    except Exception as e:
        print(f"[WARN]  Error leyendo hoja '{SHEET_NAME}': {e}")
        return []

    # Autodetectar la fila de encabezados buscando 'Rancho' y 'Flor'
    header_idx = 0
    for i in range(min(10, len(df_raw))):
        row_strs = [str(v).strip().lower() for v in df_raw.iloc[i].values]
        if 'rancho' in row_strs and 'flor' in row_strs:
            header_idx = i
            break

    df = df_raw.iloc[header_idx + 1:].copy()
    df.columns = [str(c).strip() for c in df_raw.iloc[header_idx].values]
    print(f"[OK] Metros Acumulados leído: {df.shape[0]} filas, encabezados detectados en fila {header_idx + 1}")

    cols = list(df.columns)
    COL_RANCHO   = cols[0] if len(cols) > 0 else "Rancho"
    COL_FLOR     = cols[1] if len(cols) > 1 else "Flor"
    COL_METROS   = cols[2] if len(cols) > 2 else "Metros"
    COL_PLA_ACUM = cols[3] if len(cols) > 3 else "Pla. Acum."
    COL_SEMANA   = cols[4] if len(cols) > 4 else "Semana"

    def _to_float(v):
        try:
            return round(float(str(v).replace(",", "").strip()), 2)
        except Exception:
            return 0.0

    result = []
    for _, row in df.iterrows():
        rancho_raw = str(row.get(COL_RANCHO, "")).strip()
        flor       = str(row.get(COL_FLOR,   "")).strip()
        sem_raw    = str(row.get(COL_SEMANA, "")).strip()
        metros_v   = row.get(COL_METROS,   0)
        pla_v      = row.get(COL_PLA_ACUM, 0)

        if not rancho_raw or not flor:
            continue

        # Normalizar nombre de rancho: CECILIACecilia, RAMONACampo-RM, etc.
        if rancho_raw.upper() == "RAMONA":
            rancho = "Campo-RM"
        else:
            rancho = norm_ranch(rancho_raw)
            if not rancho:
                rancho = rancho_raw  # mantener original si no mapea

        # Extraer código de semana final: '2302 - 2616' o '2616' o '2616.0'
        semana_fin = None
        if "-" in sem_raw:
            partes = sem_raw.split("-")
            try:
                semana_fin = int(float(str(partes[-1]).strip()))
            except ValueError:
                pass
        else:
            try:
                semana_fin = int(float(sem_raw.strip()))
            except ValueError:
                pass

        if semana_fin is None:
            continue

        result.append({
            "semana_fin":   semana_fin,
            "rancho":       rancho,
            "flor":         flor,
            "metros":       _to_float(metros_v),
            "pla_acum":     _to_float(pla_v),
            "semana_rango": sem_raw,
        })

    print(f"[OK] metros_acumulados: {len(result)} registros cargados")
    return result



#  Detalle de Plantas (Charolas Sembradas) 
def _extraer_plantas_metros() -> list:
    """
    Lee la hoja 'Pl.-Mtrs' del Excel de siembra.
    Columnas: Rancho, Flor, Plantas, Metros, Semana, Año.
    """
    print(f"[DL] Descargando Pl.-Mtrs desde SharePoint vía MS Graph...")
    url_limpia = SHAREPOINT_URL_SIEMBRA_DETALLE.split("?")[0]
    archivo = _descargar_con_graph(url_limpia, "Plantas-Metros")
    if archivo is None:
        print("[WARN]  No se pudo descargar el Excel de Pl.-Mtrs")
        return []

    try:
        xls = pd.ExcelFile(archivo)
    except Exception as e:
        print(f"[WARN]  No se pudo abrir el Excel de Pl.-Mtrs: {e}")
        return []

    SHEET_NAME = "Pl.-Mtrs"
    if SHEET_NAME not in xls.sheet_names:
        match = next((s for s in xls.sheet_names if s.strip().lower() == SHEET_NAME.lower()), None)
        if not match:
            print(f"[WARN]  Hoja '{SHEET_NAME}' no encontrada. Hojas disponibles: {xls.sheet_names}")
            return []
        SHEET_NAME = match

    try:
        df_raw = pd.read_excel(xls, sheet_name=SHEET_NAME, header=None).fillna("")
    except Exception as e:
        print(f"[WARN]  Error leyendo hoja '{SHEET_NAME}': {e}")
        return []

    # Detección de encabezados
    header_idx = 0
    for i in range(min(10, len(df_raw))):
        row_strs = [str(v).strip().lower() for v in df_raw.iloc[i].values]
        if 'rancho' in row_strs and 'flor' in row_strs and 'plantas' in row_strs:
            header_idx = i
            break

    df = df_raw.iloc[header_idx + 1:].copy()
    df.columns = [str(c).strip() for c in df_raw.iloc[header_idx].values]
    print(f"[OK] Pl.-Mtrs leído: {df.shape[0]} filas, detectados en fila {header_idx + 1}")

    cols = list(df.columns)
    COL_RANCHO  = cols[0] if len(cols) > 0 else "Rancho"
    COL_FLOR    = cols[1] if len(cols) > 1 else "Flor"
    COL_PLANTAS = cols[2] if len(cols) > 2 else "Plantas"
    COL_METROS  = cols[3] if len(cols) > 3 else "Metros"
    COL_SEMANA  = cols[4] if len(cols) > 4 else "Semana"
    COL_ANO     = cols[5] if len(cols) > 5 else "Año"

    def _to_float(v):
        try: return round(float(str(v).replace(",", "").strip()), 2)
        except: return 0.0

    result = []
    for _, row in df.iterrows():
        rancho_raw = str(row.get(COL_RANCHO, "")).strip()
        flor       = str(row.get(COL_FLOR,   "")).strip()
        sem        = str(row.get(COL_SEMANA, "")).strip()
        ano        = str(row.get(COL_ANO,    "")).strip()
        plantas_v  = row.get(COL_PLANTAS, 0)
        metros_v   = row.get(COL_METROS,  0)

        if not rancho_raw or not flor:
            continue

        if rancho_raw.upper() == "RAMONA": rancho = "Campo-RM"
        else:
            rancho = norm_ranch(rancho_raw)
            if not rancho: rancho = rancho_raw
        
        try:
            week_i = int(float(sem))
            year_i = int(float(ano))
        except ValueError:
            continue

        # Convertir 2023, 1 -> 2301
        semana_fin = (year_i % 100) * 100 + week_i

        result.append({
            "semana_fin": semana_fin,
            "rancho":     rancho,
            "flor":       flor,
            "plantas":    _to_float(plantas_v),
            "metros":     _to_float(metros_v)
        })

    print(f"[OK] plantas_metros: {len(result)} registros cargados")
    return result


#  Detalle de Esquejes 
def _extraer_esquejes() -> list:
    """
    Lee la hoja 'Esquejes' del Excel de siembra.
    Columnas: FLOR, CHAROLA, Cantidad, PLANTAS, SEMANA, Año.
    """
    print(f"[DL] Descargando Esquejes desde SharePoint vía MS Graph...")
    url_limpia = SHAREPOINT_URL_SIEMBRA_DETALLE.split("?")[0]
    archivo = _descargar_con_graph(url_limpia, "Esquejes")
    if archivo is None:
        print("[WARN]  No se pudo descargar el Excel para Esquejes")
        return []

    try:
        xls = pd.ExcelFile(archivo)
    except Exception as e:
        print(f"[WARN]  No se pudo abrir el Excel de Esquejes: {e}")
        return []

    SHEET_NAME = "Esquejes"
    if SHEET_NAME not in xls.sheet_names:
        match = next((s for s in xls.sheet_names if s.strip().lower() == SHEET_NAME.lower()), None)
        if not match:
            print(f"[WARN]  Hoja '{SHEET_NAME}' no encontrada. Hojas disponibles: {xls.sheet_names}")
            return []
        SHEET_NAME = match

    try:
        df_raw = pd.read_excel(xls, sheet_name=SHEET_NAME, header=None).fillna("")
    except Exception as e:
        print(f"[WARN]  Error leyendo hoja '{SHEET_NAME}': {e}")
        return []

    # Detección de encabezados
    header_idx = 0
    for i in range(min(15, len(df_raw))):
        row_strs = [str(v).strip().lower() for v in df_raw.iloc[i].values]
        if 'flor' in row_strs and 'plantas' in row_strs and 'semana' in row_strs:
            header_idx = i
            break

    df = df_raw.iloc[header_idx + 1:].copy()
    df.columns = [str(c).strip() for c in df_raw.iloc[header_idx].values]
    print(f"[OK] Esquejes leído: {df.shape[0]} filas, detectados en fila {header_idx + 1}")

    # Convertir columnas a minúsculas para encontrar sin importar mayúsculas
    col_map = {c.strip().lower(): c for c in df.columns}
    
    COL_FLOR    = col_map.get("flor", "FLOR")
    COL_PLANTAS = col_map.get("plantas", "PLANTAS")
    COL_SEMANA  = col_map.get("semana", "SEMANA")
    COL_ANO     = col_map.get("año", "Año")

    def _to_float(v):
        try: return round(float(str(v).replace(",", "").strip()), 2)
        except: return 0.0

    result = []
    for _, row in df.iterrows():
        flor       = str(row.get(COL_FLOR,   "")).strip()
        sem        = str(row.get(COL_SEMANA, "")).strip()
        plantas_v  = row.get(COL_PLANTAS, 0)

        if not flor or not sem:
            continue
            
        ano_str = str(row.get(COL_ANO, "")).strip()
        try:
            week_i = int(float(sem))
            
            if week_i > 1000:
                semana_fin = week_i
            else:
                if ano_str:
                    year_i = int(float(ano_str))
                    semana_fin = (year_i % 100) * 100 + week_i
                else:
                    continue
        except ValueError:
            continue

        result.append({
            "semana_fin": semana_fin,
            "flor":       flor,
            "plantas":    _to_float(plantas_v)
        })

    print(f"[OK] esquejes: {len(result)} registros cargados")
    return result


def _extraer_detalle_weekly() -> dict:
    """
    Lee el Excel de detalle semanal (hojas WEEKLY####) desde SharePoint.

    Columnas relevantes (0-based):
      F(5)  FLOR              clave de agrupación
      G(6)  INV.INICIAL       inv_inicial
      H-L   Recepción/rancho  tallos_proc (suma)
      O(14) PRAS A TERC       tallos_comp
      S(18) EXPORTACIÓN       parte de tallos_desp
      T(19) MUESTRAS          parte de tallos_desp
      V(21) DESECHOS          tallos_des
      W(22) INV.FINAL calc.   inv_final

    Retorna:
      {
        2552: {
          'inv_inicial': [{'flor': 'CELOSIA', 'valor': 110}, ...],
          'tallos_proc': [...],
          'tallos_comp': [...],
          'tallos_desp': [...],
          'tallos_des':  [...],
          'inv_final':   [...],
        },
        '2552': <misma referencia>,
        ...
      }
    """
    # Fuentes WEEKLY: histórico + 2026 (misma estructura, se combinan)
    _WEEKLY_SOURCES = [
        (SHAREPOINT_URL_WEEKLY,      "Excel WEEKLY histórico"),
        (SHAREPOINT_URL_WEEKLY_2026, "Excel WEEKLY 2026"),
    ]

    result: dict = {}

    def _safe(row, col_idx: int) -> float:
        """Extrae valor numérico seguro de una fila pandas."""
        if col_idx >= len(row):
            return 0.0
        v = row.iloc[col_idx]
        if v is None or v == "":
            return 0.0
        try:
            return float(str(v).replace(",", "").replace("$", "").strip())
        except (ValueError, TypeError):
            return 0.0

    def _procesar_xls(xls, label: str) -> None:
        """Procesa todas las hojas WEEKLY#### de un ExcelFile y acumula en result."""
        for sheet in xls.sheet_names:
            m = re.match(r"WEEKLY(\d{4})", sheet.strip().upper())
            if not m:
                continue
            wk_code = int(m.group(1))

            # Si ya fue cargado por otra fuente, no sobreescribir
            if wk_code in result:
                print(f"     WEEKLY{wk_code} ya existe ({label}), se omite")
                continue

            try:
                df = pd.read_excel(xls, sheet_name=sheet, header=None).fillna("")

                # Detectar columnas dinámicamente por nombre de encabezado
                c = _detectar_columnas_weekly(df)

                # Columnas de recepción por rancho (para TALLOS COSECHADOS)
                # Nombres normalizados para coincidir con el sistema
                _RANCH_COLS = [
                    (c["cec"], "Cecilia"),
                    (c["ram"], "Campo-RM"),
                    (c["isa"], "Isabela"),
                    (c["chr"], "Christina"),
                    (c["c25"], "Cecilia 25"),
                ]
                ranch_totals = {name: 0.0 for _, name in _RANCH_COLS}

                wk_data: dict = {
                    "inv_inicial": [],
                    "tallos_cos":  [],   # por flor (suma de todas las columnas de rancho)
                    "tallos_proc": [],   # EXPORTACION + MUESTRAS
                    "tallos_comp": [],
                    "tallos_desp": [],   # EXPORTACION + MUESTRAS (igual que tallos_proc)
                    "tallos_des":  [],
                    "inv_final":   [],
                }
                cos_por_flor: dict = {}   # (flor, rancho) -> valor

                for _, row in df.iterrows():
                    if len(row) <= c["flor"]:
                        continue
                    flor = str(row.iloc[c["flor"]]).strip()
                    if (
                        not flor
                        or flor in ("0", "nan", "")
                        or flor.upper() in ("FLOR", "TOTAL", "TOTALES")
                        or re.match(r"^(SEMANA|INVENTARIO|TOTAL)", flor.upper())
                    ):
                        continue

                    inv_ini = _safe(row, c["ini"])
                    export  = _safe(row, c["export"])
                    muest   = _safe(row, c["muest"])
                    des     = _safe(row, c["des"])
                    inv_fin = _safe(row, c["inv_fin"])
                    proc    = export + muest   # TALLOS PROC. TOTALES = EXPORTACION + MUESTRAS

                    # tallos_comp: un registro por proveedor (col) con su nombre
                    _prov_names = c.get("comp_names", ["COMPRAS"] * len(c["comp_cols"]))
                    for col_idx, prov_name in zip(c["comp_cols"], _prov_names):
                        val = _safe(row, col_idx)
                        if val != 0.0:
                            wk_data["tallos_comp"].append(
                                {"flor": flor, "proveedor": prov_name, "valor": val}
                            )
                    comp = sum(_safe(row, col) for col in c["comp_cols"])  # total (para compatibilidad)

                    for key, val in [
                        ("inv_inicial", inv_ini),
                        ("tallos_proc", proc),
                        ("tallos_desp", proc),
                        ("tallos_des",  des),
                        ("inv_final",   inv_fin),
                    ]:
                        if val != 0.0:
                            wk_data[key].append({"flor": flor, "valor": val})

                    # tallos_cos: acumular por flor+rancho individualmente
                    for col_idx, ranch_name in _RANCH_COLS:
                        val = _safe(row, col_idx)
                        ranch_totals[ranch_name] += val
                        if val != 0.0:
                            key2 = (flor, ranch_name)
                            cos_por_flor[key2] = cos_por_flor.get(key2, 0.0) + val

                wk_data["tallos_cos"] = [
                    {"flor": f, "rancho": rn, "valor": round(v, 0)}
                    for (f, rn), v in cos_por_flor.items()
                    if v != 0.0
                ]

                result[wk_code]      = wk_data
                result[str(wk_code)] = wk_data
                print(f"   [OK] WEEKLY{wk_code} [{label}]: "
                      f"inv_ini={len(wk_data['inv_inicial'])} "
                      f"cos={len(wk_data['tallos_cos'])} ranchos "
                      f"proc={len(wk_data['tallos_proc'])} "
                      f"comp={len(wk_data['tallos_comp'])} "
                      f"desp={len(wk_data['tallos_desp'])} "
                      f"des={len(wk_data['tallos_des'])} "
                      f"inv_fin={len(wk_data['inv_final'])} filas")

            except Exception as e:
                print(f"   [WARN] Error leyendo hoja {sheet} [{label}]: {e}")

    for url, label in _WEEKLY_SOURCES:
        bio = _descargar_con_graph(url, label)
        if bio is None:
            print(f"[WARN] No se pudo descargar {label}  se continúa con las demás fuentes.")
            continue
        try:
            xls = pd.ExcelFile(bio)
            print(f"[DL] Procesando {label} ({len(xls.sheet_names)} hojas)...")
            _procesar_xls(xls, label)
        except Exception as e:
            print(f"[ERR] Error abriendo {label}: {e}")

    print(f"[OK] detalle_weekly: {len(result)//2} semanas cargadas en total")
    return result


def _extraer_horas_transporte() -> dict:
    url = GOOGLE_DRIVE_URL_TRANSPORTE
    print(f"[DL] Descargando Horas de Transporte desde Google Drive")
    try:
        import requests
        from io import BytesIO
        r = requests.get(url, timeout=30)
        r.raise_for_status()
        df_raw = pd.read_excel(BytesIO(r.content), header=None).fillna("")
    except Exception as e:
        print(f"[WARN]  Error leyendo horas de transporte: {e}")
        return {}

    header_idx = None
    for i in range(min(10, len(df_raw))):
        row = [str(x).strip().upper() for x in df_raw.iloc[i].values]
        if "FECHA" in row and "HORAS POSCO" in row:
            header_idx = i
            break
            
    if header_idx is None:
        return {}
        
    df = df_raw.iloc[header_idx+1:].copy()
    df.columns = [str(x).strip() for x in df_raw.iloc[header_idx].values]
    
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df = df.dropna(subset=["Fecha"])
    
    col_map = {
        "Horas posco": "PosCo-RM",
        "Horas Cecilia": "Cecilia",
        "Horas Cecilia 25": "Cecilia 25",
        "Horas Cristina": "Christina",
        "Horas Isabela": "Isabela",
        "Horas Ramona": "Campo-RM",
    }
    
    result = {}
    for _, row in df.iterrows():
        fecha = row["Fecha"]
        
        # Agrupamos los 7 días de Miércoles a Martes 
        fecha_shifted = fecha - pd.Timedelta(days=2)
        yr, wk, _ = fecha_shifted.isocalendar()
        code = (yr % 100) * 100 + wk
        
        if code not in result:
            result[code] = {}
            
        for col_excel, ranch_name in col_map.items():
            if col_excel in df.columns:
                val = row[col_excel]
                try:
                    val = float(val)
                except:
                    val = 0.0
                if val == val:
                    result[code][ranch_name] = result[code].get(ranch_name, 0.0) + val
                    
    print(f"[OK] Horas transporte procesadas: {len(result)} semanas")
    return result


def _extraer_tractores() -> dict:
    url = GOOGLE_DRIVE_URL_TRANSPORTE
    print(f"[DL] Descargando Tractores desde Google Drive")
    try:
        import requests
        from io import BytesIO
        r = requests.get(url, timeout=30)
        r.raise_for_status()
        df_raw = pd.read_excel(BytesIO(r.content), sheet_name="Tractores", header=None).fillna("")
    except Exception as e:
        print(f"[WARN]  Error leyendo tractores: {e}")
        return {}

    totales_indices = []
    for idx, row in df_raw.iterrows():
        val = str(row[1]).strip().upper() if len(row) > 1 else ""
        if "TOTALES" in val:
            totales_indices.append(idx)

    parsed_data = {}
    ranch_mapping = [
        {"ranch": "Cecilia", "c_col": 5, "h_col": 6},
        {"ranch": "Cecilia 25", "c_col": 7, "h_col": 8},
        {"ranch": "Christina", "c_col": 9, "h_col": 10},
        {"ranch": "Isabela", "c_col": 11, "h_col": 12},
        {"ranch": "Campo-RM", "c_col": 13, "h_col": 14}
    ]

    def to_float(v):
        try:
            f = float(str(v).replace(",", "").strip())
            return f if f == f else 0.0
        except:
            return 0.0

    for tot_idx in totales_indices:
        header_idx = None
        for i in range(tot_idx - 1, max(0, tot_idx - 40), -1):
            val_sem = str(df_raw.iloc[i, 0]).strip().upper()
            if "SEMANA" in val_sem:
                header_idx = i
                break
        
        if header_idx is not None:
            sem_val = ""
            for r_idx in range(header_idx + 1, tot_idx):
                sem_val = str(df_raw.iloc[r_idx, 0]).strip()
                if sem_val and sem_val != "nan":
                    break
            
            try:
                sem_code = int(float(sem_val))
            except ValueError:
                continue
                
            parsed_data[sem_code] = {}
            
            for r_idx in range(header_idx + 2, tot_idx):
                row = df_raw.iloc[r_idx].values
                labor_name = str(row[1]).strip()
                
                if not labor_name or labor_name.upper() in ["nan", "TOTALES", "TOTAL", "SEMANA", "LABORES"]:
                    continue
                
                ranch_data = {}
                tiene_valores = False
                
                for r_info in ranch_mapping:
                    ranch_name = r_info["ranch"]
                    c_col = r_info["c_col"]
                    h_col = r_info["h_col"]
                    
                    c_val = row[c_col] if len(row) > c_col else 0.0
                    h_val = row[h_col] if len(row) > h_col else 0.0
                    
                    c_float = round(to_float(c_val), 2)
                    h_float = round(to_float(h_val), 4)
                    
                    if c_float > 0.01 or h_float > 0.01:
                        tiene_valores = True
                        
                    ranch_data[ranch_name] = {
                        "camas": c_float,
                        "horas": h_float
                    }
                
                if tiene_valores:
                    parsed_data[sem_code][labor_name] = ranch_data
            
            # También agregar el total general de la semana
            tot_row = df_raw.iloc[tot_idx].values
            tot_ranch_data = {}
            for r_info in ranch_mapping:
                ranch_name = r_info["ranch"]
                c_col = r_info["c_col"]
                h_col = r_info["h_col"]
                
                c_val = tot_row[c_col] if len(tot_row) > c_col else 0.0
                h_val = tot_row[h_col] if len(tot_row) > h_col else 0.0
                
                tot_ranch_data[ranch_name] = {
                    "camas": round(to_float(c_val), 2),
                    "horas": round(to_float(h_val), 4)
                }
            parsed_data[sem_code]["TOTALES"] = tot_ranch_data

    print(f"[OK] Tractores procesados: {len(parsed_data)} semanas")
    return parsed_data


def get_datos() -> dict:
    """
    [>>] VERSIÓN PARALELA  ThreadPoolExecutor
    - Hojas WK   Excel principal en SharePoint  (secuencial, es la base)
    - PR/MP/ME/MV/Conteo/Weekly/Metros/Plantas/Transporte/Tractores
       se descargan TODAS en paralelo después del WK.

    Tiempo estimado:
      Antes (secuencial):  ~60 segundos
      Ahora  (paralelo):   ~12-15 segundos
    """
    #  1. Excel WK principal  debe ir primero (es la base de todo) 
    print("\n" + "=" * 60)
    print("[DL] [1/2] DESCARGANDO EXCEL WK PRINCIPAL (base)")
    print("=" * 60)
    archivo = _descargar_con_graph(SHAREPOINT_URL_WK, "Excel WK")
    if archivo is None:
        return {"error": "No se pudo descargar el archivo WK de SharePoint."}

    try:
        xls = pd.ExcelFile(archivo)
    except Exception as e:
        return {"error": f"No se pudo abrir el Excel WK: {e}"}

    resultado = extraer_datos(xls)

    if "error" in resultado:
        return resultado

    #  2. Lanzar TODAS las fuentes secundarias en PARALELO 
    print("\n" + "=" * 60)
    print("[>>] [2/2] DESCARGANDO FUENTES SECUNDARIAS EN PARALELO")
    print("=" * 60)

    def _fetch_mv_wrapper(rows):
        return _parse_generic(rows, mv_mode=True)

    def _tarea_pr():
        return ("pr", _fetch_desde_sharepoint("PR", _parse_generic, "PR"))

    def _tarea_mp():
        return ("mp", _fetch_desde_sharepoint("MP", _parse_generic, "MP"))

    def _tarea_me():
        return ("me", _fetch_desde_sharepoint("ME", _parse_generic, "ME"))

    def _tarea_mv():
        return ("mv", _fetch_desde_sharepoint("MV", _fetch_mv_wrapper, "MV"))

    def _tarea_conteo():
        return ("conteo", _extraer_mano_obra_conteo())

    def _tarea_weekly():
        return ("weekly", _extraer_detalle_weekly())

    def _tarea_metros():
        return ("metros", _extraer_metros_acumulados())

    def _tarea_plantas():
        return ("plantas", _extraer_plantas_metros())

    def _tarea_transporte():
        return ("transporte", _extraer_horas_transporte())

    def _tarea_tractores():
        return ("tractores", _extraer_tractores())

    def _tarea_esquejes():
        return ("esquejes", _extraer_esquejes())

    tareas = [
        _tarea_pr, _tarea_mp, _tarea_me, _tarea_mv,
        _tarea_conteo, _tarea_weekly,
        _tarea_metros, _tarea_plantas,
        _tarea_transporte, _tarea_tractores, _tarea_esquejes,
    ]

    resultados_paralelos = {}
    with ThreadPoolExecutor(max_workers=10, thread_name_prefix="cfbc") as executor:
        futures = {executor.submit(t): t.__name__ for t in tareas}
        for future in as_completed(futures):
            try:
                key, value = future.result()
                resultados_paralelos[key] = value
                print(f"   [OK] [{key}] listo")
            except Exception as e:
                print(f"   [ERR] [{futures[future]}] error: {e}")

    #  3. Integrar resultados 
    print("\n" + "=" * 60)
    print("[LINK] INTEGRANDO RESULTADOS")
    print("=" * 60)

    # PR  merge con lo que ya traía el WK
    if resultados_paralelos.get("pr"):
        productos, productos_debug = resultados_paralelos["pr"]
        resultado["productos"].update(productos)
        resultado["productos_debug"].update(productos_debug)

    # MP
    if resultados_paralelos.get("mp"):
        resultado["productos_mp"], resultado["productos_mp_debug"] = resultados_paralelos["mp"]
    else:
        resultado["productos_mp"] = {}; resultado["productos_mp_debug"] = {}

    # ME
    if resultados_paralelos.get("me"):
        resultado["productos_me"], resultado["productos_me_debug"] = resultados_paralelos["me"]
    else:
        resultado["productos_me"] = {}; resultado["productos_me_debug"] = {}

    # MV
    if resultados_paralelos.get("mv"):
        resultado["productos_mv"], resultado["productos_mv_debug"] = resultados_paralelos["mv"]
    else:
        resultado["productos_mv"] = {}; resultado["productos_mv_debug"] = {}

    # Config de ranchos
    HIDDEN_RANCHES = {"Albahaca-RM", "Campo-VI"}
    resultado["config"] = {
        "ranch_order":  [k for k in RANCH_CONFIG.keys() if k not in HIDDEN_RANCHES],
        "ranch_colors": {k: v["color"] for k, v in RANCH_CONFIG.items() if k not in HIDDEN_RANCHES},
    }

    # Mano de Obra  merge de 3 fuentes (lógica original intacta)
    #  Fuente 1: Conteo Marlen (hoja "Conteo", formato nuevo YYWW) 
    # TEMPORAL: Marlen deshabilitado mientras se termina el Excel
    # marlen_data_raw = _extraer_conteo_marlen()
    # marlen_data = [r for r in marlen_data_raw if r.get("hc_total", 0) > 0]
    marlen_data = []
    marlen_keys = set()

    #  Fuente 2: Conteo antiguo BD (sólo semanas que Marlen no cubre) 
    conteo_data_raw = resultados_paralelos.get("conteo") or []
    conteo_data = [
        r for r in conteo_data_raw
        if (r.get("mxn_total", 0) > 0 or r.get("usd_total", 0) > 0 or r.get("hc_total", 0) > 0)
        and (r["year"], r["week"]) not in marlen_keys
    ]
    conteo_keys = marlen_keys | {(r["year"], r["week"]) for r in conteo_data}

    #  Fuente 3: WK Excel (sólo semanas sin conteo de ninguna fuente) 
    wk_mano_obra = resultado.get("mano_obra_data", [])
    merged_mano_obra = list(marlen_data) + list(conteo_data)
    wk_added_count = 0
    for r in wk_mano_obra:
        if (r["year"], r["week"]) not in conteo_keys:
            merged_mano_obra.append(r)
            wk_added_count += 1

    resultado["mano_obra_data"] = merged_mano_obra
    print(
        f"[OK] Merge Conteo: {len(marlen_data)} reg. Marlen "
        f"+ {len(conteo_data)} reg. BD antiguo "
        f"+ {wk_added_count} reg. WK Excel (fallback)"
    )

    # CRITICAL: inyectar semanas de mano_obra en listas globales
    # para que el frontend JS (allWeeks / rangeWeeks) las reconozca
    for r in merged_mano_obra:
        yr = r["year"]
        wk = r["week"]
        if yr not in resultado["years"]:
            resultado["years"].append(yr)
            resultado["years"].sort()

        w_per_year = resultado.get("weeks_per_year", {})
        if yr not in w_per_year:
            w_per_year[yr] = []
        if wk not in w_per_year[yr]:
            w_per_year[yr].append(wk)
            w_per_year[yr].sort()
        resultado["weeks_per_year"] = w_per_year

        key = f"{yr}-{wk}"
        if key not in resultado["week_date_ranges"]:
            resultado["week_date_ranges"][key] = r.get("date_range", f"W{wk:02d}")

    # Resto de fuentes
    resultado["metros_acumulados"] = resultados_paralelos.get("metros")    or []
    resultado["plantas_metros"]    = resultados_paralelos.get("plantas")   or []
    resultado["detalle_weekly"]    = resultados_paralelos.get("weekly")    or {}
    resultado["horas_transporte"]  = resultados_paralelos.get("transporte") or {}
    resultado["tractores"]         = resultados_paralelos.get("tractores") or {}
    resultado["esquejes_data"]     = resultados_paralelos.get("esquejes") or []

    print(f"\n[OK] get_datos() completado  {len(resultados_paralelos)}/11 fuentes integradas.")
    return resultado


# --- Construir hoja WK en blanco con estructura fija ---
def _construir_hoja_wk(ws, nombre_hoja: str):
    """
    Escribe la estructura completa de una hoja WK#### con formato IDENTICO al Excel de SharePoint.
    Colores, negritas, bordes y rellenos exactos.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    #  Helpers 
    NAVY  = "333399"   # color texto principal
    GRAY  = "44546A"   # color texto datos rancho
    WHITE = "FFFFFF"

    def _f(bold=False, size=10, color=NAVY, name="Calibri"):
        return Font(bold=bold, size=size, color=color, name=name)

    def _fill(hex_color):
        if not hex_color or hex_color in ("", "none"):
            return PatternFill(fill_type=None)
        c = hex_color.lstrip("FF") if len(hex_color) == 8 else hex_color
        if len(c) != 6:
            return PatternFill(fill_type=None)
        return PatternFill("solid", fgColor=c)

    def _al(h="general", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    thin   = Side(style="thin")
    medium = Side(style="medium")
    none_s = Side(style=None)

    def _bdr(left=None, right=None, top=None, bottom=None):
        return Border(left=left or none_s, right=right or none_s,
                      top=top or none_s,   bottom=bottom or none_s)

    # Rellenos
    fill_green  = _fill("CCFFCC")   # verde claro USD   (FFCCFFCC en real)
    fill_blue   = _fill("DAE3F3")   # azul claro encabezado
    fill_lime   = _fill("C5E0B4")   # verde lima codigo semana
    fill_orange = _fill("FFCC99")   # naranja subtotales (FFFFCC99)
    fill_yellow = _fill("FFFFCC")   # amarillo produccion
    fill_white  = _fill("FFFFFF")
    fill_kpi    = _fill("008000")   # verde oscuro KPI headers

    # Bordes reutilizables
    bdr_L_med         = _bdr(left=medium)
    bdr_L_med_R_thin  = _bdr(left=medium, right=thin)
    bdr_R_med         = _bdr(right=medium)
    bdr_L_R_thin      = _bdr(left=thin, right=thin)

    #  Ancho de columnas 
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 69.4
    ws.column_dimensions["C"].width = 14
    for col in ("D","E","F","G","H","I","J"):
        ws.column_dimensions[col].width = 11
    ws.column_dimensions["K"].width = 3
    for col in ("L","M","N","O","P","Q","R","S"):
        ws.column_dimensions[col].width = 11

    #  Fila 1 
    ws["B1"].value = "CENTRO FLORICULTOR DE BAJA CALIFORNIA, S.A. DE C.V. "
    ws["B1"].font  = _f(bold=True)

    #  Fila 2 
    ws["B2"].value = "SEMANA DE CALCULO - Mexico"
    ws["B2"].font  = _f(bold=True)
    ws["B2"].fill  = fill_blue
    ws["B2"].alignment = _al("center")

    #  Fila 3 
    code = nombre_hoja[2:] if nombre_hoja.upper().startswith("WK") else nombre_hoja
    try:
        _tc = 20 if 2502 <= int(code) <= 2520 else 19
    except (ValueError, TypeError):
        _tc = 19
    ws["B3"].value = code
    ws["B3"].font  = _f(bold=True)
    ws["B3"].fill  = fill_lime
    ws["B3"].alignment = _al("center")
    ws["C3"].value = _tc;  ws["C3"].font = _f(bold=True)
    ws["C3"].border = _bdr(bottom=medium)
    ws["D3"].value = " tipo de cambio"; ws["D3"].font = _f(bold=True)
    ws["L3"].value = _tc;  ws["L3"].font = _f(bold=True)
    ws["M3"].value = "  tipo de cambio "; ws["M3"].font = _f(bold=True)

    #  Fila 4 
    ws["B4"].value = "Del ___ al ___ de ________ 20__"
    ws["B4"].alignment = _al("center")
    ws.row_dimensions[4].height = 15

    #  Fila 5 
    ws.merge_cells("C5:J5")
    ws["C5"].value = "(MXN) Pesos Mexicanos"
    ws["C5"].alignment = _al("center")
    ws["C5"].border = _bdr(left=medium, right=medium, top=medium, bottom=thin)
    ws.merge_cells("L5:R5")
    ws["L5"].value = "US Dollars"
    ws["L5"].fill  = fill_green
    ws["L5"].alignment = _al("center")
    ws["L5"].border = _bdr(left=medium, top=medium, bottom=thin)
    ws["S5"].fill  = fill_green
    ws["S5"].border = _bdr(right=medium, top=medium, bottom=thin)

    #  Fila 6 
    ws["B6"].value = "TOTAL FINCA"
    ws["B6"].fill  = fill_white
    ws["B6"].alignment = _al("center")
    ws["B6"].border = bdr_L_med
    ws["C6"].border = _bdr(left=medium, right=thin, top=thin)
    for col in ("L","M","N","O","P","Q","R"):
        ws[f"{col}6"].fill = fill_green
        ws[f"{col}6"].alignment = _al("center")
    ws["L6"].border = _bdr(left=medium, right=thin)
    ws["S6"].fill  = fill_green
    ws["S6"].border = bdr_R_med
    ws["S6"].alignment = _al("center")
    ws.row_dimensions[6].height = 26.4

    #  Fila 7 
    ws["B7"].value = "Produccion"
    ws["B7"].fill  = fill_white
    ws["B7"].alignment = _al("center")
    ws["B7"].border = bdr_L_med
    headers_mxn = ["TOTAL","Prop-RM","PosCo-RM","Campo -RM","Isabela","Christina","Cecilia","Cecilia 25"]
    headers_usd = ["TOTAL","Prop-RM","PosCo-RM","Campo -RM","ISABELA","Christina","CECILIA","CECILIA 25"]
    for i, h in enumerate(headers_mxn):
        col = chr(ord("C")+i)
        ws[f"{col}7"].value = h
        ws[f"{col}7"].font  = _f(bold=(i==0))
        ws[f"{col}7"].alignment = _al("center")
    ws["C7"].border = _bdr(left=medium, right=thin)
    ws["J7"].border = bdr_R_med
    usd_cols = ["L","M","N","O","P","Q","R","S"]
    for i, h in enumerate(headers_usd):
        c = usd_cols[i]
        ws[f"{c}7"].value = h
        ws[f"{c}7"].fill  = fill_green
        ws[f"{c}7"].alignment = _al("center")
    ws["L7"].border = _bdr(left=medium, right=thin)
    ws["S7"].border = bdr_R_med

    #  Fila 8 
    ws["C8"].value = "SEMANAL "
    ws["C8"].alignment = _al("center")
    ws["C8"].border = _bdr(left=medium, right=thin)
    ws["L8"].value = '"WEEKLY"'
    ws["L8"].fill  = fill_green
    ws["L8"].alignment = _al("center")
    ws["L8"].border = _bdr(left=medium, right=thin)
    for col in ("M","N","O","P","Q","R"):
        ws[f"{col}8"].fill = fill_green
    ws["S8"].fill  = fill_green
    ws["S8"].border = bdr_R_med

    #  Fila 9 
    ws["B9"].value = "EJECUCION SEMANAL"
    ws["B9"].font  = _f(bold=True)
    ws["B9"].fill  = fill_white
    ws["B9"].alignment = _al("center")
    ws["B9"].border = _bdr(left=medium, bottom=thin)
    ws["C9"].border = _bdr(left=medium, right=thin, bottom=thin)
    for col in ("D","E","F","G","H","I"):
        ws[f"{col}9"].value = 1
        ws[f"{col}9"].alignment = _al("center")
        ws[f"{col}9"].border = _bdr(bottom=thin)
    ws["J9"].value = 1
    ws["J9"].alignment = _al("center")
    ws["J9"].border = _bdr(right=medium, bottom=thin)
    ws["L9"].fill  = fill_green
    ws["L9"].border = _bdr(left=medium, right=thin, bottom=thin)
    for col in ("M","N","O","P","Q","R"):
        ws[f"{col}9"].fill  = fill_green
        ws[f"{col}9"].border = _bdr(bottom=thin)
    ws["S9"].fill  = fill_green
    ws["S9"].border = _bdr(right=medium, bottom=thin)

    #  Helper fila de categoría 
    def _fila_cat(row, label, fill_usd=None, top_border=False):
        if fill_usd is None:
            fill_usd = fill_green
        ws[f"B{row}"].value = label
        ws[f"B{row}"].font  = _f()
        ws[f"B{row}"].fill  = fill_white
        ws[f"B{row}"].alignment = _al("left")
        ws[f"B{row}"].border = bdr_L_med
        ws[f"C{row}"].value = 0
        ws[f"C{row}"].font  = _f(bold=True)
        ws[f"C{row}"].border = bdr_L_med_R_thin
        ws[f"C{row}"].number_format = '#,##0;-#,##0;"-   "'
        for dc in ("D","E","F","G","H","I"):
            ws[f"{dc}{row}"].value = 0
            ws[f"{dc}{row}"].font  = _f(color=GRAY)
            ws[f"{dc}{row}"].number_format = '#,##0;-#,##0;"-   "'
        ws[f"J{row}"].value = 0
        ws[f"J{row}"].font  = _f(color=GRAY)
        ws[f"J{row}"].border = bdr_R_med
        ws[f"J{row}"].number_format = '#,##0;-#,##0;"-   "'
        ws[f"L{row}"].value = 0
        ws[f"L{row}"].font  = _f(bold=True)
        ws[f"L{row}"].fill  = fill_usd
        if top_border:
            ws[f"L{row}"].border = _bdr(left=medium, right=thin, top=thin)
        else:
            ws[f"L{row}"].border = bdr_L_med_R_thin
        ws[f"L{row}"].number_format = '#,##0;-#,##0;"-   "'
        for uc in ("M","N","O","P","Q","R"):
            ws[f"{uc}{row}"].value = 0
            ws[f"{uc}{row}"].fill  = fill_usd
            ws[f"{uc}{row}"].alignment = _al("center")
            ws[f"{uc}{row}"].number_format = '#,##0;-#,##0;" -   "'
            if top_border:
                ws[f"{uc}{row}"].border = _bdr(top=thin)
        ws[f"S{row}"].value = 0
        ws[f"S{row}"].fill  = fill_usd
        ws[f"S{row}"].alignment = _al("center")
        ws[f"S{row}"].number_format = '#,##0;-#,##0;" -   "'
        if top_border:
            ws[f"S{row}"].border = _bdr(right=medium, top=thin)
        else:
            ws[f"S{row}"].border = bdr_R_med

    def _fila_blank(row, fill_usd=None):
        if fill_usd is None:
            fill_usd = fill_green
        ws[f"B{row}"].fill  = fill_white
        ws[f"B{row}"].border = bdr_L_med
        ws[f"C{row}"].border = bdr_L_med_R_thin
        ws[f"J{row}"].border = bdr_R_med
        ws[f"L{row}"].fill  = fill_usd
        ws[f"L{row}"].border = bdr_L_med_R_thin
        for uc in ("M","N","O","P","Q","R"):
            ws[f"{uc}{row}"].fill  = fill_usd
        ws[f"S{row}"].fill  = fill_usd
        ws[f"S{row}"].border = bdr_R_med

    def _fila_subtotal(row, label):
        ws[f"B{row}"].value = label
        ws[f"B{row}"].font  = _f(bold=True)
        ws[f"B{row}"].fill  = fill_white
        ws[f"B{row}"].border = bdr_L_med
        for col in ("C","D","E","F","G","H","I"):
            ws[f"{col}{row}"].value = 0
            ws[f"{col}{row}"].font  = _f(bold=True)
            ws[f"{col}{row}"].fill  = fill_orange
            ws[f"{col}{row}"].alignment = _al("center")
            ws[f"{col}{row}"].number_format = '#,##0;-#,##0;"-   "'
        ws[f"C{row}"].border = bdr_L_med_R_thin
        ws[f"J{row}"].value = 0
        ws[f"J{row}"].font  = _f(bold=True)
        ws[f"J{row}"].fill  = fill_orange
        ws[f"J{row}"].alignment = _al("center")
        ws[f"J{row}"].border = bdr_R_med
        ws[f"J{row}"].number_format = '#,##0;-#,##0;"-   "'
        for col in ("L","M","N","O","P","Q","R"):
            ws[f"{col}{row}"].value = 0
            ws[f"{col}{row}"].font  = _f(bold=True)
            ws[f"{col}{row}"].fill  = fill_orange
            ws[f"{col}{row}"].alignment = _al("center")
            ws[f"{col}{row}"].border = _bdr(top=thin, bottom=thin)
            ws[f"{col}{row}"].number_format = '#,##0;-#,##0;"-   "'
        ws[f"L{row}"].border = _bdr(left=medium, right=thin, top=thin, bottom=thin)
        ws[f"S{row}"].value = 0
        ws[f"S{row}"].font  = _f(bold=True)
        ws[f"S{row}"].fill  = fill_orange
        ws[f"S{row}"].alignment = _al("center")
        ws[f"S{row}"].border = _bdr(right=medium, top=thin, bottom=thin)
        ws[f"S{row}"].number_format = '#,##0;-#,##0;"-   "'

    #  Filas 10-20: Materiales 
    categorias = [
        (10, "DESINFECCION Y FERTILIZACION"),
        (11, "AMPLIACION "),
        (12, "CULTIVO TIERRA, CHAROLAS"),
        (13, "MATERIAL VEGETAL"),
        (14, "PREPARACION DE SUELO"),
        (15, "FERTILIZANTES (Manejo Integrado de Riego y Fertilizacion) "),
        (16, "DESINFECCION / PLAGUICIDAS (Manejo Integrado de Plagas y Enfermedades)"),
        (17, "MANTENIMIENTO"),
        (18, "EXPANSION CECILIA 25"),
        (19, "RENOVACION DE SIEMBRA"),
        (20, "MATERIAL DE EMPAQUE"),
    ]
    for i, (row, label) in enumerate(categorias):
        _fila_cat(row, label, top_border=(i == 0))
    _fila_blank(21)
    _fila_subtotal(22, "COSTO DE MATERIALES")
    _fila_blank(23)

    #  Filas 24-59: Nominas 
    nominas = [
        (24, "NOMINA ADMON Oficina, Jefes de Finca, Ingenieros"),
        (25, "HORAS EXTR. DOM. Y FESTIVOS"),
        (26, "BONOS ASISIT, PUNTAULIDAD Y DESPENSA"),
        (27, "NOMINA PRODUCCION "),
        (28, "HORAS EXTR. DOM. Y FEST."),
        (29, "BONOS ASISIT, PUNT. Y DESP."),
        (30, "NOMINA PRODUCCION CORTE"),
        (31, "HORAS EXTR. DOM. Y FESTIVOS CORTE"),
        (32, "BONOS ASISIT, PUNTAULIDAD Y DESP. CORTE"),
        (33, "NOMINA PRODUCCION TRANSPLANTE"),
        (34, "HORAS EXTR. DOM. Y FEST. TRANSPLANTE"),
        (35, "BONOS ASISIT, PUNT. Y DESP. TRANSPLANTE"),
        (36, "NOMINA PRODUCCION MANEJO PLANTA"),
        (37, "HORAS EXTR. DOM. Y FEST. MANEJO PLANTA"),
        (38, "BONOS ASISIT, PUNT. Y DESP. MANEJO PLANTA"),
        (39, "NOMINA  HOOPS"),
        (40, "HORAS EXTR. DOM. Y FEST. HOOPS"),
        (41, "BONOS ASISIT, PUNT. Y DESP.HOOPS"),
        (42, "NOMINA  (MIPE,MIRFE,)"),
        (43, "HORAS EXTR. DOM. Y FEST. (MIPE,MIRFE)"),
        (44, "BONOS ASISIT, PUNT. Y DESP.(MIPE,MIRFE)"),
        (45, "NOMINA OPERATIVOS (TRACTORES, CAMEROS)"),
        (46, "HORAS EXTR. DOM. Y FEST. (TRACTORES, CAMEROS)"),
        (47, "BONOS ASISIT, PUNT. Y DESP. (TRACTORES, CAMEROS)"),
        (48, "NOMINA OPERATIVOS (CHOFER)"),
        (49, "HORAS EXTR. DOM. Y FEST. (CHOFER)"),
        (50, "BONOS ASISIT, PUNT. Y DESP. (CHOFER)"),
        (51, "NOMINA OPERATIVOS (VELADORES)"),
        (52, "HORAS EXTR. DOM. Y FEST. (VELADORES)"),
        (53, "BONOS ASISIT, PUNT. Y DESP. (VELADORES)"),
        (54, "NOMINA OPERATIVOS (SOLDADOR)"),
        (55, "HORAS EXTR. DOM. Y FEST. (SOLDADOR)"),
        (56, "BONOS ASISIT, PUNT. Y DESP. (SOLDADOR)"),
        (57, "NOMINA PRODUCCION Contratista y comisiones"),
        (58, "IMSS , INFONAVIT RCV"),
        (59, "1.8% al estado (1.2% tasa efectiva)"),
    ]
    for row, label in nominas:
        _fila_cat(row, label)
    _fila_blank(60)
    _fila_subtotal(61, "COSTO DE MANO DE OBRA")
    _fila_blank(62)

    #  Filas 63-70: Servicios 
    servicios = [
        (63, "ELECTRICIDAD"),
        (64, "FLETES Y ACARREOS (Flete aduana)"),
        (65, "GASTOS DE EXPORTACION "),
        (66, "CERTIFICADO DE FITOSANITARIOS"),
        (67, "Transporte de personal"),
        (68, "COMPRA DE FLOR A TERCEROS"),
        (69, "COMIDA PARA EL PERSONAL"),
        (70, "RO, TEL, RTA.ALIM."),
    ]
    for row, label in servicios:
        _fila_cat(row, label)
    _fila_blank(71)
    _fila_subtotal(72, "COSTO DE SERVICIOS")

    # Fila 73: separadora con bordes top/bottom
    ws["B73"].fill  = fill_white
    ws["B73"].border = _bdr(left=medium, top=thin, bottom=thin)
    ws["C73"].border = _bdr(left=medium, top=thin, bottom=thin)
    ws["L73"].fill  = fill_green; ws["L73"].border = bdr_L_med
    for col in ("M","N","O","P","Q","R"):
        ws[f"{col}73"].fill = fill_green
    ws["S73"].fill  = fill_green; ws["S73"].border = bdr_R_med

    #  Fila 74: COSTO DE PRODUCCION Y VENTAS 
    ws["B74"].value = "COSTO DE PRODUCCION Y VENTAS"
    ws["B74"].font  = _f(bold=True)
    ws["B74"].fill  = fill_white
    ws["B74"].border = _bdr(left=medium, bottom=medium)
    for col in ("D","E","F","G","H","I"):
        ws[f"{col}74"].value = 0
        ws[f"{col}74"].font  = _f(bold=True)
        ws[f"{col}74"].border = _bdr(bottom=medium)
        ws[f"{col}74"].number_format = '#,##0;-#,##0;"-   "'
    ws["C74"].value = 0; ws["C74"].font = _f(bold=True)
    ws["C74"].border = _bdr(left=medium, right=thin, top=thin, bottom=medium)
    ws["C74"].number_format = '#,##0;-#,##0;"-   "'
    ws["J74"].value = 0; ws["J74"].font = _f(bold=True)
    ws["J74"].border = _bdr(right=medium, bottom=medium)
    ws["J74"].number_format = '#,##0;-#,##0;"-   "'
    for col in ("M","N","O","P","Q","R"):
        ws[f"{col}74"].value = 0; ws[f"{col}74"].font = _f(bold=True)
        ws[f"{col}74"].fill  = fill_green
        ws[f"{col}74"].border = _bdr(top=thin, bottom=medium)
        ws[f"{col}74"].number_format = '#,##0;-#,##0;"-   "'
    ws["L74"].value = 0; ws["L74"].font = _f(bold=True)
    ws["L74"].fill  = fill_green
    ws["L74"].border = _bdr(left=medium, right=thin, top=thin, bottom=medium)
    ws["L74"].number_format = '#,##0;-#,##0;"-   "'
    ws["S74"].value = 0; ws["S74"].font = _f(bold=True)
    ws["S74"].fill  = fill_green
    ws["S74"].border = _bdr(right=medium, top=thin, bottom=medium)
    ws["S74"].number_format = '#,##0;-#,##0;"-   "'
    ws.row_dimensions[74].height = 15

    #  Filas 76-92: Produccion 
    produccion = [
        (76, "CAJAS PROCESADAS TOTALES"),
        (77, "INVENTARIO INICIAL"),
        (78, "TALLOS COSECHADOS"),
        (79, "TALLOS DESECHADOS"),
        (80, "TALLOS DESECHADOS sf"),
        (81, "TALLOS COMPRADOS"),
        (82, "TALLOS EN BOUQUETS O PROCESADOS"),
        (83, "TALLOS DESPACHADOS"),
        (84, "LIBRAS DESPACHADAS ALBAHACA"),
        (85, "TALLOS muestra"),
        (86, "INVENTARIO FINAL"),
        (87, "TALLOS PROCESADOS TOTALES"),
        (88, " CHAROLAS SEMBRADAS *288 PLUGS ="),
        (89, " NUMERO DE CHAROLAS SEMBRADAS "),
        (90, " NUMERO DE ESQUEJES SEMBRADOS"),
        (91, " METROS DE SIEMBRA"),
        (92, " HECTAREAS EN SIEMBRA"),
    ]
    for i, (row, label) in enumerate(produccion):
        first = (i == 0)
        last  = (i == len(produccion)-1)
        ws[f"B{row}"].value = label
        ws[f"B{row}"].fill  = fill_white
        ws[f"B{row}"].alignment = _al("left")
        b_bdr = _bdr(left=medium, top=medium) if first else (_bdr(left=medium, bottom=medium) if last else bdr_L_med)
        ws[f"B{row}"].border = b_bdr
        ws[f"C{row}"].value = 0
        ws[f"C{row}"].font  = _f(bold=True)
        c_bdr = _bdr(left=medium, right=thin, top=medium) if first else (_bdr(left=medium, right=thin, bottom=medium) if last else bdr_L_med_R_thin)
        ws[f"C{row}"].border = c_bdr
        for dc in ("D","E","F","G","H","I"):
            ws[f"{dc}{row}"].value = 0
            ws[f"{dc}{row}"].border = _bdr(top=medium) if first else (_bdr(bottom=medium) if last else _bdr())
        ws[f"J{row}"].value = 0
        ws[f"J{row}"].border = _bdr(right=medium, top=medium) if first else (_bdr(right=medium, bottom=medium) if last else bdr_R_med)
        ws[f"L{row}"].value = 0
        ws[f"L{row}"].font  = _f(bold=True)
        ws[f"L{row}"].fill  = fill_yellow
        ws[f"L{row}"].border = _bdr(left=medium, right=thin, top=medium) if first else (_bdr(left=medium, right=thin, bottom=medium) if last else bdr_L_med_R_thin)
        for uc in ("M","N","O","P","Q","R"):
            ws[f"{uc}{row}"].value = 0
            ws[f"{uc}{row}"].fill  = fill_yellow
            ws[f"{uc}{row}"].alignment = _al("center")
            ws[f"{uc}{row}"].border = _bdr(top=medium) if first else (_bdr(bottom=medium) if last else _bdr())
        ws[f"S{row}"].value = 0
        ws[f"S{row}"].fill  = fill_yellow
        ws[f"S{row}"].alignment = _al("center")
        ws[f"S{row}"].border = _bdr(right=medium, top=medium) if first else (_bdr(right=medium, bottom=medium) if last else bdr_R_med)

    ws.row_dimensions[92].height = 15

    #  Fila 93 
    ws["B93"].value = "<<< INDICADORES"
    ws["B93"].font  = _f(bold=True)
    ws.row_dimensions[93].height = 15

    #  Filas 94-121: Costos unitarios 
    ws["B94"].value = "COSTOS UNITARIOS"; ws["B94"].font = _f(bold=True)
    ws["B94"].border = _bdr(left=medium, top=medium)
    ws["L94"].fill  = fill_green; ws["L94"].border = _bdr(left=medium, right=thin, top=medium)
    for col in ("M","N","O","P","Q","R"): ws[f"{col}94"].fill = fill_green; ws[f"{col}94"].border = _bdr(top=medium)
    ws["S94"].fill  = fill_green; ws["S94"].border = _bdr(right=medium, top=medium)

    def _cu_row(row, label, bold=False, fill_b=None):
        ws[f"B{row}"].value = label; ws[f"B{row}"].font = _f(bold=bold)
        if fill_b: ws[f"B{row}"].fill = fill_b
        ws[f"B{row}"].border = bdr_L_med
        ws[f"C{row}"].value = 0; ws[f"C{row}"].font = _f(bold=True)
        ws[f"C{row}"].border = bdr_L_med_R_thin
        ws[f"L{row}"].value = 0; ws[f"L{row}"].font = _f(bold=True)
        ws[f"L{row}"].fill  = fill_green; ws[f"L{row}"].border = bdr_L_med_R_thin
        for col in ("M","N","O","P","Q","R"): ws[f"{col}{row}"].fill = fill_green
        ws[f"S{row}"].fill = fill_green; ws[f"S{row}"].border = bdr_R_med
        if fill_b:
            ws[f"C{row}"].fill = fill_b
            for col in ("L","M","N","O","P","Q","R","S"): ws[f"{col}{row}"].fill = fill_b

    _cu_row(95, "$ / Tallo Procesado", bold=True)
    _cu_row(96, "COSTOS UNITARIOS", bold=True)
    _cu_row(97, "$ / Libras Procesadas", bold=True)
    ws["L97"].border = _bdr(left=medium, right=thin, bottom=thin)

    _cu_row(98, "Materiales")
    ws["B98"].border = _bdr(left=medium, top=thin)
    ws["C98"].border = _bdr(left=medium, right=thin, top=thin)
    ws["L98"].border = _bdr(left=medium, right=thin, top=thin)
    for col in ("M","N","O","P","Q","R"): ws[f"{col}98"].border = _bdr(top=thin)
    ws["S98"].border = _bdr(right=medium, top=thin)

    _cu_row(99, "Mano de Obra")
    _cu_row(100, "Servicios (Fletes)")
    ws["B100"].border = _bdr(left=medium, bottom=thin)
    ws["C100"].border = _bdr(left=medium, right=thin, bottom=thin)
    ws["L100"].border = _bdr(left=medium, right=thin, bottom=thin)
    for col in ("M","N","O","P","Q","R"): ws[f"{col}100"].border = _bdr(bottom=thin)
    ws["S100"].border = _bdr(right=medium, bottom=thin)

    _cu_row(101, "Costo de Produccion y Ventas", bold=True, fill_b=fill_orange)
    ws["B101"].border = _bdr(left=medium, top=thin, bottom=thin)
    ws["C101"].border = _bdr(left=medium, right=thin, top=thin, bottom=thin)
    ws["L101"].border = _bdr(left=medium, right=thin, bottom=thin)
    ws["S101"].border = _bdr(right=medium, bottom=thin)

    # Spacers 102, 104, 107
    for row in (102, 104, 107):
        ws[f"C{row}"].border = bdr_L_med_R_thin
        ws[f"L{row}"].fill = fill_green; ws[f"L{row}"].border = bdr_L_med_R_thin
        for col in ("M","N","O","P","Q","R"): ws[f"{col}{row}"].fill = fill_green
        ws[f"S{row}"].fill = fill_green; ws[f"S{row}"].border = bdr_R_med

    _cu_row(103, "Material de Empaque / Tallo", bold=True)
    ws["B103"].border = _bdr(left=medium, top=thin, bottom=thin)
    ws["C103"].border = _bdr(left=medium, right=thin, top=thin, bottom=thin)
    ws["L103"].border = _bdr(left=medium, right=thin, top=thin, bottom=thin)
    for col in ("M","N","O","P","Q","R"): ws[f"{col}103"].border = _bdr(top=thin, bottom=thin)
    ws["S103"].border = _bdr(right=medium, top=thin, bottom=thin)

    _cu_row(105, "Sanidad Vegetal / Tallo", bold=True)
    ws["B105"].border = _bdr(left=medium, top=thin)
    ws["C105"].border = _bdr(left=medium, right=thin, top=thin)
    ws["L105"].border = _bdr(left=medium, right=thin, top=thin)
    for col in ("M","N","O","P","Q","R"): ws[f"{col}105"].border = _bdr(top=thin)
    ws["S105"].border = _bdr(right=medium, top=thin)

    _cu_row(106, "Fertlizacion / Tallo", bold=True)
    ws["B106"].border = _bdr(left=medium, bottom=thin)
    ws["C106"].border = _bdr(left=medium, right=thin, bottom=thin)
    ws["L106"].border = _bdr(left=medium, right=thin, bottom=thin)
    for col in ("M","N","O","P","Q","R"): ws[f"{col}106"].border = _bdr(bottom=thin)
    ws["S106"].border = _bdr(right=medium, bottom=thin)

    _cu_row(108, "Mano de Obra Prod / Tallo", bold=True)
    ws["B108"].border = _bdr(left=medium, top=thin, bottom=medium)
    ws["C108"].border = _bdr(left=medium, right=thin, top=thin, bottom=medium)
    ws["L108"].border = _bdr(left=medium, right=thin, top=thin, bottom=medium)
    for col in ("M","N","O","P","Q","R"): ws[f"{col}108"].border = _bdr(top=thin, bottom=medium)
    ws["S108"].border = _bdr(right=medium, top=thin, bottom=medium)
    ws.row_dimensions[108].height = 15
    ws.row_dimensions[109].height = 15

    #  Fila 110-121: $ / Hectarea 
    ws["B110"].value = "$ / Hectarea"; ws["B110"].font = _f(bold=True)
    ws["B110"].border = _bdr(left=medium, top=medium)
    ws["C110"].border = _bdr(left=medium, right=thin, top=medium, bottom=thin)
    ws["J110"].border = _bdr(right=medium, top=medium)
    ws["L110"].fill = fill_yellow; ws["L110"].border = _bdr(left=medium, right=thin, top=medium)
    for col in ("M","N","O","P","Q","R"):
        ws[f"{col}110"].fill = fill_yellow; ws[f"{col}110"].border = _bdr(top=medium)
    ws["S110"].fill = fill_yellow; ws["S110"].border = _bdr(right=medium, top=medium)

    def _ha_row(row, label, top_b=False, bottom_b=False, both_b=False):
        ws[f"B{row}"].value = label; ws[f"B{row}"].font = _f()
        ws[f"B{row}"].fill  = fill_white; ws[f"B{row}"].alignment = _al("left")
        if both_b:   ws[f"B{row}"].border = _bdr(left=medium, top=thin, bottom=thin)
        elif top_b:  ws[f"B{row}"].border = _bdr(left=medium, top=thin)
        elif bottom_b: ws[f"B{row}"].border = _bdr(left=medium, bottom=thin)
        else:        ws[f"B{row}"].border = bdr_L_med
        ws[f"C{row}"].value = 0; ws[f"C{row}"].font = _f(bold=True)
        ws[f"C{row}"].border = _bdr(left=medium, right=thin,
                                    top=(thin if top_b or both_b else none_s),
                                    bottom=(thin if bottom_b or both_b else none_s))
        ws[f"J{row}"].border = bdr_R_med
        ws[f"L{row}"].value = 0; ws[f"L{row}"].font = _f(bold=True)
        ws[f"L{row}"].fill  = fill_yellow
        ws[f"L{row}"].border = _bdr(left=medium, right=thin,
                                    top=(thin if top_b or both_b else none_s),
                                    bottom=(thin if bottom_b or both_b else none_s))
        for col in ("M","N","O","P","Q","R"):
            ws[f"{col}{row}"].fill = fill_yellow
            ws[f"{col}{row}"].border = _bdr(top=(thin if top_b or both_b else none_s),
                                            bottom=(thin if bottom_b or both_b else none_s))
        ws[f"S{row}"].fill = fill_yellow
        ws[f"S{row}"].border = _bdr(right=medium,
                                    top=(thin if top_b or both_b else none_s),
                                    bottom=(thin if bottom_b or both_b else none_s))

    _ha_row(111, "Materiales", top_b=True)
    _ha_row(112, "Mano de Obra")
    _ha_row(113, "Servicios (Fletes)", bottom_b=True)
    # spacer 115
    ws[f"C115"].border = bdr_L_med_R_thin
    ws["L115"].fill = fill_yellow; ws["L115"].border = bdr_L_med_R_thin
    for col in ("M","N","O","P","Q","R"): ws[f"{col}115"].fill = fill_yellow
    ws["S115"].fill = fill_yellow; ws["S115"].border = bdr_R_med

    _ha_row(114, "Costo de Produccion y Ventas", both_b=True)
    _ha_row(116, "Material de Empaque / Caja", both_b=True)
    ws["B116"].font = _f(bold=True)
    # spacer 117
    ws["C117"].border = bdr_L_med_R_thin
    ws["L117"].fill = fill_yellow; ws["L117"].border = bdr_L_med_R_thin
    for col in ("M","N","O","P","Q","R"): ws[f"{col}117"].fill = fill_yellow
    ws["S117"].fill = fill_yellow; ws["S117"].border = bdr_R_med

    _ha_row(118, "Sanidad Vegetal / Ha", top_b=True)
    ws["B118"].font = _f(bold=True)
    _ha_row(119, "Fertlizacion / Ha", bottom_b=True)
    ws["B119"].font = _f(bold=True)
    # spacer 120
    ws["C120"].border = bdr_L_med_R_thin
    ws["L120"].fill = fill_yellow; ws["L120"].border = bdr_L_med_R_thin
    for col in ("M","N","O","P","Q","R"): ws[f"{col}120"].fill = fill_yellow
    ws["S120"].fill = fill_yellow; ws["S120"].border = bdr_R_med

    ws.row_dimensions[121].height = 15
    ws["B121"].value = "Mano de Obra Prod / Ha"; ws["B121"].font = _f(bold=True)
    ws["B121"].fill = fill_white
    ws["B121"].border = _bdr(left=medium, top=thin, bottom=medium)
    ws["C121"].value = 0; ws["C121"].font = _f(bold=True)
    ws["C121"].border = _bdr(left=medium, right=thin, top=thin, bottom=medium)
    ws["L121"].value = 0; ws["L121"].font = _f(bold=True)
    ws["L121"].fill = fill_yellow
    ws["L121"].border = _bdr(left=medium, right=thin, top=thin, bottom=medium)
    for col in ("M","N","O","P","Q","R"):
        ws[f"{col}121"].fill = fill_yellow
        ws[f"{col}121"].border = _bdr(top=thin, bottom=medium)
    ws["S121"].fill = fill_yellow
    ws["S121"].border = _bdr(right=medium, top=thin, bottom=medium)

    #  KPI's 
    ws["B124"].value = "KPI's "; ws["B124"].font = _f(bold=True)

    # Proyectos de inversion
    ws["B125"].value = "Proyectos de inversion"
    ws["B125"].font  = Font(bold=True, color=WHITE, name="Calibri", size=10)
    ws["B125"].fill  = fill_kpi
    ws["B125"].alignment = _al("left")
    ws["B125"].border = _bdr(left=thin, right=thin, top=thin)
    ws["L125"].value = "Total Weekly"
    ws["L125"].font  = Font(bold=True, color=WHITE, name="Calibri", size=10)
    ws["L125"].fill  = fill_kpi
    ws["L125"].alignment = _al("center")
    ws["L125"].border = _bdr(left=thin, right=thin, top=thin, bottom=thin)

    proyectos = [
        (126, "Sistema de riego (Ramona)"),
        (127, "Sistema de riego (Isabella)"),
        (128, "Caseta (Isabella)"),
        (129, "Sistema de ventilacion"),
        (130, "Sistema de tratamiento de aguas residuales (Isabella)"),
        (131, "Arcos para invernaderos "),
        (132, "proyecto luz"),
        (133, "Construccion de Almacen (Ramona) "),
        (134, "Construccion de Almacen (Isabela) "),
        (135, "Carritos"),
        (136, "Maquinaria "),
        (137, "Chiller"),
        (138, "Cuarto frio"),
        (139, "veronicas"),
    ]
    for row, label in proyectos:
        ws[f"B{row}"].value = label
        ws[f"B{row}"].fill  = fill_white
        ws[f"B{row}"].border = _bdr(left=thin, right=thin)
        ws[f"C{row}"].value = 0
        ws[f"C{row}"].font  = Font(color="0000FF", name="Calibri", size=10)
        ws[f"C{row}"].fill  = fill_white
        ws[f"C{row}"].border = _bdr(left=thin, right=thin)
        ws[f"C{row}"].number_format = '"$"#,##0;-"$"#,##0;" $-   "'
        ws[f"J{row}"].border = _bdr(right=thin)
        ws[f"L{row}"].value = 0
        ws[f"L{row}"].font  = Font(color="0000FF", name="Calibri", size=10)
        ws[f"L{row}"].fill  = fill_white
        ws[f"L{row}"].border = _bdr(left=thin, right=thin)
        ws[f"L{row}"].number_format = '" $"#,##0;-" $"#,##0;" $-   "'
        for uc in ("M","N","O","P","Q","R","S"):
            ws[f"{uc}{row}"].value = 0
            ws[f"{uc}{row}"].font  = Font(color="0000FF", name="Calibri", size=10)
            ws[f"{uc}{row}"].fill  = fill_white
            ws[f"{uc}{row}"].border = _bdr(left=thin, right=thin)

    ws["B139"].border = _bdr(left=thin, right=thin, bottom=thin)
    ws["C139"].border = _bdr(left=thin, right=thin, bottom=thin)
    ws["J139"].border = _bdr(right=thin, bottom=thin)
    for uc in ("L","M","N","O","P","Q","R","S"):
        ws[f"{uc}139"].border = _bdr(left=thin, right=thin, bottom=thin)

    ws["B140"].value = "Total "
    ws["B140"].font  = _f(bold=True)
    ws["B140"].fill  = fill_white
    ws["B140"].border = _bdr(left=thin, right=thin, top=thin, bottom=thin)
    ws["C140"].value = 0
    ws["C140"].font  = Font(color="0000FF", name="Calibri", size=10)
    ws["C140"].border = _bdr(top=thin, bottom=thin)
    ws["C140"].number_format = '" $"#,##0;-" $"#,##0;" $-   "'
    ws["L140"].value = 0
    ws["L140"].font  = Font(color="0000FF", name="Calibri", size=10)
    ws["L140"].border = _bdr(left=thin, right=thin, bottom=thin)
    ws["L140"].number_format = '" $"#,##0;-" $"#,##0;" $-   "'

    # Logistica
    ws["B143"].value = "Logistica "
    ws["B143"].font  = Font(bold=True, color=WHITE, name="Calibri", size=10)
    ws["B143"].fill  = fill_kpi
    ws["B143"].alignment = _al("left")
    ws["B143"].border = _bdr(left=thin, top=thin)
    ws["J143"].border = _bdr(right=thin, top=thin)
    ws["L143"].value = "Total Weekly"
    ws["L143"].font  = Font(bold=True, color=WHITE, name="Calibri", size=10)
    ws["L143"].fill  = fill_kpi
    ws["L143"].alignment = _al("center")
    ws["L143"].border = _bdr(left=thin, right=thin, top=thin, bottom=thin)
    ws["N143"].value = "PosCo-RM"
    ws["N143"].font  = Font(bold=True, color=WHITE, name="Calibri", size=10)
    ws["N143"].fill  = fill_kpi
    ws["N143"].alignment = _al("center")
    ws["N143"].border = _bdr(left=thin, right=thin, top=thin, bottom=thin)

    logistica = [
        (144, "Numero de camiones despachados "),
        (145, "Numero de tarimas despachadas (montadas al camion)"),
        (146, "Numero de cajas despachadas"),
        (147, "Numero de Pies cubicos de cajas despachadas "),
        (148, "Numero de Pies cubicos promedio / camion despachado "),
        (149, "Capacidad en pies cubicos por camion "),
        (150, "Rendimiento promedio por camion "),
    ]
    for row, label in logistica:
        ws[f"B{row}"].value = label
        ws[f"B{row}"].fill  = fill_white
        ws[f"B{row}"].border = _bdr(left=thin)
        ws[f"C{row}"].value = 0
        ws[f"C{row}"].font  = Font(color="0000FF", name="Calibri", size=10)
        ws[f"C{row}"].fill  = fill_white
        ws[f"C{row}"].border = _bdr(left=thin, right=thin)
        ws[f"J{row}"].border = _bdr(right=thin)
        ws[f"L{row}"].value = 0
        ws[f"L{row}"].font  = Font(color="0000FF", name="Calibri", size=10)
        ws[f"L{row}"].fill  = fill_white
        ws[f"L{row}"].border = _bdr(left=thin, right=thin)
        ws[f"N{row}"].value = 0
        ws[f"N{row}"].font  = _f(bold=True)
        ws[f"N{row}"].fill  = fill_white
        ws[f"N{row}"].border = _bdr(right=thin)
        ws[f"S{row}"].border = _bdr(right=thin)

    kpi_groups = [
        (152, "Costo incurrido por flete, gtos expo, fitosanitarios"),
        (153, "Costo incurrido en flete, gtos expo, fitosanitarios (USD)"),
        (154, "Numero de Camiones despachados "),
        (156, "Costo incurrido promedio flete, gtos expo, fitosanitarios / pie cubico"),
        (157, "Costo incurrido en flete, gtos expo, fitosanitarios (USD)"),
        (158, "Numero de Pies cubicos de cajas despachadas"),
        (160, "Costo incurrido flete, gtos expo, fitosanitarios / cajas despachadas"),
        (161, "Costo incurrido en flete, gtos expo, fitosanitarios (USD)"),
        (162, "Numero de cajas despachadas"),
    ]
    for row, label in kpi_groups:
        ws[f"B{row}"].value = label
        ws[f"B{row}"].fill  = fill_white
        ws[f"B{row}"].border = _bdr(left=thin)
        ws[f"C{row}"].value = 0
        ws[f"C{row}"].font  = Font(color="0000FF", name="Calibri", size=10)
        ws[f"C{row}"].fill  = fill_white
        ws[f"C{row}"].border = _bdr(left=thin, right=thin)
        ws[f"L{row}"].value = 0
        ws[f"L{row}"].font  = Font(color="0000FF", name="Calibri", size=10)
        ws[f"L{row}"].fill  = fill_white
        ws[f"L{row}"].border = _bdr(left=thin, right=thin)
        ws[f"N{row}"].value = 0
        ws[f"N{row}"].border = _bdr(right=thin)

    ws["B165"].value = "Material de empaque / Caja"
    ws["B165"].font  = Font(bold=True, color=WHITE, name="Calibri", size=10)
    ws["B165"].fill  = fill_kpi
    ws["B165"].alignment = _al("left")
    ws["B165"].border = _bdr(left=thin, top=thin)

    me_rows = [
        (166, "Costo incurrido en Material de empaque / pie cubico"),
        (167, "Costo incurrido en Material de empaque (USD)"),
        (168, "Numero de Pies cubicos de cajas despachadas"),
        (170, "Costo incurrido en Material de empaque / cajas despachadas"),
        (171, "Costo incurrido en Material de empaque (USD)"),
        (172, "Numero de cajas despachadas"),
    ]
    for row, label in me_rows:
        ws[f"B{row}"].value = label
        ws[f"B{row}"].fill  = fill_white
        ws[f"B{row}"].border = _bdr(left=thin)
        ws[f"C{row}"].value = 0
        ws[f"C{row}"].font  = Font(color="0000FF", name="Calibri", size=10)
        ws[f"C{row}"].fill  = fill_white
        ws[f"C{row}"].border = _bdr(left=thin, right=thin)
        ws[f"L{row}"].value = 0
        ws[f"L{row}"].font  = Font(color="0000FF", name="Calibri", size=10)
        ws[f"L{row}"].fill  = fill_white
        ws[f"L{row}"].border = _bdr(left=thin, right=thin)
        ws[f"N{row}"].value = 0
        ws[f"N{row}"].border = _bdr(right=thin)

    # Merged cells
    merges = [
        "C5:J5", "L5:R5",
        "C153:C154", "L153:L154",
        "C157:C158", "L157:L158",
        "C161:C162", "L161:L162",
        "C167:C168", "L167:L168",
        "C171:C172", "L171:L172",
    ]
    for m in merges:
        try:
            ws.merge_cells(m)
        except Exception:
            pass

# --- Crear nueva hoja WK en SharePoint via Microsoft Graph API (con sesión) ---
def crear_hoja_wk(nombre_hoja: str, tenant_id: str, client_id: str, client_secret: str) -> dict:
    """
    Crea una nueva hoja WK#### desde cero usando una sesión de workbook de Graph API.
    Funciona aunque el archivo esté abierto por otros usuarios (no requiere lock).
    Escribe todas las celdas directamente via API, sin subir el archivo completo.
    Requiere Files.ReadWrite en la App Registration de Azure AD.
    """
    import base64 as _b64
    import time

    #  Helper: construir lista plana de celdas { address, value } 
    def _celdas_de_la_hoja(nombre):
        """Devuelve lista de dicts con address y valor para escribir via Graph API."""
        celdas = []

        def c(addr, val):
            celdas.append({"address": addr, "value": val})

        # Encabezados
        _code = nombre[2:] if str(nombre).upper().startswith("WK") else str(nombre)
        try:
            _tc = 20 if 2502 <= int(_code) <= 2520 else 19
        except (ValueError, TypeError):
            _tc = 19
        c("B1", "CENTRO FLORICULTOR DE BAJA CALIFORNIA, S.A. DE C.V. ")
        c("B2", "SEMANA DE CALCULO - Mexico")
        c("B3", nombre)
        c("C3", _tc); c("D3", " tipo de cambio")
        c("L3", _tc); c("M3", "  tipo de cambio ")
        c("B4", "Del ___ al ___ de ________ 20__")
        c("C5", "(MXN) Pesos Mexicanos")
        c("L5", "US Dollars")
        c("B6", "TOTAL FINCA")
        # Fila 7: ranchos MXN
        for col, h in zip(["C","D","E","F","G","H","I","J"],
                          ["TOTAL","Prop-RM","PosCo-RM","Campo -RM","Isabela","Christina","Cecilia","Cecilia 25"]):
            c(f"{col}7", h)
        # Fila 7: ranchos USD
        for col, h in zip(["L","M","N","O","P","Q","R","S"],
                          ["TOTAL","Prop-RM","PosCo-RM","Campo -RM","ISABELA","Christina","CECILIA","CECILIA 25"]):
            c(f"{col}7", h)
        c("B7", "Produccion"); c("C8", "SEMANAL "); c("L8", '"WEEKLY"')
        c("B9", "EJECUCION SEMANAL")
        for col in ["D","E","F","G","H","I","J"]:
            c(f"{col}9", 1)

        # Categorías de materiales (filas 10-20)
        categorias = [
            (10, "DESINFECCION Y FERTILIZACION"),
            (11, "AMPLIACION "),
            (12, "CULTIVO TIERRA, CHAROLAS"),
            (13, "MATERIAL VEGETAL"),
            (14, "PREPARACION DE SUELO"),
            (15, "FERTILIZANTES (Manejo Integrado de Riego y Fertilización) "),
            (16, "DESINFECCION / PLAGUICIDAS (Manejo Integrado de Plagas y Enfermedades)"),
            (17, "MANTENIMIENTO"),
            (18, "EXPANSION CECILIA 25"),
            (19, "RENOVACION DE SIEMBRA"),
            (20, "MATERIAL DE EMPAQUE"),
        ]
        for row, label in categorias:
            c(f"B{row}", label)

        c("B22", "COSTO DE MATERIALES")

        # Nóminas (filas 24-59)
        nominas = [
            (24, "NOMINA ADMON Oficina, Jefes de Finca, Ingenieros"),
            (25, "HORAS EXTR. DOM. Y FESTIVOS"),
            (26, "BONOS ASISIT, PUNTAULIDAD Y DESPENSA"),
            (27, "NOMINA PRODUCCION "),
            (28, "HORAS EXTR. DOM. Y FEST."),
            (29, "BONOS ASISIT, PUNT. Y DESP."),
            (30, "NOMINA PRODUCCION CORTE"),
            (31, "HORAS EXTR. DOM. Y FESTIVOS CORTE"),
            (32, "BONOS ASISIT, PUNTAULIDAD Y DESP. CORTE"),
            (33, "NOMINA PRODUCCION TRANSPLANTE"),
            (34, "HORAS EXTR. DOM. Y FEST. TRANSPLANTE"),
            (35, "BONOS ASISIT, PUNT. Y DESP. TRANSPLANTE"),
            (36, "NOMINA PRODUCCION MANEJO PLANTA"),
            (37, "HORAS EXTR. DOM. Y FEST. MANEJO PLANTA"),
            (38, "BONOS ASISIT, PUNT. Y DESP. MANEJO PLANTA"),
            (39, "NOMINA  HOOPS"),
            (40, "HORAS EXTR. DOM. Y FEST. HOOPS"),
            (41, "BONOS ASISIT, PUNT. Y DESP.HOOPS"),
            (42, "NOMINA  (MIPE,MIRFE,)"),
            (43, "HORAS EXTR. DOM. Y FEST. (MIPE,MIRFE)"),
            (44, "BONOS ASISIT, PUNT. Y DESP.(MIPE,MIRFE)"),
            (45, "NOMINA OPERATIVOS (TRACTORES, CAMEROS)"),
            (46, "HORAS EXTR. DOM. Y FEST. (TRACTORES, CAMEROS)"),
            (47, "BONOS ASISIT, PUNT. Y DESP. (TRACTORES, CAMEROS)"),
            (48, "NOMINA OPERATIVOS (CHOFER)"),
            (49, "HORAS EXTR. DOM. Y FEST. (CHOFER)"),
            (50, "BONOS ASISIT, PUNT. Y DESP. (CHOFER)"),
            (51, "NOMINA OPERATIVOS (VELADORES)"),
            (52, "HORAS EXTR. DOM. Y FEST. (VELADORES)"),
            (53, "BONOS ASISIT, PUNT. Y DESP. (VELADORES)"),
            (54, "NOMINA OPERATIVOS (SOLDADOR)"),
            (55, "HORAS EXTR. DOM. Y FEST. (SOLDADOR)"),
            (56, "BONOS ASISIT, PUNT. Y DESP. (SOLDADOR)"),
            (57, "NOMINA PRODUCCION Contratista y comisiones"),
            (58, "IMSS , INFONAVIT RCV"),
            (59, "1.8% al estado (1.2% tasa efectiva)"),
        ]
        for row, label in nominas:
            c(f"B{row}", label)

        c("B61", "COSTO DE MANO DE OBRA")

        # Servicios (filas 63-70)
        servicios = [
            (63, "ELECTRICIDAD"),
            (64, "FLETES Y ACARREOS (Flete aduana)"),
            (65, "GASTOS DE EXPORTACION "),
            (66, "CERTIFICADO DE FITOSANITARIOS"),
            (67, "Transporte de personal"),
            (68, "COMPRA DE FLOR A TERCEROS"),
            (69, "COMIDA PARA EL PERSONAL"),
            (70, "RO, TEL, RTA.ALIM."),
        ]
        for row, label in servicios:
            c(f"B{row}", label)

        c("B72", "COSTO DE SERVICIOS")
        c("B74", "COSTO DE PRODUCCION Y VENTAS")

        # Producción (filas 76-92)
        prod = [
            (76, "CAJAS PROCESADAS TOTALES"),
            (77, "INVENTARIO INICIAL"),
            (78, "TALLOS COSECHADOS"),
            (79, "TALLOS DESECHADOS"),
            (80, "TALLOS DESECHADOS sf"),
            (81, "TALLOS COMPRADOS"),
            (82, "TALLOS EN BOUQUETS O PROCESADOS"),
            (83, "TALLOS DESPACHADOS"),
            (84, "LIBRAS DESPACHADAS ALBAHACA"),
            (85, "TALLOS muestra"),
            (86, "INVENTARIO FINAL"),
            (87, "TALLOS PROCESADOS TOTALES"),
            (88, " CHAROLAS SEMBRADAS *288 PLUGS ="),
            (89, " NUMERO DE CHAROLAS SEMBRADAS "),
            (90, " NUMERO DE ESQUEJES SEMBRADOS"),
            (91, " METROS DE SIEMBRA"),
            (92, " HECTAREAS EN SIEMBRA"),
        ]
        for row, label in prod:
            c(f"B{row}", label)

        c("B93", "<<< INDICADORES")
        c("B94", "COSTOS UNITARIOS"); c("B95", "$ / Tallo Procesado")
        c("B96", "COSTOS UNITARIOS"); c("B97", "$ / Libras Procesadas")
        for row, label in [(98,"Materiales"),(99,"Mano de Obra"),(100,"Servicios (Fletes)"),
                           (101,"Costo de Produccion y Ventas"),(103,"Material de Empaque / Tallo"),
                           (105,"Sanidad Vegetal / Tallo"),(106,"Fertlizacion / Tallo"),
                           (108,"Mano de Obra Prod / Tallo")]:
            c(f"B{row}", label)
        c("B110", "$ / Hectarea")
        for row, label in [(111,"Materiales"),(112,"Mano de Obra"),(113,"Servicios (Fletes)"),
                           (114,"Costo de Produccion y Ventas"),(121,"Mano de Obra Prod / Ha")]:
            c(f"B{row}", label)

        c("B124", "KPI's ")
        c("B125", "Proyectos de inversión"); c("L125", "Total Weekly")
        proyectos = [
            (126,"Sistema de riego (Ramona)"),(127,"Sistema de riego (Isabella)"),
            (128,"Caseta (Isabella)"),(129,"Sistema de ventilacion)"),
            (130,"Sistema de tratamiento de aguas residuales (Isabella)"),
            (131,"Arcos para invernaderos "),(132,"proyecto luz"),
            (133,"Construcción de Almacén (Ramona) "),(134,"Construcción de Almacén (Isabela) "),
            (135,"Carritos"),(136,"Maquinaria "),(137,"Chiller"),
            (138,"Cuarto frio"),(139,"veronicas"),
        ]
        for row, label in proyectos:
            c(f"B{row}", label)
        c("B140", "Total ")
        c("B143", "Logística "); c("L143", "Total Weekly"); c("N143", "PosCo-RM")
        for row, label in [
            (144,"Número de camiones despachados "),(145,"Número de tarimas despachadas (montadas al camión)"),
            (146,"Número de cajas despachadas"),(147,"Número de Pies cúbicos de cajas despachadas "),
            (148,"Número de Pies cubicos promedio / camión despachado "),
            (149,"Capacidad en pies cúbicos por camión "),(150,"Rendimiento promedio por camión "),
        ]:
            c(f"B{row}", label)
        for row, label in [
            (152,"Costo incurrido por flete, gtos expo, fitosanitarios"),
            (153,"Costo incurrido en flete, gtos expo, fitosanitarios (USD)"),
            (154,"Número de Camiones despachados "),
            (156,"Costo incurrido promedio flete, gtos expo, fitosanitarios / pie cúbico"),
            (157,"Costo incurrido en flete, gtos expo, fitosanitarios (USD)"),
            (158,"Número de Pies cúbicos de cajas despachadas"),
            (160,"Costo incurrido flete, gtos expo, fitosanitarios / cajas despachadas"),
            (161,"Costo incurrido en flete, gtos expo, fitosanitarios (USD)"),
            (162,"Número de cajas despachadas"),
        ]:
            c(f"B{row}", label)
        c("B165", "Material de empaque / Caja")
        for row, label in [
            (166,"Costo incurrido en Material de empaque / pie cúbico"),
            (167,"Costo incurrido en Material de empaque (USD)"),
            (168,"Número de Pies cúbicos de cajas despachadas"),
            (170,"Costo incurrido en Material de empaque / cajas despachadas"),
            (171,"Costo incurrido en Material de empaque (USD)"),
            (172,"Número de cajas despachadas"),
        ]:
            c(f"B{row}", label)

        return celdas

    #  1. Token OAuth2 
    token_url = f'https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token'
    token_resp = requests.post(token_url, data={
        "grant_type":    "client_credentials",
        "client_id":     client_id,
        "client_secret": client_secret,
        "scope":         "https://graph.microsoft.com/.default",
    }, timeout=20)
    if token_resp.status_code != 200:
        return {"ok": False, "error": f"Error obteniendo token: {token_resp.text}"}

    token = token_resp.json().get('access_token')
    hdrs_json = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    #  2. Resolver driveId + itemId 
    encoded = _b64.b64encode(SHAREPOINT_URL_WK.encode()).decode().rstrip('=')
    encoded = 'u!' + encoded.replace('/', '_').replace('+', '-')
    res = requests.get(
        f'https://graph.microsoft.com/v1.0/shares/{encoded}/driveItem',
        headers=hdrs_json, timeout=20,
    )
    if res.status_code != 200:
        return {"ok": False, "error": f"No se pudo resolver el archivo: {res.text}"}

    item     = res.json()
    drive_id = item.get('parentReference', {}).get('driveId')
    item_id  = item.get('id')
    if not drive_id or not item_id:
        return {"ok": False, "error": "No se pudo obtener driveId o itemId."}

    wb_url = f'https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/workbook'

    #  3. Abrir sesión de workbook (permite trabajar con archivo abierto) 
    sess_resp = requests.post(
        f'{wb_url}/createSession',
        headers=hdrs_json,
        json={"persistChanges": True},
        timeout=30,
    )
    if sess_resp.status_code not in (200, 201):
        return {"ok": False, "error": f"Error abriendo sesión: {sess_resp.text}"}

    session_id = sess_resp.json().get('id')
    hdrs = {**hdrs_json, "workbook-session-id": session_id}

    try:
        #  4. Verificar que la hoja no exista 
        sheets_resp = requests.get(f'{wb_url}/worksheets', headers=hdrs, timeout=20)
        if sheets_resp.status_code != 200:
            return {"ok": False, "error": f"Error listando hojas: {sheets_resp.text}"}
        nombres = [h['name'].strip() for h in sheets_resp.json().get('value', [])]
        if nombre_hoja.upper() in [n.upper() for n in nombres]:
            return {"ok": False, "error": f"La hoja '{nombre_hoja}' ya existe."}

        #  5. Crear la hoja nueva 
        add_resp = requests.post(
            f'{wb_url}/worksheets/add',
            headers=hdrs,
            json={"name": nombre_hoja},
            timeout=20,
        )
        if add_resp.status_code not in (200, 201):
            return {"ok": False, "error": f"Error creando hoja: {add_resp.text}"}

        ws_id = add_resp.json().get('id', nombre_hoja)
        #  6. Mover al inicio (posición 0) 
        requests.patch(
            f'{wb_url}/worksheets/{nombre_hoja}',
            headers=hdrs,
            json={"position": 0},
            timeout=20,
        )

        #  7. Copiado masivo de Fórmulas y Formatos de Número (A1:Z1500) 
        import re
        prev_wk_name = None
        m = re.search(r'\d+', nombre_hoja)
        if m:
            num = int(m.group())
            prev_wk_name = nombre_hoja.replace(str(num), str(num - 1))

        copied_data = None
        if prev_wk_name and prev_wk_name.upper() in [n.upper() for n in nombres]:
            # Traer fórmulas y formatos de número explícitamente de A1 hasta Z1500
            try:
                get_resp = requests.get(
                    f"{wb_url}/worksheets/{prev_wk_name}/range(address='A1:Z1500')?$select=formulas,numberFormat",
                    headers=hdrs, timeout=60
                )
                if get_resp.status_code == 200:
                    copied_data = get_resp.json()
            except:
                pass

        if copied_data and "formulas" in copied_data:
            # Pegar el bloque de 1500 filas directo
            patch_resp = requests.patch(
                f"{wb_url}/worksheets/{nombre_hoja}/range(address='A1:Z1500')",
                headers=hdrs, 
                json={
                    "formulas": copied_data["formulas"],
                    "numberFormat": copied_data.get("numberFormat", [])
                }, 
                timeout=120
            )
            # Actualizar nombre de la semana en la celda B3 independientemente
            requests.patch(
                f"{wb_url}/worksheets/{nombre_hoja}/range(address='B3')",
                headers=hdrs, json={"values": [[nombre_hoja]]}, timeout=20
            )
            
            # NOTE: Dejamos que el código continúe hacia el Paso 8 para "pintar" la plantilla (colores/bordes), 
            # ya que leer puras fórmulas no copia los colores de fondo nativamente.

        else:
            #  Alternativa: Escribir celdas desde cero (batchUpdate vía range) 
            NROWS, NCOLS = 250, 21  # cols A(0)..S(18)
            col_idx = {c: i for i, c in enumerate("ABCDEFGHIJKLMNOPQRSTU")}
            matrix = [[""] * NCOLS for _ in range(NROWS)]
    
            for cell in _celdas_de_la_hoja(nombre_hoja):
                addr = cell["address"]
                val  = cell["value"]
                col_str = ''.join(ch for ch in addr if ch.isalpha())
                row_str = ''.join(ch for ch in addr if ch.isdigit())
                if col_str in col_idx and row_str:
                    r = int(row_str) - 1
                    col_c = col_idx[col_str]
                    if 0 <= r < NROWS and 0 <= col_c < NCOLS:
                        matrix[r][col_c] = val if val is not None else ""
    
            range_addr = f"A1:S{NROWS}"
            patch_resp = requests.patch(
                f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{range_addr}\')',
                headers=hdrs,
                json={"values": matrix},
                timeout=60,
            )
            if patch_resp.status_code not in (200, 201):
                return {"ok": False, "error": f"Error escribiendo celdas: {patch_resp.text}"}

        #  8. Aplicar formatos via Graph API 
        def fmt(rng, body):
            requests.patch(
                f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{rng}\')/format',
                headers=hdrs, json=body, timeout=30,
            )

        def fill(rng, color):
            requests.patch(
                f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{rng}\')/format/fill',
                headers=hdrs, json={"color": f"#{color}"}, timeout=30,
            )

        def font(rng, bold=False, color=None, size=None):
            body = {"bold": bold, "name": "Arial"}
            if color: body["color"] = f"#{color}"
            if size:  body["size"]  = size
            requests.patch(
                f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{rng}\')/format/font',
                headers=hdrs, json=body, timeout=30,
            )

        def border(rng, left=None, right=None, top=None, bottom=None, inner_h=None, inner_v=None):
            style_map = {"thin": "Continuous", "medium": "Medium"}
            base = f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{rng}\')/format/borders'
            for side_name, style in [
                ("EdgeLeft",left), ("EdgeRight",right), ("EdgeTop",top), ("EdgeBottom",bottom),
                ("InsideHorizontal",inner_h), ("InsideVertical",inner_v)
            ]:
                if style:
                    requests.patch(
                        f'{base}/{side_name}',
                        headers=hdrs,
                        json={"style": style_map.get(style, style), "color": "#000000"},
                        timeout=30,
                    )

        #  Colores de fondo 
        fill("B2",        "DAE3F3")   # azul claro encabezado semana
        fill("B3",        "C5E0B4")   # verde lima código semana

        # Verde claro (CCFFCC)  columnas USD L:S
        for rng in [
            "L5:U9",
            "L10:U21", "L22:U23", "L24:U60",
            "L61:U62", "L63:U71", "L72:U74",
            "L94:U100", "L102:U102", "L104:U104", "L107:U107",
        ]:
            fill(rng, "CCFFCC")

        # Naranja (FFCC99)  subtotales
        fill("C22:J22",   "FFCC99")
        fill("L22:U22",   "FFCC99")
        fill("C61:J61",   "FFCC99")
        fill("L61:U61",   "FFCC99")
        fill("C72:J72",   "FFCC99")
        fill("L72:U72",   "FFCC99")
        fill("B101:J101", "FFCC99")
        fill("L101:U101", "FFCC99")

        # Amarillo claro (FFFFCC)  producción y $/Ha
        fill("L76:U92",   "FFFFCC")
        fill("L110:U121", "FFFFCC")
        # Amarillo vivo (FFFF00)  charolas/esquejes MXN
        fill("D89:J91",   "FFFF00")

        # Verde oscuro (008000)  headers KPI
        for rng in ["B125", "L125", "B143", "L143", "N143", "B165"]:
            fill(rng, "008000")

        # Blanco explícito  sección KPI proyectos / logística (filas 126-250)
        fill("B126:U250", "FFFFFF")

        #  Color de texto navy (#333399) en todo el cuerpo + tamaño 10 
        font("B1:U250",  bold=False, color="333399", size=10)

        #  Negritas 
        font("B1:B3",    bold=True,  color="333399", size=10)
        font("B9",       bold=True,  color="333399", size=10)
        font("B22",      bold=True,  color="333399", size=10)
        font("B61",      bold=True,  color="333399", size=10)
        font("B72",      bold=True,  color="333399", size=10)
        font("B74",      bold=True,  color="333399", size=10)
        font("B93:B97",  bold=True,  color="333399", size=10)
        font("B101",     bold=True,  color="333399", size=10)
        font("B103",     bold=True,  color="333399", size=10)
        font("B105",     bold=True,  color="333399", size=10)
        font("B106",     bold=True,  color="333399", size=10)
        font("B108",     bold=True,  color="333399", size=10)
        font("B110",     bold=True,  color="333399", size=10)
        font("B116",     bold=True,  color="333399", size=10)
        font("B118",     bold=True,  color="333399", size=10)
        font("B119",     bold=True,  color="333399", size=10)
        font("B121",     bold=True,  color="333399", size=10)
        font("B124",     bold=True,  color="333399", size=10)
        font("B140",     bold=True,  color="333399", size=10)
        # Columna C subtotales / columna L subtotales USD
        for rng in ["C22:J22", "C61:J61", "C72:J72", "C74:J74",
                    "L22:U22", "L61:U61", "L72:U72", "L74:U74",
                    "L76:L92", "L95:L121",
                    "L101:U101", "L103:U103", "L105:U106", "L108:U108",
                    "L111:N114", "L116:N119", "L121:U121"]:
            font(rng, bold=True, size=10)
        # Columna C negrita en todas las filas de datos
        for rng in ["C10:C21", "C24:C60", "C63:C70",
                    "C76:C92", "C95:C121"]:
            font(rng, bold=True, color="333399", size=10)
        # KPI headers  texto blanco negrita
        for rng in ["B125", "L125", "B143", "L143", "N143", "B165"]:
            font(rng, bold=True, color="FFFFFF", size=10)
        # Texto azul en valores KPI proyectos/logística
        for rng in ["C126:C250", "L126:L250"]:
            font(rng, bold=False, color="0000FF", size=10)

        #  Bordes  estrategia simplificada (pocas llamadas) 
        
        # GRILLAS INTERNAS (Lo que faltaba para no dejarlo a medias)
        border("B5:U122", inner_h="thin", inner_v="thin")
        border("B125:U250", inner_h="thin", inner_v="thin")

        # ESTRUCTURA PRINCIPAL: 3 columnas clave con rangos grandes
        # Left medio en B (toda el área de datos)
        border("B2:B250",  left="medium")
        # Right medio en J (toda el área de datos)
        border("J2:J250",  right="medium")
        # Right thin en C (separador columna TOTAL)
        border("C5:C250",  right="thin")
        # Left medio en L + right medio en S (todo el bloque USD)
        border("L5:L250",  left="medium")
        border("U5:U250",  right="medium")
        # Left medio en L para separar C de la zona MXN izquierda también
        border("C5:C9",    left="medium")
        border("C10:C21",  left="medium")
        border("C22:C74",  left="medium")
        border("C76:C121", left="medium")

        # FILAS ESPECIALES  separadores horizontales MXN
        border("B5:J5",    top="medium", bottom="thin")
        border("B9:J9",    bottom="thin")
        border("B22:J22",  top="thin",   bottom="thin")
        border("B61:J61",  top="thin",   bottom="thin")
        border("B72:J72",  top="thin",   bottom="thin")
        border("B74:J74",  top="thin",   bottom="medium")
        border("B76:J76",  top="medium")
        border("B92:J92",  bottom="medium")
        border("B94:J94",  top="medium")
        border("B108:J108",bottom="medium")
        border("B110:J110",top="medium")
        border("B121:J121",bottom="medium")

        # FILAS ESPECIALES  separadores horizontales USD
        border("L5:U5",    top="medium", bottom="thin")
        border("L9:U9",    bottom="thin")
        border("L10:U10",  top="thin")
        border("L22:U22",  top="thin",   bottom="thin")
        border("L61:U61",  top="thin",   bottom="thin")
        border("L72:U72",  top="thin",   bottom="thin")
        border("L74:U74",  top="thin",   bottom="medium")
        border("L76:U76",  top="medium")
        border("L92:U92",  bottom="medium")
        border("L94:U94",  top="medium")
        border("L97:U97",  bottom="thin")
        border("L98:U98",  top="thin")
        border("L100:U100",bottom="thin")
        border("L101:U101",bottom="thin")
        border("L103:U103",top="thin",   bottom="thin")
        border("L105:U105",top="thin")
        border("L106:U106",bottom="thin")
        border("L108:U108",top="thin",   bottom="medium")
        border("L110:U110",top="medium")
        border("L113:U113",bottom="thin")
        border("L114:U114",top="thin",   bottom="thin")
        border("L116:U116",top="thin",   bottom="thin")
        border("L118:U118",top="thin")
        border("L119:U119",bottom="thin")
        border("L121:U121",top="thin",   bottom="medium")

        # KPI HEADERS borders
        border("B125",  left="thin", right="thin", top="thin")
        border("L125",  left="thin", right="thin", top="thin", bottom="thin")
        border("B143",  left="thin", right="thin", top="thin")
        border("J143",  right="thin", top="thin")
        border("L143",  left="thin", right="thin", top="thin", bottom="thin")
        border("N143",  left="thin", right="thin", top="thin", bottom="thin")
        border("B165",  left="thin", right="thin", top="thin")

        # PROYECTOS  outline
        border("B126:J139", left="thin",   right="thin")
        border("B139:J139", bottom="thin")
        border("L126:U139", left="thin",   right="thin")
        border("L139:U139", bottom="thin")
        border("B140",  left="thin", right="thin", top="thin", bottom="thin")
        border("L140",  left="thin", right="thin", bottom="thin")

        # LOGÍSTICA  outline extendido
        border("B144:B250", left="thin")
        border("J144:J250", right="thin")
        border("B250:J250", bottom="thin")
        
        border("L144:L250", left="thin")
        border("N144:N250", right="thin")
        border("U144:U250", right="thin")
        border("L250:U250", bottom="thin")

        #  Alineación 
        fmt("B2",    {"horizontalAlignment": "Center"})
        fmt("B3",    {"horizontalAlignment": "Center"})
        fmt("B4",    {"horizontalAlignment": "Center"})
        fmt("C5:J5", {"horizontalAlignment": "Center"})
        fmt("L5:U5", {"horizontalAlignment": "Center", "verticalAlignment": "Center"})
        fmt("B6",    {"horizontalAlignment": "Center", "verticalAlignment": "Top", "wrapText": True})
        fmt("B7",    {"horizontalAlignment": "Center", "verticalAlignment": "Top", "wrapText": True})
        fmt("C7:J7", {"horizontalAlignment": "Center", "verticalAlignment": "Top"})
        fmt("L7:U7", {"horizontalAlignment": "Center", "verticalAlignment": "Top"})
        fmt("C8",    {"horizontalAlignment": "Center"})
        fmt("L8",    {"horizontalAlignment": "Center"})
        fmt("B9",    {"horizontalAlignment": "Center"})
        fmt("L125",  {"horizontalAlignment": "Center"})
        fmt("L143",  {"horizontalAlignment": "Center"})
        fmt("N143",  {"horizontalAlignment": "Center"})

        #  Anchos de columnas 
        # Configurar anchos de columna para que coincidan con el formato esperado
        # Graph API requiere usar el endpoint de columnas específico
        column_widths = {
            "A": 3,
            "B": 69.4,
            "C": 14,
            "D": 11, "E": 11, "F": 11, "G": 11, "H": 11, "I": 11, "J": 11,
            "K": 11,
            "L": 11, "M": 11, "N": 11, "O": 11, "P": 11, "Q": 11, "R": 11, "S": 11, "T": 11, "U": 11,
        }
        for col_letter, width in column_widths.items():
            try:
                # Usamos el endpoint de formato de rango para establecer el ancho
                requests.patch(
                    f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{col_letter}:{col_letter}\')/format',
                    headers=hdrs,
                    json={"columnWidth": width * 7.5},
                    timeout=20,
                )
            except Exception as e:
                # No es crítico si falla el ajuste de ancho
                print(f"[WARN]  Error configurando ancho columna {col_letter}: {e}")

        #  Alto de filas 
        # Configurar alto de filas específicas
        row_heights = {
            3: 15.0,
            4: 15.0,
            6: 26.4,
        }
        for row_num, height in row_heights.items():
            try:
                requests.patch(
                    f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{row_num}:{row_num}\')/format',
                    headers=hdrs,
                    json={"rowHeight": height},
                    timeout=20,
                )
            except Exception as e:
                # No es crítico si falla el ajuste de alto
                print(f"[WARN]  Error configurando alto fila {row_num}: {e}")

        #  Formato de número (#,##0) para celdas de valores 
        # Aplicar formato de número con separador de miles a las celdas con valores
        number_ranges = [
            # Subtotales MXN
            "C22:J22", "C61:J61", "C72:J72", "C74:J74",
            # Subtotales USD
            "L22:U22", "L61:U61", "L72:U72", "L74:U74",
            # Valores de datos MXN
            "C10:J21", "C24:J60", "C63:J70",
            # Valores de datos USD
            "L10:U21", "L24:U60", "L63:U70",
            # Sección de producción y costos
            "C76:J92", "L76:U92",
            "C95:J121", "L95:U121",
        ]
        for rng in number_ranges:
            try:
                requests.patch(
                    f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{rng}\')/format',
                    headers=hdrs,
                    json={"numberFormat": "#,##0"},
                    timeout=20,
                )
            except Exception as e:
                print(f"[WARN]  Error configurando formato número en {rng}: {e}")

        #  Merge de celdas 
        merges = [
            # Headers principales
            "C5:J5", "L5:R5",
            # Headers columnas K-L
            "K1:L1", "K2:L2", "K4:L4",
            # Separadores K-L
            "K75:L75", "K93:L93", "K109:L109",
            # KPI headers K-L
            "K122:L122", "K123:L123", "K124:L124",
            "K141:L141", "K142:L142",
            "K164:L164",
            "K174:L174", "K175:L175",
            # KPI sections A-B
            "A123:B123",
            "A141:B141", "A142:B142",
            "A164:B164",
            "A174:B174", "A175:B175",
            # Valores combinados verticalmente (logística)
            "C153:C154", "L153:L154",
            "C157:C158", "L157:L158",
            "C161:C162", "L161:L162",
            "C167:C168", "L167:L168",
            "C171:C172", "L171:L172",
        ]
        for m in merges:
            try:
                requests.post(
                    f'{wb_url}/worksheets/{nombre_hoja}/range(address=\'{m}\')/merge',
                    headers=hdrs, json={"across": False}, timeout=20,
                )
            except Exception as e:
                print(f"[WARN]  Error merge {m}: {e}")

    finally:
        #  8. Cerrar la sesión siempre 
        requests.post(
            f'{wb_url}/closeSession',
            headers=hdrs,
            timeout=20,
        )

    return {
        "ok": True,
        "mensaje": f"Hoja '{nombre_hoja}' creada exitosamente en SharePoint.",
    }


def autorrellenar_materiales_wk(week_code: str, tenant_id: str, client_id: str, client_secret: str) -> dict:
    """
    Autorrellena las filas de materiales en MN por rancho para una WK existente.

    Categorías soportadas:
    - FERTILIZANTES                  <- PR#### tipo MIRFE
    - DESINFECCION / PLAGUICIDAS     <- PR#### tipo MIPE
    - MANTENIMIENTO                  <- MP#### total por rancho
    - MATERIAL DE EMPAQUE            <- ME#### total por rancho
    """
    import base64 as _b64

    code = _normalizar_week_code(week_code)
    if not (code.isdigit() and len(code) == 4):
        return {"ok": False, "error": "El código de semana debe ser exactamente 4 dígitos (ej: 2614)."}

    nombre_hoja = f"WK{code}"

    archivo = _descargar_con_graph(SHAREPOINT_URL_PR, "Excel PR/MP/ME")
    if archivo is None:
        return {"ok": False, "error": "No se pudo descargar el Excel de PR/MP/ME desde SharePoint."}

    try:
        xls = pd.ExcelFile(archivo)
    except Exception as e:
        return {"ok": False, "error": f"No se pudo abrir el Excel de PR/MP/ME: {e}"}

    resumen = {}
    fuentes_ok = []
    fuentes_faltantes = []
    omitidos = {}

    for categoria, cfg in WK_MATERIAL_AUTOFILL.items():
        prefix = cfg["prefix"]
        tipo = cfg["tipo"]
        sheet_name = _buscar_hoja_por_prefijo(xls.sheet_names, prefix, code)
        if not sheet_name:
            resumen[categoria] = {"row": cfg["row"], "totales": {rn: 0.0 for rn in WK_MXN_RANCH_COLS}}
            missing_name = f"{prefix}{code}"
            if missing_name not in fuentes_faltantes:
                fuentes_faltantes.append(missing_name)
            continue

        vals = _leer_hoja(xls, sheet_name, rango_filas=500, rango_cols=11)
        parsed = _parse_generic(vals)
        totales, omitidos_cat = _sumar_gasto_por_rancho(parsed, tipo=tipo)
        resumen[categoria] = {"row": cfg["row"], "totales": totales}
        if sheet_name not in fuentes_ok:
            fuentes_ok.append(sheet_name)
        if omitidos_cat:
            omitidos[categoria] = omitidos_cat

    if not fuentes_ok:
        return {
            "ok": False,
            "error": f"No se encontraron las hojas PR{code}, MP{code} ni ME{code} en el Excel fuente.",
        }

    token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    token_resp = requests.post(token_url, data={
        "grant_type": "client_credentials",
        "client_id": client_id,
        "client_secret": client_secret,
        "scope": "https://graph.microsoft.com/.default",
    }, timeout=20)
    if token_resp.status_code != 200:
        return {"ok": False, "error": f"Error obteniendo token: {token_resp.text[:300]}"}

    token = token_resp.json().get("access_token")
    hdrs_json = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    encoded = _b64.b64encode(SHAREPOINT_URL_WK.encode()).decode().rstrip("=")
    encoded = "u!" + encoded.replace("/", "_").replace("+", "-")
    res = requests.get(
        f"https://graph.microsoft.com/v1.0/shares/{encoded}/driveItem",
        headers=hdrs_json,
        timeout=20,
    )
    if res.status_code != 200:
        return {"ok": False, "error": f"No se pudo resolver el archivo WK: {res.text[:300]}"}

    item = res.json()
    drive_id = item.get("parentReference", {}).get("driveId")
    item_id = item.get("id")
    if not drive_id or not item_id:
        return {"ok": False, "error": "No se pudo obtener driveId o itemId del Excel WK."}

    wb_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/workbook"
    sess_resp = requests.post(
        f"{wb_url}/createSession",
        headers=hdrs_json,
        json={"persistChanges": True},
        timeout=30,
    )
    if sess_resp.status_code not in (200, 201):
        return {"ok": False, "error": f"Error abriendo sesión del workbook WK: {sess_resp.text[:300]}"}

    session_id = sess_resp.json().get("id")
    hdrs = {**hdrs_json, "workbook-session-id": session_id}

    try:
        sheets_resp = requests.get(f"{wb_url}/worksheets", headers=hdrs, timeout=20)
        if sheets_resp.status_code != 200:
            return {"ok": False, "error": f"Error listando hojas WK: {sheets_resp.text[:300]}"}

        target_sheet = None
        normalized_target = nombre_hoja.replace(" ", "").upper()
        for ws in sheets_resp.json().get("value", []):
            ws_name = str(ws.get("name", "")).strip()
            if ws_name.replace(" ", "").upper() == normalized_target:
                target_sheet = ws_name
                break

        if not target_sheet:
            return {"ok": False, "error": f"La hoja '{nombre_hoja}' no existe en el Excel WK."}

        cols = list(WK_MXN_RANCH_COLS.values())
        for categoria, info in resumen.items():
            row = info["row"]
            values = [[info["totales"].get(rn, 0.0) for rn in WK_MXN_RANCH_COLS]]
            patch_resp = requests.patch(
                f"{wb_url}/worksheets/{target_sheet}/range(address='E{row}:K{row}')",
                headers=hdrs,
                json={"values": values},
                timeout=30,
            )
            if patch_resp.status_code not in (200, 201):
                return {
                    "ok": False,
                    "error": f"No se pudo escribir {categoria} en {target_sheet}: {patch_resp.text[:300]}",
                }

            requests.patch(
                f"{wb_url}/worksheets/{target_sheet}/range(address='E{row}:K{row}')/format",
                headers=hdrs,
                json={"numberFormat": "#,##0"},
                timeout=20,
            )
            requests.patch(
                f"{wb_url}/worksheets/{target_sheet}/range(address='E{row}:K{row}')/format/font",
                headers=hdrs,
                json={"color": "#C00000", "bold": True, "name": "Arial"},
                timeout=20,
            )

    finally:
        requests.post(f"{wb_url}/closeSession", headers=hdrs, timeout=20)

    fuentes_txt = ", ".join(fuentes_ok) if fuentes_ok else "ninguna"
    faltantes_txt = f" Faltantes: {', '.join(fuentes_faltantes)}." if fuentes_faltantes else ""
    omitidos_txt = ""
    if omitidos:
        partes = []
        for categoria, datos in omitidos.items():
            ranchos = ", ".join(f"{rn}={round(val, 2)}" for rn, val in datos.items())
            partes.append(f"{categoria}: {ranchos}")
        omitidos_txt = f" Ranchos omitidos sin columna WK: {' | '.join(partes)}."

    return {
        "ok": True,
        "mensaje": f"Materiales MN autorrellenados en '{nombre_hoja}'. Fuentes: {fuentes_txt}.{faltantes_txt}{omitidos_txt}",
    }


def autorrellenar_material_vegetal_wk(week_code: str, tenant_id: str, client_id: str, client_secret: str) -> dict:
    """
    Autorrellena la fila de MATERIAL VEGETAL (fila 14, columnas E:K) en la hoja WK####
    usando los datos de la hoja MV#### del Excel secundario de SharePoint.

    Usa mv_mode=True en el parser para detectar correctamente los ranchos
    con nombres completos (Propagacion, Cristina, Cecilia25, etc.).
    """
    import base64 as _b64

    code = _normalizar_week_code(week_code)
    if not (code.isdigit() and len(code) == 4):
        return {"ok": False, "error": "El código de semana debe ser exactamente 4 dígitos (ej: 2614)."}

    nombre_hoja = f"WK{code}"
    mv_sheet_name_expected = f"MV{code}"

    #  1. Descargar Excel secundario (PR/MP/ME/MV) 
    archivo = _descargar_con_graph(SHAREPOINT_URL_PR, "Excel MV")
    if archivo is None:
        return {"ok": False, "error": "No se pudo descargar el Excel de MV desde SharePoint."}

    try:
        xls = pd.ExcelFile(archivo)
    except Exception as e:
        return {"ok": False, "error": f"No se pudo abrir el Excel de MV: {e}"}

    sheet_name = _buscar_hoja_por_prefijo(xls.sheet_names, "MV", code)
    if not sheet_name:
        return {
            "ok": False,
            "error": f"No se encontró la hoja '{mv_sheet_name_expected}' en el Excel fuente. "
                     f"Súbela primero desde el panel 'Subir PR / MP / ME / MV'.",
        }

    #  2. Parsear MV con mv_mode=True 
    vals = _leer_hoja(xls, sheet_name, rango_filas=500, rango_cols=11)
    parsed = _parse_generic(vals, mv_mode=True)
    totales, omitidos = _sumar_gasto_por_rancho(parsed, tipo=None)

    mv_row = WK_MATERIAL_AUTOFILL["MATERIAL VEGETAL"]["row"]  #  14

    #  3. Autenticar contra Microsoft Graph 
    token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    token_resp = requests.post(token_url, data={
        "grant_type": "client_credentials",
        "client_id": client_id,
        "client_secret": client_secret,
        "scope": "https://graph.microsoft.com/.default",
    }, timeout=20)
    if token_resp.status_code != 200:
        return {"ok": False, "error": f"Error obteniendo token: {token_resp.text[:300]}"}

    token = token_resp.json().get("access_token")
    hdrs_json = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    #  4. Resolver el Excel WK principal 
    encoded = _b64.b64encode(SHAREPOINT_URL_WK.encode()).decode().rstrip("=")
    encoded = "u!" + encoded.replace("/", "_").replace("+", "-")
    res = requests.get(
        f"https://graph.microsoft.com/v1.0/shares/{encoded}/driveItem",
        headers=hdrs_json,
        timeout=20,
    )
    if res.status_code != 200:
        return {"ok": False, "error": f"No se pudo resolver el archivo WK: {res.text[:300]}"}

    item = res.json()
    drive_id = item.get("parentReference", {}).get("driveId")
    item_id = item.get("id")
    if not drive_id or not item_id:
        return {"ok": False, "error": "No se pudo obtener driveId o itemId del Excel WK."}

    wb_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/workbook"
    sess_resp = requests.post(
        f"{wb_url}/createSession",
        headers=hdrs_json,
        json={"persistChanges": True},
        timeout=30,
    )
    if sess_resp.status_code not in (200, 201):
        return {"ok": False, "error": f"Error abriendo sesión del workbook WK: {sess_resp.text[:300]}"}

    session_id = sess_resp.json().get("id")
    hdrs = {**hdrs_json, "workbook-session-id": session_id}

    try:
        #  5. Verificar que exista la hoja WK#### 
        sheets_resp = requests.get(f"{wb_url}/worksheets", headers=hdrs, timeout=20)
        if sheets_resp.status_code != 200:
            return {"ok": False, "error": f"Error listando hojas WK: {sheets_resp.text[:300]}"}

        target_sheet = None
        normalized_target = nombre_hoja.replace(" ", "").upper()
        for ws in sheets_resp.json().get("value", []):
            ws_name = str(ws.get("name", "")).strip()
            if ws_name.replace(" ", "").upper() == normalized_target:
                target_sheet = ws_name
                break

        if not target_sheet:
            return {"ok": False, "error": f"La hoja '{nombre_hoja}' no existe en el Excel WK."}

        #  6. Escribir totales MV en la fila 14 (columnas E:K) 
        values = [[totales.get(rn, 0.0) for rn in WK_MXN_RANCH_COLS]]
        patch_resp = requests.patch(
            f"{wb_url}/worksheets/{target_sheet}/range(address='E{mv_row}:K{mv_row}')",
            headers=hdrs,
            json={"values": values},
            timeout=30,
        )
        if patch_resp.status_code not in (200, 201):
            return {
                "ok": False,
                "error": f"No se pudo escribir MATERIAL VEGETAL en {target_sheet}: {patch_resp.text[:300]}",
            }

        # Formato numérico y estilo rojo-negrita igual que el resto de materiales
        requests.patch(
            f"{wb_url}/worksheets/{target_sheet}/range(address='E{mv_row}:K{mv_row}')/format",
            headers=hdrs,
            json={"numberFormat": "#,##0"},
            timeout=20,
        )
        requests.patch(
            f"{wb_url}/worksheets/{target_sheet}/range(address='E{mv_row}:K{mv_row}')/format/font",
            headers=hdrs,
            json={"color": "#C00000", "bold": True, "name": "Arial"},
            timeout=20,
        )

    finally:
        requests.post(f"{wb_url}/closeSession", headers=hdrs, timeout=20)

    omitidos_txt = ""
    if omitidos:
        partes = [f"{rn}={round(val, 2)}" for rn, val in omitidos.items()]
        omitidos_txt = f" Ranchos omitidos sin columna WK: {', '.join(partes)}."

    return {
        "ok": True,
        "mensaje": (
            f"[OK] MATERIAL VEGETAL autorrellenado en '{nombre_hoja}' (fila {mv_row}) "
            f"usando '{sheet_name}'.{omitidos_txt}"
        ),
    }


def autorrellenar_nomina_wk(week_code: str, tenant_id: str, client_id: str, client_secret: str) -> dict:
    """
    Autorrellena filas base de nómina en una hoja WK#### usando la hoja BD
    del Excel de nómina en SharePoint. Lee exclusivamente la columna MN ####.
    """
    import base64 as _b64

    code = _normalizar_week_code(week_code)
    if not (code.isdigit() and len(code) == 4):
        return {"ok": False, "error": "El código de semana debe ser exactamente 4 dígitos (ej: 2615)."}

    archivo = _descargar_con_graph(SHAREPOINT_URL_NOMINA, "Excel Nómina")
    if archivo is None:
        return {"ok": False, "error": "No se pudo descargar el Excel de nómina desde SharePoint."}

    try:
        df = pd.read_excel(archivo, sheet_name="BD", header=5)
    except Exception as e:
        return {"ok": False, "error": f"No se pudo abrir la hoja BD del Excel de nómina: {e}"}

    df.columns = [str(c).strip() for c in df.columns]
    monto_col = f"MN {code}"
    required = {"FINCA", "ESTATUS", "DEPARTAMENTO", monto_col}
    missing = required - set(df.columns)
    if missing:
        return {"ok": False, "error": f"Faltan columnas en BD: {', '.join(sorted(missing))}."}

    df = df[["FINCA", "ESTATUS", "DEPARTAMENTO", monto_col]].copy()
    df["FINCA"] = df["FINCA"].fillna("").astype(str).str.strip().str.upper()
    df["ESTATUS"] = df["ESTATUS"].fillna("").astype(str).str.strip().str.upper()
    df["DEPARTAMENTO"] = df["DEPARTAMENTO"].fillna("").astype(str).str.strip().str.upper()
    df[monto_col] = pd.to_numeric(df[monto_col], errors="coerce").fillna(0.0)

    df = df[
        (df["FINCA"] != "")
        & (df["DEPARTAMENTO"] != "")
        & (df[monto_col] != 0)
    ].copy()

    resumen = {row: {rn: 0.0 for rn in WK_MXN_RANCH_COLS} for row in NOMINA_WK_ROWS}
    omitidos = []
    usados = 0

    for _, row in df.iterrows():
        departamento = row["DEPARTAMENTO"]
        wk_row = _nomina_wk_row_from_departamento(departamento)
        if wk_row is None:
            continue

        rancho = _nomina_wk_ranch_from_bd(row["FINCA"], departamento)
        if not rancho or rancho not in WK_MXN_RANCH_COLS:
            omitidos.append(f"{departamento} [{row['FINCA']}]={round(float(row[monto_col]), 2)}")
            continue

        resumen[wk_row][rancho] = round(resumen[wk_row][rancho] + float(row[monto_col]), 2)
        usados += 1

    rows_con_datos = [row for row, vals in resumen.items() if any(vals.values())]
    if not rows_con_datos:
        return {
            "ok": False,
            "error": f"No se encontraron filas compatibles para automatizar usando la columna '{monto_col}'.",
        }

    token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    token_resp = requests.post(token_url, data={
        "grant_type": "client_credentials",
        "client_id": client_id,
        "client_secret": client_secret,
        "scope": "https://graph.microsoft.com/.default",
    }, timeout=20)
    if token_resp.status_code != 200:
        return {"ok": False, "error": f"Error obteniendo token: {token_resp.text[:300]}"}

    token = token_resp.json().get("access_token")
    hdrs_json = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    encoded = _b64.b64encode(SHAREPOINT_URL_WK.encode()).decode().rstrip("=")
    encoded = "u!" + encoded.replace("/", "_").replace("+", "-")
    res = requests.get(
        f"https://graph.microsoft.com/v1.0/shares/{encoded}/driveItem",
        headers=hdrs_json,
        timeout=20,
    )
    if res.status_code != 200:
        return {"ok": False, "error": f"No se pudo resolver el archivo WK: {res.text[:300]}"}

    item = res.json()
    drive_id = item.get("parentReference", {}).get("driveId")
    item_id = item.get("id")
    if not drive_id or not item_id:
        return {"ok": False, "error": "No se pudo obtener driveId o itemId del Excel WK."}

    nombre_hoja = f"WK{code}"
    wb_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/workbook"
    sess_resp = requests.post(
        f"{wb_url}/createSession",
        headers=hdrs_json,
        json={"persistChanges": True},
        timeout=30,
    )
    if sess_resp.status_code not in (200, 201):
        return {"ok": False, "error": f"Error abriendo sesión del workbook WK: {sess_resp.text[:300]}"}

    session_id = sess_resp.json().get("id")
    hdrs = {**hdrs_json, "workbook-session-id": session_id}

    try:
        sheets_resp = requests.get(f"{wb_url}/worksheets", headers=hdrs, timeout=20)
        if sheets_resp.status_code != 200:
            return {"ok": False, "error": f"Error listando hojas WK: {sheets_resp.text[:300]}"}

        target_sheet = None
        normalized_target = nombre_hoja.replace(" ", "").upper()
        for ws in sheets_resp.json().get("value", []):
            ws_name = str(ws.get("name", "")).strip()
            if ws_name.replace(" ", "").upper() == normalized_target:
                target_sheet = ws_name
                break

        if not target_sheet:
            return {"ok": False, "error": f"La hoja '{nombre_hoja}' no existe en el Excel WK."}

        for wk_row in rows_con_datos:
            values = [[resumen[wk_row].get(rn, 0.0) for rn in WK_MXN_RANCH_COLS]]
            patch_resp = requests.patch(
                f"{wb_url}/worksheets/{target_sheet}/range(address='E{wk_row}:K{wk_row}')",
                headers=hdrs,
                json={"values": values},
                timeout=30,
            )
            if patch_resp.status_code not in (200, 201):
                return {
                    "ok": False,
                    "error": f"No se pudo escribir la fila {wk_row} en {target_sheet}: {patch_resp.text[:300]}",
                }

            requests.patch(
                f"{wb_url}/worksheets/{target_sheet}/range(address='E{wk_row}:K{wk_row}')/format",
                headers=hdrs,
                json={"numberFormat": "#,##0"},
                timeout=20,
            )
            requests.patch(
                f"{wb_url}/worksheets/{target_sheet}/range(address='E{wk_row}:K{wk_row}')/format/font",
                headers=hdrs,
                json={"color": "#C00000", "bold": True, "name": "Arial"},
                timeout=20,
            )
    finally:
        requests.post(f"{wb_url}/closeSession", headers=hdrs, timeout=20)

    filas_txt = ", ".join(f"{row}={NOMINA_WK_ROWS[row]}" for row in rows_con_datos)
    omitidos_txt = ""
    if omitidos:
        omitidos_txt = f" Omitidos sin cruce WK: {' | '.join(omitidos[:8])}"
        if len(omitidos) > 8:
            omitidos_txt += f" | +{len(omitidos) - 8} más"

    return {
        "ok": True,
        "mensaje": (
            f"Nómina MN autorrellenada en '{nombre_hoja}' usando '{monto_col}'. "
            f"Filas actualizadas: {filas_txt}. Registros sumados: {usados}.{omitidos_txt}"
        ),
    }


def autorrellenar_conteo_marlen(
    week_code: str,
    tt_file,
    tenant_id: str,
    client_id: str,
    client_secret: str,
    posco_file=None,
    vivero_file=None,
) -> dict:
    """
    Lee la hoja TEMP del TT Nómina (header fila 7, columna DEPARTAMENTO = '{ACTIVIDAD} {FINCA}')
    y escribe los conteos en la hoja "Conteo" del SharePoint de Marlen.

    Fincas activas: ISABELA, RAMONA, CECILIA, CECILIA 25, CHRISTINA.

    Prefijo TT  ÁREA del Conteo:
      CORTE            CORTE
      TRANSPLANTE      TRASPLANTE
      MANEJO           MANEJO PLANTA
      HOOPS            HOOPS
      MIPE + MIRFE     MIPE / MIRFE       (sumados)
      CAMERO + TRACTORISTA  TRACTORES/CAMEROS (sumados)
      VELADOR          VELADORES
      SOLDADOR         SOLDADORES
      CHOFER           TRANSPORTE
    """
    import base64 as _b64
    from io import BytesIO as _BytesIO
    from collections import defaultdict

    code = str(week_code).strip()
    if not (code.isdigit() and len(code) == 4):
        return {"ok": False, "error": "El código de semana debe ser 4 dígitos (ej: 2615)."}

    code_int = int(code)

    #  Normalización de área (consistente en build y lookup) 
    def _norm_area(s: str) -> str:
        """Normaliza nombre de área: upper, colapsa espacios+slashes a espacio."""
        return re.sub(r"[\s/]+", " ", str(s).upper()).strip()

    #  Mapa prefijo TT  ÁREA exacta en hoja Conteo 
    _PREFIX_TO_AREA = {
        "CORTE":        "CORTE",
        "TRANSPLANTE":  "TRASPLANTE",
        "MANEJO":       "MANEJO PLANTA",
        "HOOPS":        "HOOPS",
        "MIPE":         "MIPE / MIRFE",
        "MIRFE":        "MIPE / MIRFE",
        "CAMERO":       "TRACTORES/CAMEROS",
        "TRACTORISTA":  "TRACTORES/CAMEROS",
        "VELADOR":      "VELADORES",
        "SOLDADOR":     "SOLDADORES",
        "CHOFER":       "TRANSPORTE",
    }
    # Pre-normalizar los valores para que el lookup sea consistente
    _PREFIX_TO_AREA_NORM = {k: _norm_area(v) for k, v in _PREFIX_TO_AREA.items()}

    #  Config de fincas: (palabras_a_filtrar, palabras_a_excluir, strips, ubic_conteo) 
    # Orden importa: CECILIA 25 antes de CECILIA para que el strip sea correcto
    _FINCAS = [
        {
            "label":    "CECILIA 25",
            "include":  ["CECILIA 25"],
            "exclude":  [],
            "strips":   ["CECILIA 25"],
            "ubic":     "CECILIA 25",
        },
        {
            "label":    "CECILIA",
            "include":  ["CECILIA"],
            "exclude":  ["CECILIA 25"],
            "strips":   ["CECILIA"],
            "ubic":     "CECILIA",
        },
        {
            "label":    "RAMONA",
            "include":  ["RAMONA"],
            "exclude":  [],
            "strips":   ["RAMONA"],
            "ubic":     "RAMONA",
        },
        {
            "label":    "ISABELA",
            "include":  ["ISABELA", "ISABELLA"],
            "exclude":  [],
            "strips":   ["ISABELA", "ISABELLA"],
            "ubic":     "ISABELA",
        },
        {
            "label":    "CHRISTINA",
            "include":  ["CHRISTINA", "CRHISTINA"],   # cubre el typo del TT
            "exclude":  [],
            "strips":   ["CHRISTINA", "CRHISTINA"],
            "ubic":     "CHRISTINA",
        },
    ]

    #  1. Leer hoja TEMP del TT Nómina 
    try:
        if hasattr(tt_file, "read"):
            raw = tt_file.read()
            if hasattr(tt_file, "seek"):
                try:
                    tt_file.seek(0)
                except Exception:
                    pass
        else:
            raw = tt_file.getvalue()

        df_tt = pd.read_excel(_BytesIO(raw), sheet_name="TEMP", header=7).fillna("")
    except Exception as e:
        return {"ok": False, "error": f"No se pudo leer la hoja TEMP del TT Nómina: {e}"}

    if "DEPARTAMENTO" not in df_tt.columns:
        return {
            "ok": False,
            "error": f"No se encontró columna DEPARTAMENTO. Columnas: {list(df_tt.columns)[:10]}",
        }

    dept_col = df_tt["DEPARTAMENTO"].astype(str).str.strip().str.upper()

    #  2. Contar por (finca, área) 
    # conteos[(ubic_conteo, area)] = int
    conteos: dict = defaultdict(int)
    sin_map: list = []

    for cfg in _FINCAS:
        # Máscara include
        mask = pd.Series([False] * len(df_tt), index=df_tt.index)
        for kw in cfg["include"]:
            mask = mask | dept_col.str.contains(kw, na=False)
        # Máscara exclude
        for kw in cfg["exclude"]:
            mask = mask & ~dept_col.str.contains(kw, na=False)

        df_finca = df_tt[mask]
        if df_finca.empty:
            print(f"   [i]  {cfg['label']}: sin empleados en este TT")
            continue

        print(f"   [OK] {cfg['label']}: {len(df_finca)} empleados")

        for dept in dept_col[mask]:
            # Extraer prefijo quitando el nombre de la finca
            prefix = dept
            for s in cfg["strips"]:
                prefix = prefix.replace(s, "").strip()
            prefix = prefix.strip()

            area_norm = _PREFIX_TO_AREA_NORM.get(prefix)
            if area_norm:
                conteos[(cfg["ubic"].upper(), area_norm)] += 1
            else:
                sin_map.append(f"{cfg['label']}:{dept}")

    #  ADMON standalone  ADMINISTRACION / ING. Y ADMON. 
    admon_mask = dept_col == "ADMON"
    admon_count = int(admon_mask.sum())
    if admon_count:
        conteos[("ADMINISTRACION", _norm_area("ING. Y ADMON."))] += admon_count
        print(f"   [OK] ADMINISTRACION (ADMON): {admon_count} empleados")

    if not conteos:
        return {"ok": False, "error": "No se encontraron empleados para ninguna finca activa en el TT."}

    print(f"\n[OK] Conteos TT  {sum(conteos.values())} empleados totales:")
    for (ubic, area), cnt in sorted(conteos.items()):
        print(f"   {ubic:12s} | {area:30s}  {cnt}")
    if sin_map:
        print(f"   [WARN]  Sin mapear: {sin_map}")

    #  3. Descargar hoja Conteo para localizar celdas 
    archivo_conteo = _descargar_con_graph(SHAREPOINT_URL_CONTEO_MARLEN, "Conteo Marlen")
    if archivo_conteo is None:
        return {"ok": False, "error": "No se pudo descargar el archivo Conteo de SharePoint."}

    try:
        df_c = pd.read_excel(archivo_conteo, sheet_name="Conteo", header=None).fillna("")
    except Exception as e:
        return {"ok": False, "error": f"No se pudo abrir la hoja \'Conteo\': {e}"}

    # Detectar fila de encabezados
    header_idx = None
    for i in range(min(10, len(df_c))):
        row_up = [str(v).strip().upper() for v in df_c.iloc[i].values]
        if "SEM" in row_up and ("UBICACIÓN" in row_up or "UBICACION" in row_up):
            header_idx = i
            break

    if header_idx is None:
        return {"ok": False, "error": "No se encontró la fila de encabezados en la hoja \'Conteo\'."}

    hdrs_c = [str(v).strip().upper() for v in df_c.iloc[header_idx].values]

    def _ci(keywords):
        for kw in keywords:
            if kw in hdrs_c:
                return hdrs_c.index(kw)
        return None

    ci_sem  = _ci(["SEM", "SEMANA"])
    ci_ubic = _ci(["UBICACIÓN", "UBICACION"])
    ci_area = _ci(["ÁREA / DEPARTAMENTO", "AREA / DEPARTAMENTO",
                   "ÁREA/DEPARTAMENTO",   "AREA/DEPARTAMENTO", "ÁREA", "AREA"])
    ci_cont = _ci(["CONTEO"])

    if any(c is None for c in [ci_sem, ci_ubic, ci_area, ci_cont]):
        return {
            "ok": False,
            "error": f"Columnas faltantes en Conteo: sem={ci_sem} ubic={ci_ubic} area={ci_area} conteo={ci_cont}",
        }

    def _col_letter(n):
        s = ""
        while n > 0:
            n, r = divmod(n - 1, 26)
            s = chr(65 + r) + s
        return s

    cont_col_letter = _col_letter(ci_cont + 1)

    # Recorrer filas Conteo buscando semana + fincas activas
    ubics_activas = {cfg["ubic"].upper() for cfg in _FINCAS} | {"ADMINISTRACION"}
    filas_a_escribir = []   # [(excel_row, ubic, area, valor)]

    for i in range(header_idx + 1, len(df_c)):
        row = df_c.iloc[i].values

        sem_raw  = str(row[ci_sem]).strip()  if ci_sem  < len(row) else ""
        ubic_raw = str(row[ci_ubic]).strip().upper() if ci_ubic < len(row) else ""
        area_raw = str(row[ci_area]).strip().upper() if ci_area < len(row) else ""

        try:
            if int(float(sem_raw)) != code_int:
                continue
        except (ValueError, TypeError):
            continue

        if ubic_raw not in ubics_activas:
            continue

        # Normalizar área igual que como se construyeron las claves de conteos
        area_norm = _norm_area(area_raw)
        valor     = conteos.get((ubic_raw, area_norm), 0)
        excel_row = i + 1
        filas_a_escribir.append((excel_row, ubic_raw, area_norm, valor))

    if not filas_a_escribir:
        return {
            "ok": False,
            "error": (
                f"No se encontraron filas en la hoja \'Conteo\' para la semana {code_int} "
                f"con las fincas activas. Verifica que la columna Sem tenga el valor {code_int}."
            ),
        }

    print(f"\n[OK] Filas a escribir en Conteo: {len(filas_a_escribir)}")

    #  4. Autenticar con Microsoft Graph y escribir 
    token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    token_resp = requests.post(token_url, data={
        "grant_type":    "client_credentials",
        "client_id":     client_id,
        "client_secret": client_secret,
        "scope":         "https://graph.microsoft.com/.default",
    }, timeout=20)
    if token_resp.status_code != 200:
        return {"ok": False, "error": f"Error obteniendo token: {token_resp.text[:300]}"}

    token     = token_resp.json().get("access_token")
    hdrs_json = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    encoded = _b64.b64encode(SHAREPOINT_URL_CONTEO_MARLEN.encode()).decode().rstrip("=")
    encoded = "u!" + encoded.replace("/", "_").replace("+", "-")
    res_sp  = requests.get(
        f"https://graph.microsoft.com/v1.0/shares/{encoded}/driveItem",
        headers=hdrs_json, timeout=20,
    )
    if res_sp.status_code != 200:
        return {"ok": False, "error": f"No se pudo resolver el archivo Conteo: {res_sp.text[:300]}"}

    item     = res_sp.json()
    drive_id = item.get("parentReference", {}).get("driveId")
    item_id  = item.get("id")
    if not drive_id or not item_id:
        return {"ok": False, "error": "No se pudo obtener driveId/itemId del archivo Conteo."}

    wb_url    = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/workbook"
    sess_resp = requests.post(
        f"{wb_url}/createSession",
        headers=hdrs_json, json={"persistChanges": True}, timeout=30,
    )
    if sess_resp.status_code not in (200, 201):
        return {"ok": False, "error": f"Error abriendo sesión: {sess_resp.text[:300]}"}

    session_id = sess_resp.json().get("id")
    hdrs       = {**hdrs_json, "workbook-session-id": session_id}

    escritas = 0
    errores  = []
    try:
        for excel_row, ubic, area, valor in filas_a_escribir:
            addr    = f"{cont_col_letter}{excel_row}"
            patch_r = requests.patch(
                f"{wb_url}/worksheets/Conteo/range(address=\'{addr}\')",
                headers=hdrs,
                json={"values": [[valor]]},
                timeout=20,
            )
            if patch_r.status_code in (200, 201):
                escritas += 1
                requests.patch(
                    f"{wb_url}/worksheets/Conteo/range(address=\'{addr}\')/format",
                    headers=hdrs, json={"numberFormat": "#,##0"}, timeout=10,
                )
            else:
                errores.append(f"Fila {excel_row} ({ubic}|{area}): {patch_r.text[:80]}")
    finally:
        requests.post(f"{wb_url}/closeSession", headers=hdrs, timeout=20)

    if errores:
        return {
            "ok": False,
            "error": f"Se escribieron {escritas}/{len(filas_a_escribir)} celdas. Errores: {'; '.join(errores[:3])}",
        }

    fincas_str = ", ".join(sorted({ubic for _, ubic, _, _ in filas_a_escribir}))
    return {
        "ok": True,
        "mensaje": (
            f"[OK] Conteo WK{code} actualizado: {escritas} celdas escritas "
            f"({fincas_str}). "
            f"{sum(conteos.values())} empleados totales."
            + (f" Sin mapear: {sin_map}" if sin_map else "")
        ),
    }



#  Descarga de una hoja WK#### como xlsx con formato completo 
def get_sheet_xlsx(week_code: str) -> bytes | None:
    """
    Descarga el Excel de SharePoint y extrae la hoja WK{week_code}
    como un archivo .xlsx independiente con formato completo.
    """
    archivo = _descargar_con_graph(SHAREPOINT_URL_WK, "Excel WK")
    if archivo is None:
        return None

    archivo_bytes = archivo.getvalue()
    sheet_name = f"WK{week_code}"

    try:
        wb = openpyxl.load_workbook(BytesIO(archivo_bytes))

        target = None
        for sname in wb.sheetnames:
            normalized = re.sub(r'\s+', '', sname.strip()).upper()
            if normalized == sheet_name.upper():
                target = sname
                break

        if target is None:
            return None

        src_ws = wb[target]
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = target

        for row in src_ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font          = copy(cell.font)
                    new_cell.border        = copy(cell.border)
                    new_cell.fill          = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection    = copy(cell.protection)
                    new_cell.alignment     = copy(cell.alignment)

        for merge in src_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merge))

        for col_letter, col_dim in src_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width  = col_dim.width
            new_ws.column_dimensions[col_letter].hidden = col_dim.hidden
        for row_num, row_dim in src_ws.row_dimensions.items():
            new_ws.row_dimensions[row_num].height = row_dim.height
            new_ws.row_dimensions[row_num].hidden = row_dim.hidden

        buf = BytesIO()
        new_wb.save(buf)
        buf.seek(0)
        return buf.read()

    except Exception as e:
        print(f"[WARN]  Error extrayendo hoja {sheet_name}: {e}")
        return None


#  Subir hojas PR / ME / MP al Excel secundario de SharePoint 
def insertar_hojas_pr_me_mp(
    semana_code: str,
    tenant_id: str,
    client_id: str,
    client_secret: str,
    pr_file=None,
    mp_file=None,
    me_file1=None,
    me_file2=None,
    mv_file=None,
) -> dict:
    """
    Inserta hojas PR####, MP####, ME#### y/o MV#### en el Excel secundario de SharePoint
    usando Microsoft Graph API (misma técnica que crear_hoja_wk).

    Args:
        semana_code : código de 4 dígitos, ej "2613"
        tenant_id   : Azure AD tenant ID
        client_id   : App registration client ID
        client_secret: App registration client secret
        pr_file     : BytesIO o file-like del Excel PR (opcional)
        mp_file     : BytesIO o file-like del Excel MP (opcional)
        me_file1    : BytesIO o file-like del primer Excel ME (opcional)
        me_file2    : BytesIO o file-like del segundo Excel ME (opcional)
        mv_file     : BytesIO o file-like del Excel MV - Material Vegetal (opcional)

    Returns:
        dict con claves "PR", "MP", "ME", "MV"  cada una con {"ok": bool, "msg": str}
    """
    import base64 as _b64
    import time

    code = semana_code.strip().upper()
    # Aceptar "WK2613" o "2613"
    if code.startswith("WK"):
        code = code[2:]

    resultado = {}

    #  Helper: leer Excel  matriz de valores 
    def _read_matrix(f, max_rows=700, max_cols=20):
        """Lee un archivo Excel y devuelve lista de listas de valores (str/num/None).
        Acepta: Streamlit UploadedFile, BytesIO, o cualquier file-like object.
        Soporta .xlsx (openpyxl) y .xls antiguo (xlrd).
        """
        from io import BytesIO as _BytesIO
        try:
            # Leer todos los bytes en memoria para garantizar compatibilidad
            if hasattr(f, "read"):
                raw = f.read()
                if hasattr(f, "seek"):
                    try:
                        f.seek(0)
                    except Exception:
                        pass
            elif isinstance(f, (bytes, bytearray)):
                raw = bytes(f)
            else:
                raw = f.getvalue()

            if not raw:
                raise RuntimeError("El archivo está vacío o no se pudo leer.")

            file_obj = _BytesIO(raw)

            # Intentar openpyxl (.xlsx)
            try:
                wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
                ws = wb.active
                rows = []
                for row in ws.iter_rows(max_row=max_rows, max_col=max_cols, values_only=True):
                    rows.append([v if v is not None else "" for v in row])
                wb.close()
                return rows
            except Exception as xlsx_err:
                # Si falla openpyxl intentar con xlrd (.xls legacy)
                file_obj.seek(0)
                try:
                    import xlrd
                    wb_xls = xlrd.open_workbook(file_contents=file_obj.read())
                    ws_xls = wb_xls.sheet_by_index(0)
                    rows = []
                    for ri in range(min(ws_xls.nrows, max_rows)):
                        row = []
                        for ci in range(min(ws_xls.ncols, max_cols)):
                            cell = ws_xls.cell(ri, ci)
                            row.append(cell.value if cell.value != "" else "")
                        rows.append(row)
                    return rows
                except ImportError:
                    raise RuntimeError(
                        f"No se pudo leer el archivo. "
                        f"Si es un archivo .xls antiguo, instala xlrd (pip install xlrd). "
                        f"Error original: {xlsx_err}"
                    )
                except Exception as xls_err:
                    raise RuntimeError(
                        f"No se pudo leer como .xlsx ({xlsx_err}) ni como .xls ({xls_err}). "
                        f"Asegúrate de que el archivo no esté corrupto o protegido con contraseña."
                    )
        except RuntimeError:
            raise
        except Exception as e:
            raise RuntimeError(f"Error leyendo Excel: {e}")



    #  Helper: obtener token OAuth2 
    def _get_token():
        token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
        r = requests.post(token_url, data={
            "grant_type":    "client_credentials",
            "client_id":     client_id,
            "client_secret": client_secret,
            "scope":         "https://graph.microsoft.com/.default",
        }, timeout=20)
        if r.status_code != 200:
            raise RuntimeError(f"Error obteniendo token: {r.text[:300]}")
        return r.json()["access_token"]

    #  Helper: resolver driveId + itemId desde URL de SharePoint 
    def _resolver_item(token, url):
        encoded = _b64.b64encode(url.encode()).decode().rstrip("=")
        encoded = "u!" + encoded.replace("/", "_").replace("+", "-")
        hdrs = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
        r = requests.get(
            f"https://graph.microsoft.com/v1.0/shares/{encoded}/driveItem",
            headers=hdrs, timeout=20,
        )
        if r.status_code != 200:
            raise RuntimeError(f"No se pudo resolver el archivo PR/ME/MP: {r.text[:300]}")
        item     = r.json()
        drive_id = item.get("parentReference", {}).get("driveId")
        item_id  = item.get("id")
        if not drive_id or not item_id:
            raise RuntimeError("No se pudo obtener driveId o itemId del Excel secundario.")
        return drive_id, item_id

    #  Helper: abrir sesión de workbook 
    def _abrir_sesion(wb_url, hdrs_json):
        r = requests.post(
            f"{wb_url}/createSession",
            headers=hdrs_json,
            json={"persistChanges": True},
            timeout=30,
        )
        if r.status_code not in (200, 201):
            raise RuntimeError(f"Error abriendo sesión: {r.text[:300]}")
        return r.json()["id"]

    #  Helper: crear e insertar una hoja 
    def _crear_hoja(wb_url, hdrs, nombre_hoja, matrix):
        """
        1. Verifica que la hoja no exista.
        2. Crea la hoja nueva.
        3. Escribe la matriz de valores via PATCH range.
        """
        # Verificar existencia
        sheets_r = requests.get(f"{wb_url}/worksheets", headers=hdrs, timeout=20)
        if sheets_r.status_code != 200:
            raise RuntimeError(f"Error listando hojas: {sheets_r.text[:300]}")
        existentes = [h["name"].strip().upper() for h in sheets_r.json().get("value", [])]
        if nombre_hoja.upper() in existentes:
            raise RuntimeError(f"La hoja '{nombre_hoja}' ya existe en SharePoint.")

        # Crear hoja
        add_r = requests.post(
            f"{wb_url}/worksheets/add",
            headers=hdrs,
            json={"name": nombre_hoja},
            timeout=20,
        )
        if add_r.status_code not in (200, 201):
            raise RuntimeError(f"Error creando hoja '{nombre_hoja}': {add_r.text[:300]}")

        time.sleep(0.5)  # pequeña pausa para que SharePoint registre la hoja

        # Preparar matriz (asegurar que todas las filas tienen el mismo número de cols)
        if not matrix:
            raise RuntimeError("Matriz vacía  el archivo parece estar vacío.")

        max_cols = max(len(r) for r in matrix)
        # Normalizar a max_cols columnas y serializar valores a strings seguros para JSON
        def _safe(v):
            if v is None or v == "":
                return ""
            if isinstance(v, float) and (v != v):  # NaN check
                return ""
            return v

        padded = [
            [_safe(v) for v in (row + [""] * (max_cols - len(row)))]
            for row in matrix
        ]
        nrows = len(padded)
        # Convertir número de columna a letra Excel (A, B, ..., Z, AA, ...)
        def _col_letter(n):  # n = 1-indexed
            s = ""
            while n > 0:
                n, r = divmod(n - 1, 26)
                s = chr(65 + r) + s
            return s

        end_col = _col_letter(max_cols)
        range_addr = f"A1:{end_col}{nrows}"

        patch_r = requests.patch(
            f"{wb_url}/worksheets/{nombre_hoja}/range(address='{range_addr}')",
            headers=hdrs,
            json={"values": padded},
            timeout=120,
        )
        if patch_r.status_code not in (200, 201):
            raise RuntimeError(
                f"Error escribiendo datos en '{nombre_hoja}': {patch_r.text[:400]}"
            )

        return nrows

    #  Obtener token y resolver Excel secundario UNA sola vez 
    # (reutilizamos para todas las hojas que se vayan a crear)
    token = None
    drive_id = None
    item_id  = None
    wb_url   = None
    session_id = None
    hdrs_json  = None
    hdrs       = None

    def _init_conexion():
        nonlocal token, drive_id, item_id, wb_url, session_id, hdrs_json, hdrs
        if wb_url is not None:
            return  # ya inicializado
        token     = _get_token()
        hdrs_json = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
        drive_id, item_id = _resolver_item(token, SHAREPOINT_URL_PR)
        wb_url    = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/workbook"
        session_id = _abrir_sesion(wb_url, hdrs_json)
        hdrs       = {**hdrs_json, "workbook-session-id": session_id}

    #  Helper: Limpiar archivos crudos de CONTPAQ 
    def _limpiar_matriz(matrix, mv_mode=False):
        """Toma la matriz con basura y extrae exactamente las 5 columnas.
        mv_mode=True: detecta producto desde col 6 primero, si vacía usa col 5.
        """
        import re
        cleaned = [["FECHA", "UBICACION", "PRODUCTO", "UNIDADES", "GASTO"]]
        for row in matrix:
            # Asegurar que la fila tenga al menos 12 elementos rellenando con vacíos
            r = list(row) + [""] * max(0, 12 - len(row))
            
            ubicacion = str(r[2]).strip().upper()
            ubicacion_cln = re.sub(r'\s+', '', ubicacion)
            
            if not ubicacion_cln or len(ubicacion_cln) < 5:
                continue
            if not re.match(r'^[A-Z0-9]+$', ubicacion_cln):
                continue
                
            fecha = r[0]
            
            # MV: col 6 tiene prioridad; si vacía usa col 5.
            # Resto: col 5 tiene prioridad; si vacía usa col 6.
            prod_c = str(r[5]).strip()
            prod_n = str(r[6]).strip()
            if mv_mode:
                prod = prod_n if prod_n else prod_c
            else:
                prod = prod_c if prod_c else prod_n
            
            # Unidades y Gasto
            unid = str(r[7]).strip()
            gasto = str(r[9]).strip()
            
            # Para evitar que filas 2 y 3 con "basura" pasen el filtro (metadatos de CONTPAQ), 
            # aseguramos que unidades o gasto contengan un número real.
            def is_num(v):
                if not v: return False
                try:
                    float(v.replace(',', '').replace('$', '').strip())
                    return True
                except ValueError: return False
                
            if not is_num(unid) and not is_num(gasto):
                continue
            
            cleaned.append([fecha, ubicacion_cln, prod, unid, gasto])
            
        print(f"   [Debug Clean] Matriz original tenía {len(matrix)} filas, la limpia tiene {len(cleaned)} filas")
        return cleaned if len(cleaned) > 1 else matrix

    #  Procesar PR 
    if pr_file is not None:
        nombre = f"PR{code}"
        try:
            _init_conexion()
            matrix = _read_matrix(pr_file)
            matrix = _limpiar_matriz(matrix)
            filas  = _crear_hoja(wb_url, hdrs, nombre, matrix)
            resultado["PR"] = {"ok": True, "msg": f"[OK] {nombre} creada ({filas} filas limpias)"}
            print(f"[OK] {nombre} insertada con {filas} filas.")
        except Exception as e:
            resultado["PR"] = {"ok": False, "msg": f"[ERR] PR  {e}"}
            print(f"[ERR] Error PR: {e}")
    else:
        resultado["PR"] = {"ok": None, "msg": " PR  no se subió archivo"}

    #  Procesar MP 
    if mp_file is not None:
        nombre = f"MP{code}"
        try:
            _init_conexion()
            matrix = _read_matrix(mp_file)
            matrix = _limpiar_matriz(matrix)
            filas  = _crear_hoja(wb_url, hdrs, nombre, matrix)
            resultado["MP"] = {"ok": True, "msg": f"[OK] {nombre} creada ({filas} filas limpias)"}
            print(f"[OK] {nombre} insertada con {filas} filas.")
        except Exception as e:
            resultado["MP"] = {"ok": False, "msg": f"[ERR] MP  {e}"}
            print(f"[ERR] Error MP: {e}")
    else:
        resultado["MP"] = {"ok": None, "msg": " MP  no se subió archivo"}

    #  Procesar ME (fusión de hasta 2 archivos) 
    if me_file1 is not None or me_file2 is not None:
        nombre = f"ME{code}"
        try:
            _init_conexion()
            matrix = []
            if me_file1 is not None:
                matrix += _read_matrix(me_file1)
            if me_file2 is not None:
                matrix += _read_matrix(me_file2)
                
            matrix_limpia = _limpiar_matriz(matrix)
            filas = _crear_hoja(wb_url, hdrs, nombre, matrix_limpia)
            
            resultado["ME"] = {
                "ok":  True,
                "msg": f"[OK] {nombre} creada ({filas} filas limpias totales)",
            }
            print(f"[OK] {nombre} insertada: {filas} filas.")
        except Exception as e:
            resultado["ME"] = {"ok": False, "msg": f"[ERR] ME  {e}"}
            print(f"[ERR] Error ME: {e}")
    else:
        resultado["ME"] = {"ok": None, "msg": " ME  no se subió archivo"}

    #  Procesar MV (Material Vegetal) 
    if mv_file is not None:
        nombre = f"MV{code}"
        try:
            _init_conexion()
            matrix = _read_matrix(mv_file)
            matrix = _limpiar_matriz(matrix, mv_mode=True)
            filas  = _crear_hoja(wb_url, hdrs, nombre, matrix)
            resultado["MV"] = {"ok": True, "msg": f"[OK] {nombre} creada ({filas} filas limpias)"}
            print(f"[OK] {nombre} insertada con {filas} filas.")
        except Exception as e:
            resultado["MV"] = {"ok": False, "msg": f"[ERR] MV  {e}"}
            print(f"[ERR] Error MV: {e}")
    else:
        resultado["MV"] = {"ok": None, "msg": " MV  no se subió archivo"}

    return resultado


def autorrellenar_siembra_wk(week_code: str, tenant_id: str, client_id: str, client_secret: str) -> dict:
    """
    Autorrellena las filas de siembra en una hoja WK####:
    - Fila 89: NUMERO DE CHAROLAS SEMBRADAS (desde Plantas-Metros)
    - Fila 91: METROS DE SIEMBRA (desde Mtrs Acumulados)
    """
    import base64 as _b64

    code = _normalizar_week_code(week_code)
    if not (code.isdigit() and len(code) == 4):
        return {"ok": False, "error": "El código de semana debe ser exactamente 4 dígitos (ej: 2614)."}

    nombre_hoja = f"WK{code}"
    code_int = int(code)

    metros_data = _extraer_metros_acumulados()
    plantas_data = _extraer_plantas_metros()

    totales_metros = {rn: 0.0 for rn in WK_MXN_RANCH_COLS}
    totales_charolas = {rn: 0.0 for rn in WK_MXN_RANCH_COLS}
    usados_m = 0
    usados_p = 0

    for item in metros_data:
        if item.get("semana_fin") == code_int:
            rn = item.get("rancho")
            if rn in totales_metros:
                totales_metros[rn] += float(item.get("metros", 0.0))
                usados_m += 1

    for item in plantas_data:
        if item.get("semana_fin") == code_int:
            rn = item.get("rancho")
            if rn in totales_charolas:
                totales_charolas[rn] += float(item.get("plantas", 0.0))
                usados_p += 1

    if usados_m == 0 and usados_p == 0:
        return {"ok": False, "error": f"No se encontraron datos de siembra para la semana {code_int}."}

    token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    token_resp = requests.post(token_url, data={
        "grant_type": "client_credentials",
        "client_id": client_id,
        "client_secret": client_secret,
        "scope": "https://graph.microsoft.com/.default",
    }, timeout=20)
    if token_resp.status_code != 200:
        return {"ok": False, "error": f"Error obteniendo token: {token_resp.text[:300]}"}

    token = token_resp.json().get("access_token")
    hdrs_json = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    encoded = _b64.b64encode(SHAREPOINT_URL_WK.encode()).decode().rstrip("=")
    encoded = "u!" + encoded.replace("/", "_").replace("+", "-")
    res = requests.get(
        f"https://graph.microsoft.com/v1.0/shares/{encoded}/driveItem",
        headers=hdrs_json,
        timeout=20,
    )
    if res.status_code != 200:
        return {"ok": False, "error": f"No se pudo resolver el archivo WK: {res.text[:300]}"}

    item = res.json()
    drive_id = item.get("parentReference", {}).get("driveId")
    item_id = item.get("id")
    if not drive_id or not item_id:
        return {"ok": False, "error": "No se pudo obtener driveId o itemId del Excel WK."}

    wb_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/workbook"
    sess_resp = requests.post(
        f"{wb_url}/createSession",
        headers=hdrs_json,
        json={"persistChanges": True},
        timeout=30,
    )
    if sess_resp.status_code not in (200, 201):
        return {"ok": False, "error": f"Error abriendo sesión del workbook WK: {sess_resp.text[:300]}"}

    session_id = sess_resp.json().get("id")
    hdrs = {**hdrs_json, "workbook-session-id": session_id}

    try:
        sheets_resp = requests.get(f"{wb_url}/worksheets", headers=hdrs, timeout=20)
        if sheets_resp.status_code != 200:
            return {"ok": False, "error": f"Error listando hojas WK: {sheets_resp.text[:300]}"}

        target_sheet = None
        normalized_target = nombre_hoja.replace(" ", "").upper()
        for ws in sheets_resp.json().get("value", []):
            ws_name = str(ws.get("name", "")).strip()
            if ws_name.replace(" ", "").upper() == normalized_target:
                target_sheet = ws_name
                break

        if not target_sheet:
            return {"ok": False, "error": f"La hoja '{nombre_hoja}' no existe en el Excel WK."}

        # Escribir Charolas (fila 89)
        values_charolas = [[totales_charolas.get(rn, 0.0) for rn in WK_MXN_RANCH_COLS]]
        patch_resp1 = requests.patch(
            f"{wb_url}/worksheets/{target_sheet}/range(address='E89:K89')",
            headers=hdrs,
            json={"values": values_charolas},
            timeout=30,
        )
        if patch_resp1.status_code in (200, 201):
            requests.patch(
                f"{wb_url}/worksheets/{target_sheet}/range(address='E89:K89')/format",
                headers=hdrs, json={"numberFormat": "#,##0"}, timeout=20,
            )
            requests.patch(
                f"{wb_url}/worksheets/{target_sheet}/range(address='E89:K89')/format/font",
                headers=hdrs, json={"color": "#000000", "bold": True, "name": "Arial"}, timeout=20,
            )

        # Escribir Metros (fila 91)
        values_metros = [[totales_metros.get(rn, 0.0) for rn in WK_MXN_RANCH_COLS]]
        patch_resp2 = requests.patch(
            f"{wb_url}/worksheets/{target_sheet}/range(address='E91:K91')",
            headers=hdrs,
            json={"values": values_metros},
            timeout=30,
        )
        if patch_resp2.status_code in (200, 201):
            requests.patch(
                f"{wb_url}/worksheets/{target_sheet}/range(address='E91:K91')/format",
                headers=hdrs, json={"numberFormat": "#,##0"}, timeout=20,
            )
            requests.patch(
                f"{wb_url}/worksheets/{target_sheet}/range(address='E91:K91')/format/font",
                headers=hdrs, json={"color": "#000000", "bold": True, "name": "Arial"}, timeout=20,
            )

    finally:
        requests.post(f"{wb_url}/closeSession", headers=hdrs, timeout=20)

    return {
        "ok": True,
        "mensaje": f"[OK] Siembra autorrellenada en '{nombre_hoja}'. Charolas (fil 89): {usados_p} cols sumadas. Metros (fil 91): {usados_m} cols sumadas.",
    }
